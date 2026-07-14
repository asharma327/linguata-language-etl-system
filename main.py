from fastapi import FastAPI, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import pymysql
import re
import traceback
import json
import io
import base64
import os
import logging
from pathlib import Path
from google.genai import types
from typing import Optional
from mangum import Mangum
from dotenv import load_dotenv

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("linguata")
import boto3
from PIL import Image as PILImage
from openai import OpenAI
import openai

load_dotenv()

app = FastAPI()


# --- Pydantic Models ---

class DBConfig(BaseModel):
    host: str
    user: str
    password: str
    database: str


class CloneSchemaRequest(BaseModel):
    source: DBConfig
    dest: DBConfig


class CefrRange(BaseModel):
    min: int
    max: int
    cefr_level: str


class InsertLessonsRequest(BaseModel):
    db: DBConfig
    folder: str | None = None
    files: list[str] | None = None
    cefr_level: str = "A1"                    # fallback if no mapping matches
    cefr_mapping: list[CefrRange] | None = None


class LessonImageTarget(BaseModel):
    lesson_id: int
    question_ids: list[int] | None = None   # if None + vocab lesson → all questions
    additional_prompt: str | None = None    # overrides global prompt for this lesson


class GrammarAudioTarget(BaseModel):
    lesson_id: int
    article_ids: list[int] | None = None  # if None → all articles for the lesson



class GenerateArticleQuestionsRequest(BaseModel):
    db: DBConfig
    s3_bucket: str
    s3_output_prefix: str                             # S3 prefix where question JSONs are written
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    openai_api_key: str | None = None
    num_questions: int = 10
    model: str = "gpt-4o-mini"
    temperature: float = 0.4
    force: bool = False                               # re-process even if output JSON already in S3
    limit: int | None = None                          # max number of lessons to process
    lesson_type: str | None = None                   # generate for all lessons of this type
    lesson_ids: list[int] | None = None              # generate for specific lesson IDs


class GenerateUnitImagesRequest(BaseModel):
    db: DBConfig
    s3_bucket: str
    s3_prefix: str                                    # e.g. "spanish/unit_images"
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    openai_api_key: str | None = None
    units: list[int] | None = None                   # specific unit numbers; None = auto-detect all
    force: bool = False                               # re-generate even if image already exists in S3
    image_model: str = "gpt-image-1"
    theme_model: str = "gpt-4o-mini"
    temperature: float = 0.4


class IngestUnitImagesRequest(BaseModel):
    db: DBConfig
    s3_bucket: str
    s3_prefix: str                                    # prefix under which unit .png files live
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    force: bool = False                               # re-insert even if UnitImages row already exists
    lesson_ids: list[int] | None = None              # if provided, link existing unit images to these lessons only


class BackfillAnswerTextRequest(BaseModel):
    db: DBConfig
    folder: str | None = None       # path to directory containing JSON files
    filename: str | None = None     # single JSON file path
    limit: int | None = None        # max number of files to process


class ReplaceSpeakingArticlesRequest(BaseModel):
    db: DBConfig
    folder: str
    limit: int | None = None


class GenerateWritingLessonsRequest(BaseModel):
    db: DBConfig
    units: list[int]
    num_questions: int = 5
    model: str = "gpt-4o-mini"
    temperature: float = 0.7
    openai_api_key: str | None = None
    cefr_mapping: list[CefrRange] | None = None   # unit-range → CEFR level; overrides DB value
    cefr_level: str = "A1"                         # fallback if no mapping and no DB value
    force: bool = False                             # re-generate even if writing lesson already exists


class ConvertToListeningRequest(BaseModel):
    db: DBConfig
    lesson_ids: list[int]
    s3_bucket: str
    s3_prefix: str                                    # e.g. "spanish/listening_audio"
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    openai_api_key: str | None = None                 # falls back to OPENAI_API_KEY env var


# --- Helpers ---

def connect_to_db(config: DBConfig, use_database: bool = True):
    """Open a PyMySQL connection. If use_database=False, connects without selecting a DB."""
    kwargs = dict(
        host=config.host,
        user=config.user,
        password=config.password,
        cursorclass=pymysql.cursors.DictCursor,
    )
    if use_database:
        kwargs["database"] = config.database
    return pymysql.connect(**kwargs)


def resolve_cefr(filename: str, mapping: list[CefrRange] | None, fallback: str) -> str:
    """Extract unit or lesson index from filename and return the matching CEFR level."""
    if mapping:
        match = re.search(r"^unit(\d+)", filename, re.IGNORECASE)
        if match:
            unit = int(match.group(1))
            for r in mapping:
                if r.min <= unit <= r.max:
                    return r.cefr_level
        # Italian FAST-style: "... Volume 1 - Lesson 01 ..." — map by lesson number
        m_lesson = re.search(r"Lesson\s+(\d+)", filename, re.IGNORECASE)
        if m_lesson:
            lesson_n = int(m_lesson.group(1))
            for r in mapping:
                if r.min <= lesson_n <= r.max:
                    return r.cefr_level
    return fallback


def make_create_if_not_exists(ddl: str) -> str:
    """Transform 'CREATE TABLE `name`' → 'CREATE TABLE IF NOT EXISTS `name`'."""
    return re.sub(
        r"CREATE TABLE\s+(`[^`]+`|\S+)",
        r"CREATE TABLE IF NOT EXISTS \1",
        ddl,
        count=1,
        flags=re.IGNORECASE,
    )


# --- Routes ---

@app.get("/")
def root():
    return {"status": "healthy"}

@app.get("/test")
def test():
    return {"status": "ok"}


@app.post("/clone-schema")
def clone_schema(body: CloneSchemaRequest):
    """
    Reads all table DDLs from the source MySQL database and recreates them
    in the destination database (created if it doesn't exist). No data is copied.
    """
    # 1. Connect to source and collect DDLs
    try:
        src_conn = connect_to_db(body.source, use_database=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to source database: {e}")

    table_ddls: dict[str, str] = {}
    try:
        with src_conn.cursor() as cur:
            cur.execute("SHOW TABLES")
            rows = cur.fetchall()
            # Result rows look like: {"Tables_in_<db>": "table_name"}
            table_names = [list(row.values())[0] for row in rows]

            for table in table_names:
                cur.execute(f"SHOW CREATE TABLE `{table}`")
                result = cur.fetchone()
                # Key is either "Create Table" or "Create View"
                ddl = result.get("Create Table") or result.get("Create View", "")
                table_ddls[table] = make_create_if_not_exists(ddl)
    except Exception as e:
        src_conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to read source schema: {e}")
    finally:
        src_conn.close()

    # 2. Connect to destination host (no database selected yet)
    try:
        dest_conn = connect_to_db(body.dest, use_database=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to destination host: {e}")

    cloned_tables = []
    errors = []
    dest_db = body.dest.database

    try:
        with dest_conn.cursor() as cur:
            # 3. Create destination database if it doesn't exist
            cur.execute(f"CREATE DATABASE IF NOT EXISTS `{dest_db}`")
            cur.execute(f"USE `{dest_db}`")

            # 4. Disable FK checks so tables can be created regardless of reference order
            cur.execute("SET FOREIGN_KEY_CHECKS=0")

            # 5. Create each table
            for table, ddl in table_ddls.items():
                try:
                    cur.execute(ddl)
                    cloned_tables.append(table)
                except Exception as table_err:
                    errors.append({"table": table, "error": str(table_err)})

            cur.execute("SET FOREIGN_KEY_CHECKS=1")

        dest_conn.commit()
    except Exception as e:
        dest_conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to create destination schema: {e}\n{traceback.format_exc()}")
    finally:
        dest_conn.close()

    response = {
        "status": "ok" if not errors else "partial",
        "created_database": dest_db,
        "tables_cloned": cloned_tables,
        "tables_total": len(table_ddls),
    }
    if errors:
        response["errors"] = errors

    return JSONResponse(content=response, status_code=200 if not errors else 207)


@app.post("/insert-lessons")
def insert_lessons(body: InsertLessonsRequest):
    """
    Reads all JSON files from the given folder and inserts each into
    Lesson, Article, Question, and Answer tables. Each file is its own
    transaction — a failure in one file does not affect others.
    """
    if not body.folder and not body.files:
        raise HTTPException(status_code=400, detail="Provide at least one of 'folder' or 'files'")

    json_files: list[Path] = []

    if body.folder:
        folder = Path(body.folder)
        if not folder.is_dir():
            raise HTTPException(status_code=400, detail=f"Folder not found: {body.folder}")
        json_files.extend(sorted(folder.rglob("*.json")))

    if body.files:
        for f in body.files:
            p = Path(f)
            if not p.is_file():
                raise HTTPException(status_code=400, detail=f"File not found: {f}")
            if p not in json_files:
                json_files.append(p)

    if not json_files:
        raise HTTPException(status_code=400, detail="No JSON files found")

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    files_processed = []
    files_failed = []
    total_lessons = total_articles = total_questions = total_answers = 0

    for json_file in json_files:
        file_log = {
            "file": json_file.name,
            "status": None,
            "lesson_id": None,
            "articles_inserted": 0,
            "questions_inserted": 0,
            "answers_inserted": 0,
            "errors": [],
        }

        # --- Parse JSON ---
        try:
            raw_data = json.loads(json_file.read_text(encoding="utf-8-sig"))
        except Exception as e:
            file_log["status"] = "failed"
            file_log["errors"].append(f"JSON parse error: {e}")
            files_failed.append(file_log)
            continue

        # Support both a single lesson object {...} and a list of lessons [...]
        lesson_entries = raw_data if isinstance(raw_data, list) else [raw_data]

        # --- Insert into DB (one transaction per file) ---
        try:
            with conn.cursor() as cur:
              for data in lesson_entries:

                # 1. Lesson
                has_question = 1 if data.get("questions_and_answers") else 0
                cefr = resolve_cefr(json_file.name, body.cefr_mapping, body.cefr_level)
                file_log["cefr_level"] = cefr
                cur.execute(
                    "INSERT INTO Lesson (title, type, cefr_level, has_question, metadata) "
                    "VALUES (%s, %s, %s, %s, %s)",
                    (
                        data["title"],
                        data["type"],
                        cefr,
                        has_question,
                        json.dumps(data.get("metadata")) if data.get("metadata") else None,
                    ),
                )
                lesson_id = cur.lastrowid
                file_log["lesson_id"] = lesson_id

                # 2. Articles
                for article in data.get("articles", []):
                    cur.execute(
                        "INSERT INTO Article (lesson_id, sequence_id, content) VALUES (%s, %s, %s)",
                        (lesson_id, article["sequence_id"], article["text"]),
                    )
                    file_log["articles_inserted"] += 1

                # 3. Questions + Answers
                for qa in data.get("questions_and_answers", []):
                    question_text = qa.get("question_text") or qa.get("question", "")

                    inferred_type = qa.get("type") or (
                        "multiple_choice" if "answers" in qa else "short_answer"
                    )
                    cur.execute(
                        "INSERT INTO Question (lesson_id, sequence_id, question_text, type, has_answer) "
                        "VALUES (%s, %s, %s, %s, %s)",
                        (
                            lesson_id,
                            qa["sequence_id"],
                            question_text,
                            inferred_type,
                            int(qa.get("has_answer", False)),
                        ),
                    )
                    question_id = cur.lastrowid
                    file_log["questions_inserted"] += 1

                    # Answers — handle all formats:
                    # A) answers[].answer_text  (vocabulary short_answer)
                    # B) flat answer + is_correct on the question  (grammar short_answer)
                    # C) answers[].answer  (multiple_choice)
                    if "answers" in qa:
                        for ans in qa["answers"]:
                            answer_text = ans.get("answer_text") or ans.get("answer") or ans.get("text", "")
                            cur.execute(
                                "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                "VALUES (%s, %s, %s, %s)",
                                (lesson_id, question_id, answer_text, int(ans.get("is_correct", False))),
                            )
                            file_log["answers_inserted"] += 1
                    elif "answer" in qa:
                        raw = qa["answer"]
                        # Format D: answer is a list of objects e.g. [{"answer": "...", "is_correct": true}]
                        if isinstance(raw, list):
                            for ans in raw:
                                cur.execute(
                                    "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                    "VALUES (%s, %s, %s, %s)",
                                    (lesson_id, question_id, ans.get("answer", ""), int(ans.get("is_correct", True))),
                                )
                                file_log["answers_inserted"] += 1
                        elif isinstance(raw, dict):
                            # Format E: answer is a dict with "text" and "is_correct" keys
                            cur.execute(
                                "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                "VALUES (%s, %s, %s, %s)",
                                (lesson_id, question_id, raw.get("text") or raw.get("answer", ""), int(raw.get("is_correct", True))),
                            )
                            file_log["answers_inserted"] += 1
                        else:
                            # Format B: answer is a plain string
                            cur.execute(
                                "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                "VALUES (%s, %s, %s, %s)",
                                (lesson_id, question_id, raw, int(qa.get("is_correct", True))),
                            )
                            file_log["answers_inserted"] += 1

            conn.commit()
            file_log["status"] = "success"
            total_lessons += 1
            total_articles += file_log["articles_inserted"]
            total_questions += file_log["questions_inserted"]
            total_answers += file_log["answers_inserted"]
            files_processed.append(file_log)

        except Exception as e:
            conn.rollback()
            file_log["status"] = "failed"
            file_log["errors"].append(str(e))
            files_failed.append(file_log)

    conn.close()

    all_succeeded = len(files_failed) == 0
    response = {
        "status": "ok" if all_succeeded else "partial",
        "summary": {
            "files_found": len(json_files),
            "files_succeeded": len(files_processed),
            "files_failed": len(files_failed),
            "lessons_inserted": total_lessons,
            "articles_inserted": total_articles,
            "questions_inserted": total_questions,
            "answers_inserted": total_answers,
        },
        "files": files_processed + files_failed,
    }
    return JSONResponse(content=response, status_code=200 if all_succeeded else 207)


# --- Image generation helpers ---

def _resize_to_256_png_bytes(raw_bytes: bytes) -> bytes:
    with PILImage.open(io.BytesIO(raw_bytes)) as im:
        im = im.convert("RGBA")
        im = im.resize((256, 256), PILImage.LANCZOS)
        out = io.BytesIO()
        im.save(out, format="PNG", optimize=True)
        return out.getvalue()


def _safe_slug(s: str) -> str:
    return "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in (s or "")).strip("_") or "item"


def _upload_to_s3_public(data_bytes: bytes, key: str, bucket: str, region: str,
                          access_key: str | None, secret_key: str | None,
                          content_type: str = "image/png") -> str:
    kwargs = {"region_name": region}
    if access_key and secret_key:
        kwargs["aws_access_key_id"] = access_key
        kwargs["aws_secret_access_key"] = secret_key
    s3 = boto3.client("s3", **kwargs)
    s3.put_object(Bucket=bucket, Key=key, Body=data_bytes, ContentType=content_type)
    return f"https://{bucket}.s3.amazonaws.com/{key}"


def _insert_image_row(conn, lesson_id: int, question_id: int,
                      image_url: str, image_metadata_json: str, sequence_id=None):
    sql = """
        INSERT INTO Image (lesson_id, question_id, sequence_id, image_url, image_metadata)
        VALUES (%s, %s, %s, %s, CAST(%s AS JSON))
    """
    with conn.cursor() as cur:
        cur.execute(sql, (lesson_id, question_id, sequence_id, image_url, image_metadata_json))


def _insert_audio_row(conn, lesson_id: int, article_id: int,
                      audio_url: str, audio_metadata_json: str, sequence_id=None):
    sql = """
        INSERT INTO Audio (lesson_id, article_id, sequence_id, audio_url, audio_metadata)
        VALUES (%s, %s, %s, %s, CAST(%s AS JSON))
    """
    with conn.cursor() as cur:
        cur.execute(sql, (lesson_id, article_id, sequence_id, audio_url, audio_metadata_json))


# --- Listening-questions helpers ---

def _make_s3_client(region: str, access_key: str | None, secret_key: str | None):
    kwargs: dict = {"region_name": region}
    if access_key and secret_key:
        kwargs["aws_access_key_id"] = access_key
        kwargs["aws_secret_access_key"] = secret_key
    return boto3.client("s3", **kwargs)


def _list_s3_mp3_keys(s3_client, bucket: str, prefix: str) -> list[str]:
    """Return sorted list of all .mp3 object keys under prefix (handles pagination)."""
    keys = []
    paginator = s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix.rstrip("/") + "/"):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.lower().endswith(".mp3"):
                keys.append(key)
    return sorted(keys)


def _s3_key_exists(s3_client, bucket: str, key: str) -> bool:
    try:
        s3_client.head_object(Bucket=bucket, Key=key)
        return True
    except Exception:
        return False


def _transcribe_from_s3(s3_client, bucket: str, key: str,
                         openai_client: OpenAI, model: str) -> str:
    """Download an MP3 from S3 and return its Whisper transcription text."""
    obj = s3_client.get_object(Bucket=bucket, Key=key)
    audio_bytes = obj["Body"].read()
    filename = Path(key).name
    # Pass as (name, file_object) tuple — accepted by all recent OpenAI SDK versions
    audio_file = (filename, io.BytesIO(audio_bytes))
    transcript = openai_client.audio.transcriptions.create(model=model, file=audio_file)
    return transcript.text


def _extract_unit_number_from_filename(filename: str) -> int | None:
    """Return the unit number found in the filename, or None.

    Handles formats like:
        Spanish Basic Course - Volume 2 - Unit 28C.mp3              →  28
        FSI - German ... - Unit 01 1.1.mp3                        →  1
        FSI - Italian FAST - Volume 1 - Lesson 07.mp3             →  7  (lesson index)
        unit05_listening_a.mp3                                     →  5
    """
    # Italian FAST: lesson number must win over "Volume 1" (otherwise every file would be 1)
    m_lesson = re.search(r"Lesson\s+(\d+)", filename, re.IGNORECASE)
    if m_lesson:
        try:
            return int(m_lesson.group(1))
        except ValueError:
            pass

    m = re.search(r"Unit\s+(\d+)", filename, re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            pass
    m2 = re.search(r"(^|[^\d])(\d+)([^\d]|$)", filename)
    if m2:
        try:
            return int(m2.group(2))
        except ValueError:
            pass
    return None


def _build_listening_lesson_title(filename: str) -> str:
    """Derive a normalised lesson title from an audio filename.

    Examples
    --------
    FSI - Italian FAST - Volume 1 - Lesson 01.mp3     →  Volume 1 - Lesson 01
    FSI ... Unit 01 1.1.mp3                           →  unit1_listening_1
    FSI ... Unit 02 2.3.mp3                           →  unit2_listening_3
    Spanish Basic Course - Volume 2 - Unit 28C.mp3    →  unit28_listening_c
    Spanish Basic Course - Volume 1 - Unit 02A.mp3   →  unit2_listening_a
    some_other_file.mp3                              →  some_other_file  (fallback)
    """
    stem = Path(filename).stem

    # Italian FAST (S3): "Volume 1 - Lesson 01" — store human-readable segment in DB
    m_it = re.search(r"Volume\s+(\d+)\s+-\s+Lesson\s+(\d+)", stem, re.IGNORECASE)
    if m_it:
        vol = int(m_it.group(1))
        lesson_n = int(m_it.group(2))
        return f"Volume {vol} - Lesson {lesson_n:02d}"

    # German-style (FSI): "Unit 01 1.1" → unit1_listening_1 (suffix = segment after the dot)
    m_de = re.search(r"Unit\s+(\d+)\s+(\d+)\.(\d+)", stem, re.IGNORECASE)
    if m_de:
        unit_num = int(m_de.group(1))
        minor = int(m_de.group(3))
        return f"unit{unit_num}_listening_{minor}"

    # Spanish-style: letter immediately after unit digits, e.g. Unit 02A
    m_es = re.search(r"Unit\s+(\d+)([A-Za-z])\b", stem, re.IGNORECASE)
    if m_es:
        unit_num = int(m_es.group(1))
        return f"unit{unit_num}_listening_{m_es.group(2).lower()}"

    # Legacy: Unit NN optional single letter (may have nothing after Unit NN)
    m = re.search(r"Unit\s+(\d+)([A-Za-z]?)", stem, re.IGNORECASE)
    if m:
        unit_num = int(m.group(1))
        part = m.group(2).lower()
        part_suffix = f"_{part}" if part else ""
        return f"unit{unit_num}_listening{part_suffix}"
    return stem


def _build_listening_mcq_messages(transcript_text: str, unit_number: int | None,
                                   num_questions: int) -> list[dict]:
    system = (
        "You are a careful item-writer for language learning. "
        "Create high-quality multiple-choice questions based ONLY on the provided transcript. "
        "Do not invent facts not supported by the transcript."
    )
    unit_repr = "null" if unit_number is None else str(unit_number)
    unit_type = "integer" if unit_number is not None else "null"
    user = (
        f"TRANSCRIPT (for Unit {unit_repr})\n"
        f"--------------------------------\n"
        f"{transcript_text}\n"
        f"--------------------------------\n\n"
        f"TASK:\n"
        f"- Produce {num_questions} multiple choice questions in ENGLISH derived ONLY from the transcript.\n"
        f"- Each question must have exactly 4 plausible answer choices.\n"
        f"- Include a field \"correct_answer\" as an integer index 0, 1, 2, or 3 pointing to the correct choice.\n"
        f"- Do NOT include explanations, rationales, or extra fields.\n"
        f"- Set \"unit_number\" to {unit_repr} ({unit_type} value).\n\n"
        f"OUTPUT FORMAT:\n"
        f"Return ONLY a JSON array (no surrounding prose) where each element has exactly:\n"
        f"  - \"question\" : string\n"
        f"  - \"answer_choices\" : array of exactly 4 strings\n"
        f"  - \"correct_answer\" : integer (0, 1, 2, or 3)\n"
        f"  - \"unit_number\" : {unit_repr}\n\n"
        f"IMPORTANT: valid JSON only, no code fences, exactly {num_questions} items."
    )
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]


def _parse_mcq_json_response(content: str) -> list:
    """Parse MCQ JSON, stripping accidental code fences if present."""
    try:
        return json.loads(content)
    except json.JSONDecodeError:
        cleaned = content.strip()
        if cleaned.startswith("```"):
            cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned)
            cleaned = re.sub(r"\s*```$", "", cleaned)
        return json.loads(cleaned)


def _validate_mcq_list(data: list, num_questions: int) -> None:
    if not isinstance(data, list):
        raise ValueError("Model did not return a JSON array.")
    if len(data) != num_questions:
        raise ValueError(f"Expected {num_questions} items; got {len(data)}.")
    for i, item in enumerate(data):
        for key in ("question", "answer_choices", "correct_answer", "unit_number"):
            if key not in item:
                raise ValueError(f"Item {i} missing key: '{key}'.")
        choices = item["answer_choices"]
        if not isinstance(choices, list) or len(choices) != 4:
            raise ValueError(f"Item {i} 'answer_choices' must be a list of exactly 4 strings.")
        ca = item["correct_answer"]
        if not isinstance(ca, int) or ca < 0 or ca > 3:
            raise ValueError(f"Item {i} 'correct_answer' must be an integer 0–3.")


def _insert_listening_lesson(conn, title: str, lesson_type: str,
                              cefr_level: str, mcq_items: list,
                              audio_url: str) -> int:
    """Insert Lesson + Audio + Questions + Answers for one audio file's MCQs. Returns lesson_id."""
    with conn.cursor() as cur:
        cur.execute(
            "INSERT INTO Lesson (title, type, cefr_level, has_question) VALUES (%s, %s, %s, 1)",
            (title, lesson_type, cefr_level),
        )
        lesson_id = cur.lastrowid

        # Store the source audio link
        audio_metadata = json.dumps({"source": "s3"})
        cur.execute(
            "INSERT INTO Audio (lesson_id, sequence_id, audio_url, audio_metadata) "
            "VALUES (%s, %s, %s, CAST(%s AS JSON))",
            (lesson_id, 1, audio_url, audio_metadata),
        )

        for seq, item in enumerate(mcq_items, start=1):
            cur.execute(
                "INSERT INTO Question (lesson_id, sequence_id, question_text, type, has_answer) "
                "VALUES (%s, %s, %s, 'multiple_choice', 1)",
                (lesson_id, seq, item["question"]),
            )
            question_id = cur.lastrowid
            for idx, choice_text in enumerate(item["answer_choices"]):
                cur.execute(
                    "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                    "VALUES (%s, %s, %s, %s)",
                    (lesson_id, question_id, choice_text, 1 if idx == item["correct_answer"] else 0),
                )
    return lesson_id


# --- Article-questions helpers ---

def _build_article_mcq_messages(articles_content: str, lesson_title: str,
                                  num_questions: int) -> list[dict]:
    system = (
        "You are a careful item-writer for language learning. "
        "Create high-quality multiple-choice questions based ONLY on the provided article content. "
        "Do not invent facts not supported by the articles."
    )
    user = (
        f"LESSON: {lesson_title}\n"
        f"ARTICLE CONTENT:\n"
        f"--------------------------------\n"
        f"{articles_content}\n"
        f"--------------------------------\n\n"
        f"TASK:\n"
        f"- Produce {num_questions} multiple choice questions in ENGLISH derived ONLY from the content above.\n"
        f"- Each question must have exactly 4 plausible answer choices.\n"
        f"- Include a field \"correct_answer\" as an integer index 0, 1, 2, or 3 pointing to the correct choice.\n"
        f"- Do NOT include explanations, rationales, or extra fields.\n\n"
        f"OUTPUT FORMAT:\n"
        f"Return ONLY a JSON array (no surrounding prose) where each element has exactly:\n"
        f"  - \"question\" : string\n"
        f"  - \"answer_choices\" : array of exactly 4 strings\n"
        f"  - \"correct_answer\" : integer (0, 1, 2, or 3)\n\n"
        f"IMPORTANT: valid JSON only, no code fences, exactly {num_questions} items."
    )
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": user},
    ]


def _validate_article_mcq_list(data: list, num_questions: int) -> None:
    if not isinstance(data, list):
        raise ValueError("Model did not return a JSON array.")
    if len(data) != num_questions:
        raise ValueError(f"Expected {num_questions} items; got {len(data)}.")
    for i, item in enumerate(data):
        for key in ("question", "answer_choices", "correct_answer"):
            if key not in item:
                raise ValueError(f"Item {i} missing key: '{key}'.")
        choices = item["answer_choices"]
        if not isinstance(choices, list) or len(choices) != 4:
            raise ValueError(f"Item {i} 'answer_choices' must be a list of exactly 4 strings.")
        ca = item["correct_answer"]
        if not isinstance(ca, int) or ca < 0 or ca > 3:
            raise ValueError(f"Item {i} 'correct_answer' must be an integer 0–3.")


def _insert_article_lesson_questions(conn, lesson_id: int, mcq_items: list) -> None:
    """Insert Questions + Answers into an existing lesson and set has_question=1."""
    with conn.cursor() as cur:
        cur.execute("UPDATE Lesson SET has_question = 1 WHERE lesson_id = %s", (lesson_id,))
        for seq, item in enumerate(mcq_items, start=1):
            cur.execute(
                "INSERT INTO Question (lesson_id, sequence_id, question_text, type, has_answer) "
                "VALUES (%s, %s, %s, 'multiple_choice', 1)",
                (lesson_id, seq, item["question"]),
            )
            question_id = cur.lastrowid
            for idx, choice_text in enumerate(item["answer_choices"]):
                cur.execute(
                    "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                    "VALUES (%s, %s, %s, %s)",
                    (lesson_id, question_id, choice_text, 1 if idx == item["correct_answer"] else 0),
                )


# --- Unit-image helpers ---

def _extract_unit_num_from_title(title: str) -> int | None:
    """Return the unit number from a lesson title like 'unit14_grammar', or None."""
    m = re.match(r"^unit(\d+)", title, re.IGNORECASE)
    return int(m.group(1)) if m else None


def _s3_unit_image_exists(s3_client, bucket: str, prefix: str, unit_number: int) -> bool:
    """Return True if any object at prefix/ starts with unit{n}_."""
    search_prefix = f"{prefix.rstrip('/')}/unit{unit_number}_"
    paginator = s3_client.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix=search_prefix):
        if page.get("Contents"):
            return True
    return False


def _build_unit_theme_messages(content_summary: str, unit_number: int) -> list[dict]:
    system = "You are a curriculum analyst for a language learning app."
    user = (
        f"Below is content from Unit {unit_number} of a language course, "
        f"including lesson articles and vocabulary Q&A.\n\n"
        f"CONTENT:\n"
        f"--------------------------------\n"
        f"{content_summary}\n"
        f"--------------------------------\n\n"
        f"TASK: Identify 1 to 3 words that best capture the central THEME of this unit "
        f"(e.g. 'greetings', 'family home', 'food market'). "
        f"Output ONLY the theme words separated by spaces, nothing else. No punctuation."
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


# --- Vocab-audio helpers ---

# Vocabulary questions that have no Audio row yet (joined on question_id)
_VOCAB_NO_AUDIO_SQL = """
    SELECT
        q.question_id,
        q.lesson_id,
        q.sequence_id,
        q.question_text,
        l.title AS lesson_title,
        a.answer_text
    FROM Question q
    JOIN Lesson l ON q.lesson_id = l.lesson_id
    LEFT JOIN Answer a ON a.question_id = q.question_id AND a.is_correct = 1
    LEFT JOIN Audio au ON au.question_id = q.question_id
    WHERE l.type = 'vocabulary'
      AND au.question_id IS NULL
    ORDER BY q.lesson_id, q.sequence_id
"""


def _generate_tts_bytes(openai_client: OpenAI, text: str,
                         model: str, voice: str, instructions: str | None) -> bytes:
    """Call OpenAI TTS and return raw MP3 bytes (fully in-memory, no disk I/O)."""
    kwargs: dict = {"model": model, "voice": voice, "input": text}
    if instructions:
        kwargs["instructions"] = instructions
    return openai_client.audio.speech.create(**kwargs).content


def _translate_text(translate_client, text: str, source_lang: str, target_lang: str) -> str:
    """Translate text using AWS Translate and return the translated string."""
    response = translate_client.translate_text(
        Text=text,
        SourceLanguageCode=source_lang,
        TargetLanguageCode=target_lang,
    )
    return response["TranslatedText"]


def _insert_vocab_audio_row(conn, lesson_id: int, question_id: int,
                             sequence_id: int, audio_url: str,
                             tts_model: str, voice: str) -> None:
    """Insert one Audio row keyed by question_id (no article_id for vocab questions).

    Falls back to an insert without audio_metadata if the column does not yet exist
    in the target database (schema version mismatch). Use /migrate-audio-table to add
    the column and avoid the fallback path.
    """
    metadata = json.dumps({"source": "tts", "tts_model": tts_model, "voice": voice})
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO Audio (lesson_id, question_id, sequence_id, audio_url, audio_metadata) "
                "VALUES (%s, %s, %s, %s, CAST(%s AS JSON))",
                (lesson_id, question_id, sequence_id, audio_url, metadata),
            )
    except pymysql.err.OperationalError as e:
        if e.args[0] == 1054:  # Unknown column — schema is missing audio_metadata
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO Audio (lesson_id, question_id, sequence_id, audio_url) "
                    "VALUES (%s, %s, %s, %s)",
                    (lesson_id, question_id, sequence_id, audio_url),
                )
        else:
            raise

class GenerateImagesRequest(BaseModel):
    db: DBConfig
    units: list[int] | None = None                  # path 1: select vocab lessons by unit number
    lessons: list[LessonImageTarget] | None = None  # path 2: explicit lesson/question targets
    titles: list[str] | None = None
    translate: bool = True
    additional_prompt: str | None = None            # global; per-lesson value overrides this
    limit_questions: int | None = None                     # max questions to process (units path)
    translate: bool = False                         # True if question_text is NOT already in English
    s3_bucket: str
    s3_prefix: str = "spanish"
    aws_region: str = "us-east-1"
    gemini_api_key: str | None = None
    openai_api_key: str | None = None               # falls back to OPENAI_API_KEY env var
    aws_access_key_id: str | None = None            # falls back to AWS_ACCESS_KEY_ID env var
    aws_secret_access_key: str | None = None        # falls back to AWS_SECRET_ACCESS_KEY env var

# --- VocabToPictures class (unchanged) ---
from openai import OpenAI
from google import genai

class VocabToPictures:
    """Generate images for vocabulary words using OpenAI, return bytes (no local save)."""

    def __init__(
        self,
        openai_api_key: str | None = None,
        gemini_api_key: str | None = None,
        model: str = "models/gemini-3.1-flash-lite-image",
        size: str = "1024x1024",
    ):

        self.openai = OpenAI(
            api_key=openai_api_key or os.getenv("OPENAI_API_KEY")
        )

        self.gemini = genai.Client(
            api_key=gemini_api_key or os.getenv("GEMINI_API_KEY")
        )

        self.model = model
        self.size = size

    def generate_one(self, word: str, translate: bool = False,
                     additional_prompt: str | None = None) -> dict | None:
        try:
            if translate:
                resp = self.openai.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": (
                        "Here is a word or phrase from a language learning textbook's vocabulary section. "
                        "Respond with ONLY the English translation. No leading or trailing text.\n\n"
                        f"Vocab:\n{word}\n\nTranslation:"
                    )}],
                    temperature=0,
                )
                english = (resp.choices[0].message.content or "").strip()
                if not english:
                    raise RuntimeError("Translation returned empty")
            else:
                english = word

            gen_prompt = f"""
                    Create a vocabulary-learning image for a language-learning application.

                    Original vocabulary:
                    "{word}"

                    English meaning:
                    "{english}"

                    Primary objective:
                    Help a beginner language learner immediately understand and remember the vocabulary item.

                    Decision rule:

                    1. If the vocabulary can be clearly represented visually:

                       * Create a realistic image of the concept.
                       * Use a realistic photographic style.
                       * Show one clear primary subject.
                       * Make the meaning obvious without requiring text.
                       * Prefer visual learning over written text.
                       * Avoid unnecessary objects and background clutter.

                    2. If the vocabulary is very difficult to represent visually like one word vocabulary (e.g. "is", "because", "however", "to be"):

                       * Create a clean educational flashcard, classroom whiteboard, notebook page, teaching poster, or vocabulary card to display the word and its meaning.
                       * Show BOTH the original vocabulary and the English meaning.
                       * Display them on in the format:

                         [original vocabulary] → [English meaning]

                       Examples:

                       está → is
                       estar → to be
                       porque → because
                       cependant → however
                       しかし → however

                       * Make the text large, readable, and centered.
                       * The vocabulary card should be the primary focus of the image.
                       * Use a clean classroom or language-learning environment.
                       * Make the image look like a professional educational resource.

                    Try your best to avoid 2nd rule. Only use it when the word is truly unrepresentable visually. When in doubt, prefer rule 1.

                    General requirements:

                    * Use realistic photographic style.
                    * Use natural colors and lighting.
                    * Keep the composition simple and easy to understand.
                    * Avoid visual clutter.
                    * Avoid artistic abstraction.
                    * Avoid visual metaphors.
                    * Avoid logos.
                    * Avoid watermarks.
                    * Avoid unrelated objects.
                    * Ensure the image remains clear and recognizable when resized to 256x256.
                    * Optimize for vocabulary acquisition and learner comprehension rather than artistic appearance.

                    Generate the single most educationally effective image possible.
                    """
            if additional_prompt:
                gen_prompt += f"\n\n{additional_prompt}"

            # result = self.client.images.generate(model=self.model, prompt=gen_prompt, size=self.size)

            # if result.data and hasattr(result.data[0], "b64_json") and result.data[0].b64_json:
            #     return {
            #         "image_bytes": base64.b64decode(result.data[0].b64_json),
            #         "original_text": word,
            #         "translated_text": english,
            #         "model": self.model,
            #         "requested_size": self.size,
            #     }

            from google.genai import types

            response = self.gemini.models.generate_content(
                model=self.model,
                contents=gen_prompt,
                config=types.GenerateContentConfig(
                    response_modalities=["TEXT", "IMAGE"],
                    image_config=types.ImageConfig(
                        aspect_ratio="1:1",
                        # image_size="1K", 
                    )
                ),
            )

            image_bytes = None

            # Gemini can return multiple parts (text, images, etc.)
            # We look for the first part that contains image data.
            if response.candidates and response.candidates[0].content:
                for part in response.candidates[0].content.parts:
                    inline = getattr(part, "inline_data", None)
                    if inline and getattr(inline, "data", None):
                        raw = inline.data # google-genai usually returns bytes; guard for base64-string SDK variants

                        if isinstance(raw, str):
                            raw = base64.b64decode(raw)
                        image_bytes = raw
                        break
                
            if image_bytes is None:
                raise RuntimeError("Gemini returned no image.")

            return {
                "image_bytes": image_bytes,
                "original_text": word,
                "translated_text": english,
                "model": self.model,
                "requested_size": self.size,
            }

        # except openai.PermissionDeniedError:
            # print(f"OpenAI permission denied. Skipping: {word}")
        except Exception as e:
            print(f"generate_one failed for '{word}': {e}")

        return None


def get_image_settings(database_name: str):
    database_name = database_name.lower().strip()

    settings = {
        "hindi":    {"additional_prompt": "\nWhen appropriate, use objects, environments, clothing, and daily-life settings commonly seen in India.\nAvoid stereotypes.\n"},
        "japanese": {"additional_prompt": "\nWhen appropriate, use objects, environments, and daily-life settings commonly seen in Japan.\nAvoid stereotypes.\n"},
        "german":   {"additional_prompt": "\nWhen appropriate, use objects and environments commonly seen in Germany.\nAvoid stereotypes.\n"},
        "extractiondb":   {"additional_prompt": "\nWhen appropriate, use objects and environments commonly seen in France.\nAvoid stereotypes.\n"},
        "italian":  {"additional_prompt": "\nWhen appropriate, use objects and environments commonly seen in Italy.\nAvoid stereotypes.\n"},
        "spanish":  {"additional_prompt": "\nUse culturally neutral imagery unless regional context is important.\n"},
        "chinese":  {"additional_prompt": "\nWhen appropriate, use objects, environments, and daily-life settings commonly seen in China.\nAvoid stereotypes.\n"},
    }
    if database_name not in settings:
        raise ValueError(f"Unsupported database: {database_name}")
    return settings[database_name]


# NOTE: GenerateImagesRequest needs:
#     titles: list[str] | None = None        # one or many lessons (exact title match)
#     translate: bool = True                 # default ON: question_text -> English before image
# (units and lessons stay as they are. limit_questions is now a GLOBAL cap.)
# Requires (already in main.py from the streaming routes): _emit, StreamingResponse.

# The blank-aware image condition, reused by every path.
_IMG_MISSING = "(i.image_id IS NULL OR i.image_url IS NULL OR TRIM(i.image_url) = '')"


@app.post("/generate-lesson-images")
def generate_lesson_images(body: GenerateImagesRequest):
    """
    Streams NDJSON progress while generating vocab images and storing them in S3 + Image table.

    Three input modes (combinable):
      - titles:  exact lesson titles (one or many)
      - units:   all vocab lessons whose title starts with unit{N}
      - lessons: explicit [{lesson_id, question_ids?, additional_prompt?}]

    Only questions whose image is MISSING are processed (no Image row, or a row whose
    image_url is NULL/blank). A blank row is UPDATED in place; otherwise a row is INSERTed.

    limit_questions is a GLOBAL cap applied after the full work list is built and ordered
    (lesson_id, sequence_id) — so limit_questions=1 generates only the first question,
    regardless of which input mode you used.
    """

    def stream():
        if not body.titles and not body.units and not body.lessons:
            yield _emit("error", message="Provide at least one of 'titles', 'units', or 'lessons'")
            return

        try:
            conn = connect_to_db(body.db)
            database_name = body.db.database
            language_image_prompt = get_image_settings(database_name)["additional_prompt"]
        except Exception as e:
            yield _emit("error", message=f"Could not connect to database: {e}")
            return
        
        yield _emit("start", action="generate-lesson-images", database=database_name,
                    translate=body.translate,
                    titles=body.titles or None, units=body.units or None,
                    lessons=[t.lesson_id for t in body.lessons] if body.lessons else None)
        
        work_items: list[dict] = []
        seen: set = set()  # (lesson_id, question_id) dedup across modes

        def _add(row, title, add_prompt):
            key = (row["lesson_id"], row["question_id"])
            if key in seen:
                return
            seen.add(key)
            work_items.append({
                "lesson_id": row["lesson_id"],
                "question_id": row["question_id"],
                "sequence_id": row.get("sequence_id", 0),
                "question_text": row["question_text"],
                "lesson_title": title,
                "additional_prompt": add_prompt,
            })

        try:
            with conn.cursor() as cur:

                # Path 1: titles (exact match)
                if body.titles:
                    ph = ",".join(["%s"] * len(body.titles))
                    cur.execute(f"""
                        SELECT l.lesson_id, l.title, q.question_id, q.sequence_id, q.question_text
                        FROM Lesson l
                        JOIN Question q ON q.lesson_id = l.lesson_id
                        LEFT JOIN Image i ON i.lesson_id = l.lesson_id AND i.question_id = q.question_id
                        WHERE l.type = 'vocabulary' AND l.title IN ({ph}) AND {_IMG_MISSING}
                    """, tuple(body.titles))
                    for r in cur.fetchall():
                        _add(r, r["title"], body.additional_prompt)

                # Path 2: units (prefix match)
                if body.units:
                    like_clauses = " OR ".join(["l.title LIKE %s" for _ in body.units])
                    like_params = [f"unit{u}%" for u in body.units]
                    cur.execute(f"""
                        SELECT l.lesson_id, l.title, q.question_id, q.sequence_id, q.question_text
                        FROM Lesson l
                        JOIN Question q ON q.lesson_id = l.lesson_id
                        LEFT JOIN Image i ON i.lesson_id = l.lesson_id AND i.question_id = q.question_id
                        WHERE ({like_clauses}) AND l.type = 'vocabulary' AND {_IMG_MISSING}
                    """, tuple(like_params))
                    for r in cur.fetchall():
                        _add(r, r["title"], body.additional_prompt)

                # Path 3: explicit lessons
                if body.lessons:
                    for target in body.lessons:
                        cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE lesson_id = %s",
                                    (target.lesson_id,))
                        lesson_row = cur.fetchone()
                        if not lesson_row:
                            yield _emit("lesson_error", lesson_id=target.lesson_id, error="lesson not found")
                            continue

                        effective_prompt = target.additional_prompt or body.additional_prompt
                        is_vocab = lesson_row["type"] == "vocabulary"

                        if target.question_ids:
                            fmt = ",".join(["%s"] * len(target.question_ids))
                            cur.execute(f"""
                                SELECT q.lesson_id, q.question_id, q.sequence_id, q.question_text
                                FROM Question q
                                LEFT JOIN Image i ON i.lesson_id = q.lesson_id AND i.question_id = q.question_id
                                WHERE q.lesson_id = %s AND q.question_id IN ({fmt}) AND {_IMG_MISSING}
                            """, (target.lesson_id, *target.question_ids))
                        elif is_vocab:
                            cur.execute(f"""
                                SELECT q.lesson_id, q.question_id, q.sequence_id, q.question_text
                                FROM Question q
                                LEFT JOIN Image i ON i.lesson_id = q.lesson_id AND i.question_id = q.question_id
                                WHERE q.lesson_id = %s AND {_IMG_MISSING}
                            """, (target.lesson_id,))
                        else:
                            yield _emit("lesson_error", lesson_id=target.lesson_id,
                                        error=f"lesson type '{lesson_row['type']}' — specify question_ids for non-vocab")
                            continue

                        for r in cur.fetchall():
                            _add(r, lesson_row["title"], effective_prompt)

        except Exception as e:
            conn.close()
            yield _emit("error", message=f"Failed to build work list: {e}")
            return

        # order globally, then apply the global cap
        work_items.sort(key=lambda w: (w["lesson_id"], w["sequence_id"]))
        if body.limit_questions is not None:
            work_items = work_items[: int(body.limit_questions)]

        yield _emit("found", total=len(work_items))
        if not work_items:
            conn.close()
            yield _emit("summary", total_items=0, succeeded=0, failed=0)
            return

        # generator = VocabToPictures(api_key=body.openai_api_key, model="gpt-image-1", size="1024x1024")
        generator = VocabToPictures(
            openai_api_key=body.openai_api_key,
            gemini_api_key=body.gemini_api_key,
            model="models/gemini-3.1-flash-lite-image",
            size="1024x1024",
        )
        succeeded = failed = 0

        try:
            for idx, item in enumerate(work_items, start=1):
                yield _emit("processing", n=idx, total=len(work_items),
                            lesson_title=item["lesson_title"], question_id=item["question_id"],
                            sequence_id=item["sequence_id"], question_text=item["question_text"])
                try:
                    effective_prompt = (item["additional_prompt"] or "") + "\n\n" + language_image_prompt

                    gen = generator.generate_one(
                        word=item["question_text"],
                        translate=body.translate,
                        additional_prompt=effective_prompt,
                    )
                    if not gen or not gen.get("image_bytes"):
                        raise RuntimeError("No image returned from generator")

                    png256 = _resize_to_256_png_bytes(gen["image_bytes"])
                    key = (f"{body.s3_prefix}/"
                           f"{item['lesson_id']}_{item['question_id']}_"
                           f"{_safe_slug(item['lesson_title'])}.png")
                    public_url = _upload_to_s3_public(
                        png256, key, body.s3_bucket, body.aws_region,
                        body.aws_access_key_id, body.aws_secret_access_key,
                    )
                    meta = {
                        "original_text": gen["original_text"],
                        "translated_text": gen["translated_text"],
                        "model": gen["model"],
                        "requested_size": gen["requested_size"],
                        "final_size": "256x256",
                        "s3_bucket": body.s3_bucket,
                        "s3_key": key,
                        "public_url": public_url,
                        "lesson_title": item["lesson_title"],
                    }
                    meta_json = json.dumps(meta, ensure_ascii=False)

                    # UPDATE blank row in place, else INSERT
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT image_id FROM Image WHERE lesson_id=%s AND question_id=%s "
                            "ORDER BY image_id ASC LIMIT 1",
                            (item["lesson_id"], item["question_id"]),
                        )
                        existing = cur.fetchone()

                    action = "updated" if existing else "inserted"
                    if existing:
                        with conn.cursor() as cur:
                            try:
                                cur.execute(
                                    "UPDATE Image SET image_url=%s, image_metadata=CAST(%s AS JSON) "
                                    "WHERE image_id=%s",
                                    (public_url, meta_json, existing["image_id"]),
                                )
                            except pymysql.err.OperationalError as e:
                                if e.args[0] == 1054:  # older schema without image_metadata
                                    cur.execute("UPDATE Image SET image_url=%s WHERE image_id=%s",
                                                (public_url, existing["image_id"]))
                                else:
                                    raise
                    else:
                        _insert_image_row(
                            conn,
                            lesson_id=item["lesson_id"],
                            question_id=item["question_id"],
                            image_url=public_url,
                            image_metadata_json=meta_json,
                        )
                    conn.commit()

                    succeeded += 1
                    yield _emit("success", question_id=item["question_id"],
                                sequence_id=item["sequence_id"], lesson_title=item["lesson_title"],
                                image_url=public_url, row=action)

                except Exception as e:
                    conn.rollback()
                    failed += 1
                    yield _emit("failed", question_id=item["question_id"],
                                sequence_id=item["sequence_id"], lesson_title=item["lesson_title"],
                                error=str(e))

            yield _emit("summary", total_items=len(work_items), succeeded=succeeded, failed=failed)
            logger.info("generate-lesson-images done | items=%d ok=%d fail=%d",
                        len(work_items), succeeded, failed)

        except Exception as e:
            conn.rollback()
            logger.error("generate-lesson-images crashed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e))
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


class GenerateGrammarAudioRequest(BaseModel):
    db: DBConfig
    units: list[int] | None = None                    # path 1: select grammar lessons by unit number
    lessons: list[GrammarAudioTarget] | None = None   # path 2: explicit lesson/article targets
    titles: list[str] | None = None
    limit_articles: int | None = None     # change from `int = 200` so None means "no cap"  # max articles to process (units path)
    s3_bucket: str
    s3_prefix: str = "spanish"
    aws_region: str = "us-east-1"
    openai_api_key: str | None = None                 # falls back to OPENAI_API_KEY env var
    aws_access_key_id: str | None = None              # falls back to AWS_ACCESS_KEY_ID env var
    aws_secret_access_key: str | None = None          # falls back to AWS_SECRET_ACCESS_KEY env var

# --- GrammarScriptToAudio class ---

def get_grammar_tts_settings(database_name: str):
    database_name = database_name.lower().strip()

    settings = {

        "hindi": {
            "voice": "marin",
            "instructions": """
                Speak as an experienced Hindi language instructor from India and speaking in English.

                Use authentic Indian pronunciation, rhythm, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
                        },

                        "spanish": {
                            "voice": "marin",
                            "instructions": """
                Speak as an experienced Spanish language instructor and speaking in English.

                Use authentic Spanish pronunciation, rhythm, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
                        },

                        "japanese": {
                            "voice": "marin",
                            "instructions": """
                Speak as an experienced Japanese language instructor from Japan and speaking in English.

                Use authentic Japanese pronunciation, rhythm, pitch patterns, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
                        },

                        "german": {
                            "voice": "marin",
                            "instructions": """
                Speak as an experienced German language instructor from Germany and speaking in English.

                Use authentic German pronunciation, rhythm, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
                        },

                        "extractiondb": {
                            "voice": "marin",
                            "instructions": """
                Speak as an experienced French language instructor from France and speaking in English.

                Use authentic French pronunciation, rhythm, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
                        },

                        "italian": {
                            "voice": "marin",
                            "instructions": """
                Speak as an experienced Italian language instructor from Italy and speaking in English.

                Use authentic Italian pronunciation, rhythm, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
                        },

                        "chinese": {
                            "voice": "marin",
                            "instructions": """
                Speak as an experienced Mandarin Chinese language instructor and speaking in English.

                Use authentic Mandarin pronunciation, tones, rhythm, and intonation.

                This audio is for adult language learners.

                Speak clearly and confidently.

                Use a calm, patient, professional teaching style.

                Speak at approximately 85% of normal conversational speed.

                Pause naturally between ideas.

                Emphasize important grammar words and example phrases.

                Do not sound rushed.

                Maintain natural speech and avoid sounding robotic.

                The lesson should feel like a premium language-learning course.
                """
        }
    }

    if database_name not in settings:
        raise ValueError(f"Unsupported database: {database_name}")

    return settings[database_name]

class GrammarScriptToAudio:
    """
    Converts a grammar article into an instructional audio clip entirely in memory.
    Step 1 — generate_script(): rewrites raw content into a 100–120 word narration.
    Step 2 — generate_audio_bytes(): synthesises the script to MP3 bytes via TTS.
    """

    SCRIPT_MODEL = "gpt-5-mini"
    TTS_MODEL = "gpt-4o-mini-tts"

    def __init__(self, database_name: str, api_key: str | None = None):
        self.client = OpenAI(api_key=api_key or os.getenv("OPENAI_API_KEY"))
        tts_settings = get_grammar_tts_settings(database_name)
        self.voice = tts_settings["voice"]
        self.instructions = tts_settings["instructions"]

    def generate_script(self, content: str) -> str:
        """Rewrite grammar article content into a 100–120 word instructional script."""
        prompt = f"""
            You are an expert language-learning instructor creating short grammar lesson audio for a premium language-learning application aimed at adult professionals.

            Transform the grammar content below into a spoken lesson script that should be in english.

            Requirements:

            - Length: approximately 70–100 words.
            - Speak directly to the learner.
            - Start immediately with the grammar concept.
            - Do not include greetings.
            - Do not include introductions.
            - Do not include lesson titles.
            - Do not include conclusions or summaries.
            - Do not say:
              - "Today we will learn..."
              - "In this lesson..."
              - "Let's learn..."
              - "Let's talk about..."
            - Explain the grammar rule clearly and efficiently.
            - Use simple, natural language.
            - Include one or two short examples if helpful.
            - Prioritize practical understanding over technical terminology.
            - Sound like a skilled language tutor.
            - Keep the pace concise and engaging.
            - Avoid repetition.
            - The script must sound natural when spoken aloud.
            - Output only the final narration script.

            Grammar Content:

            {content}
        """
        response = self.client.responses.create(model=self.SCRIPT_MODEL, input=prompt)
        return response.output_text.strip()

    def generate_audio_bytes(self, script: str) -> bytes:
        """Synthesise script to MP3 and return raw bytes (no disk I/O)."""
        response = self.client.audio.speech.create(
            model=self.TTS_MODEL,
            voice=self.voice,
            input=script,
            instructions=self.instructions,
        )
        return response.content

@app.post("/generate-grammar-audio")
def generate_grammar_audio(body: GenerateGrammarAudioRequest):
    """
    Streams NDJSON progress while generating instructional audio for grammar lesson articles
    and uploading them to S3. Returns the S3 URL per article.

    Three input modes (combinable):
      - titles:  exact grammar lesson titles (one or many)
      - units:   all grammar articles whose lesson title matches ^unit{N}_
      - lessons: explicit [{lesson_id, article_ids?}]

    NOTE: this route does NOT write to the database — it only generates the MP3 and returns
    its URL (grammar audio feeds the external video step). There is no skip-if-present, so
    re-running regenerates everything in scope.

    limit_articles is a GLOBAL cap applied after the work list is built and ordered
    (lesson_id, sequence_id) — limit_articles=1 generates only the first article.
    """

    def stream():
        if not body.titles and not body.units and not body.lessons:
            yield _emit("error", message="Provide at least one of 'titles', 'units', or 'lessons'")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"Could not connect to database: {e}")
            return

        yield _emit("start", action="generate-grammar-audio", database=body.db.database,
                    titles=body.titles or None, units=body.units or None,
                    lessons=[t.lesson_id for t in body.lessons] if body.lessons else None)

        work_items: list[dict] = []
        seen: set = set()  # (lesson_id, article_id) dedup across modes

        def _add(row, title):
            key = (row["lesson_id"], row["article_id"])
            if key in seen:
                return
            seen.add(key)
            work_items.append({
                "lesson_id": row["lesson_id"],
                "article_id": row["article_id"],
                "sequence_id": row.get("sequence_id", 0),
                "content": row["content"],
                "lesson_title": title,
            })

        try:
            with conn.cursor() as cur:

                # Path 1: titles (exact match)
                if body.titles:
                    ph = ",".join(["%s"] * len(body.titles))
                    cur.execute(f"""
                        SELECT a.article_id, a.lesson_id, a.sequence_id, a.content, l.title
                        FROM Article a
                        JOIN Lesson l ON l.lesson_id = a.lesson_id
                        WHERE l.type = 'grammar' AND l.title IN ({ph})
                    """, tuple(body.titles))
                    for r in cur.fetchall():
                        _add(r, r["title"])

                # Path 2: units (REGEXP prefix, exact unit boundary)
                if body.units:
                    regexp_clauses = " OR ".join(["l.title REGEXP %s" for _ in body.units])
                    regexp_params = [f"^unit{u}_" for u in body.units]
                    cur.execute(f"""
                        SELECT a.article_id, a.lesson_id, a.sequence_id, a.content, l.title
                        FROM Article a
                        JOIN Lesson l ON l.lesson_id = a.lesson_id
                        WHERE ({regexp_clauses}) AND l.type = 'grammar'
                    """, tuple(regexp_params))
                    for r in cur.fetchall():
                        _add(r, r["title"])

                # Path 3: explicit lessons
                if body.lessons:
                    for target in body.lessons:
                        cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE lesson_id = %s",
                                    (target.lesson_id,))
                        lesson_row = cur.fetchone()
                        if not lesson_row:
                            yield _emit("lesson_error", lesson_id=target.lesson_id, error="lesson not found")
                            continue
                        if lesson_row["type"] != "grammar":
                            yield _emit("lesson_error", lesson_id=target.lesson_id,
                                        error=f"lesson type '{lesson_row['type']}' — only grammar supported")
                            continue

                        if target.article_ids:
                            fmt = ",".join(["%s"] * len(target.article_ids))
                            cur.execute(
                                f"SELECT article_id, lesson_id, sequence_id, content FROM Article "
                                f"WHERE lesson_id = %s AND article_id IN ({fmt})",
                                (target.lesson_id, *target.article_ids),
                            )
                        else:
                            cur.execute(
                                "SELECT article_id, lesson_id, sequence_id, content FROM Article "
                                "WHERE lesson_id = %s",
                                (target.lesson_id,),
                            )
                        for r in cur.fetchall():
                            _add(r, lesson_row["title"])

        except Exception as e:
            conn.close()
            yield _emit("error", message=f"Failed to build work list: {e}")
            return

        # order globally, then apply the global cap
        work_items.sort(key=lambda w: (w["lesson_id"], w["sequence_id"]))
        if body.limit_articles is not None:
            work_items = work_items[: int(body.limit_articles)]

        yield _emit("found", total=len(work_items))
        if not work_items:
            conn.close()
            yield _emit("summary", total_items=0, succeeded=0, failed=0)
            return

        generator = GrammarScriptToAudio(database_name=body.db.database, api_key=body.openai_api_key)
        succeeded = failed = 0

        try:
            for idx, item in enumerate(work_items, start=1):
                yield _emit("processing", n=idx, total=len(work_items),
                            lesson_title=item["lesson_title"], lesson_id=item["lesson_id"],
                            article_id=item["article_id"], sequence_id=item["sequence_id"])
                try:
                    # 1) rewrite content into instructional script
                    script = generator.generate_script(item["content"])
                    # 2) synthesise to MP3 bytes (no disk I/O)
                    audio_bytes = generator.generate_audio_bytes(script)
                    # 3) upload to S3
                    key = (f"{body.s3_prefix}/"
                           f"{item['lesson_id']}_{item['article_id']}_"
                           f"{_safe_slug(item['lesson_title'])}.mp3")
                    public_url = _upload_to_s3_public(
                        audio_bytes, key, body.s3_bucket, body.aws_region,
                        body.aws_access_key_id, body.aws_secret_access_key,
                        content_type="audio/mpeg",
                    )
                    succeeded += 1
                    yield _emit("success", lesson_id=item["lesson_id"], article_id=item["article_id"],
                                sequence_id=item["sequence_id"], lesson_title=item["lesson_title"],
                                audio_url=public_url)

                except Exception as e:
                    failed += 1
                    yield _emit("failed", lesson_id=item["lesson_id"], article_id=item["article_id"],
                                sequence_id=item["sequence_id"], lesson_title=item["lesson_title"],
                                error=str(e))

            yield _emit("summary", total_items=len(work_items), succeeded=succeeded, failed=failed)
            logger.info("generate-grammar-audio done | items=%d ok=%d fail=%d",
                        len(work_items), succeeded, failed)

        except Exception as e:
            logger.error("generate-grammar-audio crashed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e))
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


class GenerateListeningQuestionsRequest(BaseModel):
    db: DBConfig
    s3_bucket: str
    s3_audio_prefix: str                              # S3 prefix where MP3 files live
    s3_output_prefix: str                             # S3 prefix where question JSONs are written
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    openai_api_key: str | None = None
    num_questions: int = 10
    model: str = "gpt-4o-mini"
    transcription_model: str = "whisper-1"
    temperature: float = 0.4
    force: bool = False                               # re-process even if output JSON already in S3
    limit: int | None = None                          # max number of audio files to process
    lesson_type: str = "listening"                    # Lesson.type value for inserted lessons
    cefr_level: str = "A1"                           # fallback when no mapping matches
    cefr_mapping: list[CefrRange] | None = None      # unit-range → CEFR level (same as other routes)


# =============================================================================
#  /generate-listening-questions  (streaming)
#
#  For each MP3 under s3_bucket/s3_audio_prefix/:
#    1. transcribe (Whisper)
#    2. generate MCQs (GPT)
#    3. upload the question JSON to s3_output_prefix/<stem>.json
#    4. insert Lesson + Audio + Questions + Answers
#
#  Skips a file if its output JSON already exists in S3 (unless force=True) AND
#  now also skips if a lesson with that title already exists in the DB (prevents
#  duplicate lessons when S3 and DB drift) — unless force=True.
#
#  Streams NDJSON so a long run (many Whisper calls) doesn't sit silent and get
#  cut by the proxy/ALB, and so you can watch progress.
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse / OpenAI /
#         _make_s3_client / _list_s3_mp3_keys / _s3_key_exists / _transcribe_from_s3 /
#         _build_listening_lesson_title / _extract_unit_number_from_filename /
#         _build_listening_mcq_messages / _parse_mcq_json_response / _validate_mcq_list /
#         _insert_listening_lesson / resolve_cefr / Path / json / os.
# =============================================================================


@app.post("/generate-listening-questions")
def generate_listening_questions(body: GenerateListeningQuestionsRequest):

    def stream():
        effective_api_key = body.openai_api_key or os.getenv("OPENAI_API_KEY")
        if not effective_api_key:
            yield _emit("error", message="OpenAI API key is required (send openai_api_key in the request).")
            return

        openai_client = OpenAI(api_key=effective_api_key)
        s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)

        try:
            audio_keys = _list_s3_mp3_keys(s3, body.s3_bucket, body.s3_audio_prefix)
        except Exception as e:
            yield _emit("error", message=f"Failed to list S3 objects: {e}")
            return

        if not audio_keys:
            yield _emit("error",
                        message=f"No .mp3 files found under s3://{body.s3_bucket}/{body.s3_audio_prefix}")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"Could not connect to database: {e}")
            return

        yield _emit("start", action="generate-listening-questions",
                    database=body.db.database, audio_files=len(audio_keys),
                    force=body.force, limit=body.limit)

        output_prefix = body.s3_output_prefix.rstrip("/")
        succeeded = failed = skipped = 0
        processed = 0

        try:
            for idx, audio_key in enumerate(audio_keys, start=1):
                # limit caps files actually PROCESSED (skips don't count) — stop once reached
                if body.limit is not None and processed >= int(body.limit):
                    yield _emit("limit_reached", limit=int(body.limit))
                    break

                stem = Path(audio_key).stem
                out_key = f"{output_prefix}/{stem}.json"
                title = _build_listening_lesson_title(Path(audio_key).name)

                # skip: output JSON already in S3
                if not body.force and _s3_key_exists(s3, body.s3_bucket, out_key):
                    skipped += 1
                    yield _emit("skipped", n=idx, total=len(audio_keys), audio_key=audio_key,
                                lesson_title=title, reason="output JSON already exists in S3")
                    continue

                # skip: lesson with this title already in the DB (prevents duplicate lessons)
                if not body.force:
                    with conn.cursor() as cur:
                        cur.execute("SELECT lesson_id FROM Lesson WHERE title = %s", (title,))
                        existing = cur.fetchone()
                    if existing:
                        skipped += 1
                        yield _emit("skipped", n=idx, total=len(audio_keys), audio_key=audio_key,
                                    lesson_title=title, lesson_id=existing["lesson_id"],
                                    reason="lesson already exists in DB")
                        continue

                processed += 1
                yield _emit("processing", n=idx, total=len(audio_keys),
                            audio_key=audio_key, lesson_title=title)

                try:
                    # 1. transcribe
                    transcript = _transcribe_from_s3(
                        s3, body.s3_bucket, audio_key, openai_client, body.transcription_model
                    )
                    yield _emit("transcribed", lesson_title=title, chars=len(transcript or ""))

                    # 2. generate MCQs
                    unit_number = _extract_unit_number_from_filename(Path(audio_key).name)
                    messages = _build_listening_mcq_messages(transcript, unit_number, body.num_questions)
                    resp = openai_client.chat.completions.create(
                        model=body.model, messages=messages, temperature=body.temperature,
                    )
                    raw_content = resp.choices[0].message.content.strip()

                    # 3. parse + validate
                    mcq_items = _parse_mcq_json_response(raw_content)
                    _validate_mcq_list(mcq_items, body.num_questions)
                    yield _emit("questions_generated", lesson_title=title, questions=len(mcq_items))

                    # 4. upload JSON to S3
                    s3.put_object(
                        Bucket=body.s3_bucket,
                        Key=out_key,
                        Body=json.dumps(mcq_items, ensure_ascii=False, indent=2).encode("utf-8"),
                        ContentType="application/json",
                    )

                    # 5. insert Lesson + Audio + Questions + Answers
                    audio_url = (f"https://{body.s3_bucket}.s3.{body.aws_region}.amazonaws.com/{audio_key}")
                    cefr = resolve_cefr(Path(audio_key).name, body.cefr_mapping, body.cefr_level)
                    lesson_id = _insert_listening_lesson(
                        conn, title, body.lesson_type, cefr, mcq_items, audio_url
                    )
                    conn.commit()

                    succeeded += 1
                    yield _emit("success", n=idx, total=len(audio_keys), lesson_title=title,
                                lesson_id=lesson_id, questions=len(mcq_items),
                                cefr_level=cefr, audio_url=audio_url, output_key=out_key)

                except Exception as e:
                    conn.rollback()
                    failed += 1
                    yield _emit("failed", n=idx, total=len(audio_keys), audio_key=audio_key,
                                lesson_title=title, error=str(e))

            yield _emit("summary", total_audio_files=len(audio_keys),
                        processed=succeeded + failed, succeeded=succeeded,
                        failed=failed, skipped=skipped)
            logger.info("generate-listening-questions done | files=%d ok=%d fail=%d skip=%d",
                        len(audio_keys), succeeded, failed, skipped)

        except Exception as e:
            conn.rollback()
            logger.error("generate-listening-questions crashed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e))
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


class GenerateVocabAudioRequest(BaseModel):
    db: DBConfig
    titles: list[str] | None = None 
    s3_bucket: str
    s3_prefix: str                                    # S3 prefix where MP3s are uploaded
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    openai_api_key: str | None = None
    limit: int | None = None                          # cap number of questions to process
    tts_model: str = "gpt-4o-mini-tts"
    voice: str = "alloy"
    tts_instructions: str | None = (
        "Speak slowly and clearly with a warm, friendly, teacher-like tone. "
        "When the text is English, use standard native English with no foreign accent "
        "(clear, neutral pronunciation)."
    )
    source_language: str = "en"                       # AWS Translate source language code
    target_language: str = "hi"                       # AWS Translate target language code (e.g. "hi", "es")


def get_tts_settings(database_name: str):
    database_name = database_name.lower().strip()

    settings = {
        "hindi": {
            "voice": "nova",
            "instructions": """
                Speak as a native Hindi speaker from India.

                This audio is for a language-learning application.

                Use authentic Indian Hindi pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak slowly.

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
                        },

        "japanese": {
            "voice": "nova",
            "instructions": """
                Speak as a native Japanese speaker from Japan.

                This audio is for a language-learning application.

                Use authentic Japanese pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak slowly

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
                        },

       "german": {
           "voice": "nova",
           "instructions": """
                Speak as a native German speaker from Germany.

                This audio is for a language-learning application.

                Use authentic German pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak slowly

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
                        },

        "extractiondb": {
            "voice": "nova",
            "instructions": """
                Speak as a native French speaker from France.

                This audio is for a language-learning application.

                Use authentic French pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak slowly

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
                        },

        "italian": {
            "voice": "nova",
            "instructions": """
                Speak as a native Italian speaker from Italy.

                This audio is for a language-learning application.

                Use authentic Italian pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak slowly

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
                        },

        "spanish": {
            "voice": "nova",
            "instructions": """
                Speak as a native Spanish speaker from Spain.

                This audio is for a language-learning application.

                Use authentic Spanish pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak Slowly

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
                        },

        "chinese": {
            "voice": "nova",
            "instructions": """
                Speak as a native Chinese speaker from China.

                This audio is for a language-learning application.

                Use authentic Chinese pronunciation, pitch patterns, rhythm, and intonation.

                Do not use an American or foreign accent.

                Speak clearly, calmly, patiently, like a teacher.

                Use a slightly slower pace than everyday conversation so learners can clearly hear each word.

                Pause naturally between words and phrases.

                Maintain natural pronunciation and rhythm.

                Do not sound rushed.

                Speak slowly

                Do not translate, explain, or add words.

                Only speak the provided text.
                """
        }
    }

    if database_name not in settings:
        raise ValueError(f"Unsupported database: {database_name}")

    return settings[database_name]

# NOTE: GenerateVocabAudioRequest needs:  title: str | None = None
# (source_language / target_language / voice / tts_instructions are no longer
#  used by this route — audio is spoken from question_text with per-language
#  settings from get_tts_settings — but they can stay on the model harmlessly.)


import re

# --- Japanese-only text normalizer (strip romaji/English before TTS) -------------
_JP_PATTERN = re.compile(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF\u30FC。、「」・？！…]+')
_JP_CHAR = re.compile(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]')


import re

# --- Japanese-only text normalizer (strip romaji/English before TTS) -------------
_JP_PATTERN = re.compile(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF\u30FC。、「」・？！…]+')
_JP_CHAR = re.compile(r'[\u3040-\u309F\u30A0-\u30FF\u4E00-\u9FFF]')


def normalize_japanese_tts(text: str) -> str | None:
    """
    Return only the Japanese portion of `text` so TTS never reads romaji/English glosses.
    Returns None if there's no Japanese to speak -> caller skips + reports (a missing file
    is safer than one that reads romaji aloud).
    """
    if not text:
        return None
    matches = _JP_PATTERN.findall(text)
    if not matches:
        return None
    joined = " ".join(m.strip() for m in matches if m.strip())
    if not joined or not _JP_CHAR.search(joined):
        return None
    return joined


@app.post("/generate-vocab-audio")
def generate_vocab_audio(body: GenerateVocabAudioRequest):
    """
    Streams NDJSON progress while generating audio for vocabulary questions whose audio
    is MISSING (no Audio row, or a row whose audio_url is NULL/blank).

    Scope: body.titles (one or many lessons); omit titles to scan the whole DB.
    Audio is spoken from the QUESTION TEXT, using per-language voice + instructions.
    For the Japanese DB, romaji/English is stripped so only Japanese is spoken; a
    question with no Japanese content is skipped and reported (never read as romaji).
    """

    def stream():
        effective_api_key = body.openai_api_key or os.getenv("OPENAI_API_KEY")
        if not effective_api_key:
            yield _emit("error", message="OpenAI API key is required."); return

        openai_client = OpenAI(api_key=effective_api_key)
        s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"Could not connect to database: {e}")
            return

        database_name = body.db.database
        is_japanese = database_name.strip().lower() == "japanese"
        try:
            tts_settings = get_tts_settings(database_name)
        except ValueError as e:
            conn.close()
            yield _emit("error", message=str(e))
            return

        voice = tts_settings["voice"]
        tts_instructions = tts_settings["instructions"]
        yield _emit("start", action="generate-vocab-audio",
                    database=database_name, japanese_normalization=is_japanese,
                    scope=(body.titles if body.titles else "whole_db"))

        # Vocab questions whose audio is missing: no Audio row, OR a row with blank/NULL url.
        sql = """
            SELECT q.question_id, q.lesson_id, q.sequence_id, q.question_text,
                   l.title AS lesson_title
            FROM Question q
            JOIN Lesson l ON q.lesson_id = l.lesson_id
            LEFT JOIN Audio au ON au.question_id = q.question_id
            WHERE l.type = 'vocabulary'
              AND (au.question_id IS NULL OR au.audio_url IS NULL OR TRIM(au.audio_url) = '')
        """
        params: list = []
        if body.titles:
            placeholders = ",".join(["%s"] * len(body.titles))
            sql += f" AND l.title IN ({placeholders})"
            params.extend(body.titles)
        sql += " ORDER BY q.lesson_id, q.sequence_id"
        if body.limit is not None:
            sql += " LIMIT %s"
            params.append(int(body.limit))

        try:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                rows = cur.fetchall()
        except Exception as e:
            conn.close()
            yield _emit("error", message=f"Failed to query vocab questions: {e}"); return

        yield _emit("found", total=len(rows))
        if not rows:
            conn.close()
            yield _emit("summary", total_found=0, succeeded=0, failed=0, skipped=0)
            return

        prefix = body.s3_prefix.rstrip("/")
        succeeded = failed = skipped = 0

        try:
            for idx, row in enumerate(rows, start=1):
                question_id   = row["question_id"]
                lesson_id     = row["lesson_id"]
                sequence_id   = row["sequence_id"]
                question_text = row["question_text"]
                lesson_title  = row["lesson_title"]

                safe_title = _safe_slug(lesson_title)
                s3_key    = f"{prefix}/{lesson_id}_{question_id}_{safe_title}.mp3"
                audio_url = f"https://{body.s3_bucket}.s3.{body.aws_region}.amazonaws.com/{s3_key}"

                yield _emit("processing", n=idx, total=len(rows),
                            lesson_title=lesson_title, question_id=question_id,
                            sequence_id=sequence_id, question_text=question_text)

                try:
                    if not (question_text or "").strip():
                        raise ValueError("question_text is blank - nothing to speak")

                    # Japanese: strip romaji/English so TTS speaks only the Japanese.
                    # If there's no Japanese content, fall back to the original text (generate normally).
                    text_to_speak = question_text
                    if is_japanese:
                        jp = normalize_japanese_tts(question_text)
                        if jp is not None:
                            text_to_speak = jp
                            if jp != question_text:
                                yield _emit("normalized", question_id=question_id,
                                            original=question_text, spoken=jp)

                    # 1. TTS from the (normalized) text
                    audio_bytes = _generate_tts_bytes(
                        openai_client, text_to_speak, body.tts_model, voice, tts_instructions,
                    )

                    # 2. Upload (deterministic key -> overwrites any stale file)
                    s3.put_object(Bucket=body.s3_bucket, Key=s3_key,
                                  Body=audio_bytes, ContentType="audio/mpeg")

                    # 3. UPDATE blank row in place, or INSERT if none
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT audio_id FROM Audio WHERE question_id = %s ORDER BY audio_id ASC LIMIT 1",
                            (question_id,),
                        )
                        existing = cur.fetchone()

                    action = "updated" if existing else "inserted"
                    if existing:
                        metadata = json.dumps({"source": "tts", "tts_model": body.tts_model,
                                               "voice": voice, "spoken_text": text_to_speak})
                        with conn.cursor() as cur:
                            try:
                                cur.execute(
                                    "UPDATE Audio SET audio_url = %s, audio_metadata = CAST(%s AS JSON) "
                                    "WHERE audio_id = %s",
                                    (audio_url, metadata, existing["audio_id"]),
                                )
                            except pymysql.err.OperationalError as e:
                                if e.args[0] == 1054:
                                    cur.execute("UPDATE Audio SET audio_url = %s WHERE audio_id = %s",
                                                (audio_url, existing["audio_id"]))
                                else:
                                    raise
                    else:
                        _insert_vocab_audio_row(
                            conn, lesson_id, question_id, sequence_id, audio_url, body.tts_model, voice,
                        )

                    conn.commit()
                    succeeded += 1
                    yield _emit("success", question_id=question_id, sequence_id=sequence_id,
                                lesson_title=lesson_title, audio_url=audio_url, row=action)

                except Exception as e:
                    conn.rollback()
                    failed += 1
                    yield _emit("failed", question_id=question_id, sequence_id=sequence_id,
                                lesson_title=lesson_title, error=str(e))

            yield _emit("summary", total_found=len(rows),
                        succeeded=succeeded, failed=failed, skipped=skipped)
            logger.info("generate-vocab-audio done | found=%d ok=%d fail=%d skip=%d",
                        len(rows), succeeded, failed, skipped)

        except Exception as e:
            conn.rollback()
            logger.error("generate-vocab-audio crashed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e))
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


@app.post("/generate-unit-images")
def generate_unit_images(body: GenerateUnitImagesRequest):
    """
    For each unit (auto-detected from lesson titles or from the provided units list):
      1. Gathers all Article content and Q&A text from every lesson in that unit.
      2. If no content is found, skips the unit.
      3. Sends the content to GPT to derive a 1–3 word theme.
      4. Generates an image representing that theme via the image model.
      5. Uploads the image to s3_bucket/s3_prefix/unit{n}_{theme_slug}.png.

    Units that already have an image at the prefix are skipped unless force=True.
    """
    effective_api_key = body.openai_api_key or os.getenv("OPENAI_API_KEY")
    if not effective_api_key:
        raise HTTPException(status_code=400, detail="OpenAI API key is required.")

    openai_client = OpenAI(api_key=effective_api_key)
    s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    # --- Resolve unit numbers ---
    try:
        if body.units:
            unit_numbers = sorted(set(body.units))
        else:
            with conn.cursor() as cur:
                cur.execute("SELECT DISTINCT title FROM Lesson WHERE title REGEXP '^unit[0-9]'")
                titles = [row["title"] for row in cur.fetchall()]
            unit_numbers = sorted({
                n for t in titles
                if (n := _extract_unit_num_from_title(t)) is not None
            })
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to resolve unit numbers: {e}")

    prefix = body.s3_prefix.rstrip("/")
    results = []
    succeeded = failed = skipped = no_content = 0

    for unit_number in unit_numbers:
        log: dict = {
            "unit": unit_number,
            "theme": None,
            "image_url": None,
            "status": None,
            "error": None,
        }

        try:
            # 1. Skip if an image already exists in S3 and force=False
            if not body.force and _s3_unit_image_exists(s3, body.s3_bucket, prefix, unit_number):
                log["status"] = "skipped"
                skipped += 1
                results.append(log)
                continue

            # 2. Gather all content for this unit
            unit_like = f"unit{unit_number}\\_%"
            with conn.cursor() as cur:
                # Articles
                cur.execute(
                    "SELECT a.content FROM Article a "
                    "JOIN Lesson l ON l.lesson_id = a.lesson_id "
                    "WHERE l.title LIKE %s ORDER BY l.lesson_id, a.sequence_id",
                    (unit_like,),
                )
                articles = [row["content"] for row in cur.fetchall()]

                # Q&A
                cur.execute(
                    "SELECT q.question_text, ans.answer_text "
                    "FROM Question q "
                    "JOIN Lesson l ON l.lesson_id = q.lesson_id "
                    "LEFT JOIN Answer ans ON ans.question_id = q.question_id AND ans.is_correct = 1 "
                    "WHERE l.title LIKE %s ORDER BY l.lesson_id, q.sequence_id",
                    (unit_like,),
                )
                qa_rows = cur.fetchall()

            content_parts = []
            if articles:
                content_parts.append("=== Articles ===\n" + "\n\n".join(articles))
            if qa_rows:
                qa_lines = [
                    f"Q: {r['question_text']}" + (f"  A: {r['answer_text']}" if r["answer_text"] else "")
                    for r in qa_rows
                ]
                content_parts.append("=== Vocabulary Q&A ===\n" + "\n".join(qa_lines))

            if not content_parts:
                log["status"] = "skipped_no_content"
                no_content += 1
                results.append(log)
                continue

            # Truncate to avoid token limits (~6000 chars is safe for theme extraction)
            combined = "\n\n".join(content_parts)[:6000]

            # 3. Derive theme via GPT
            messages = _build_unit_theme_messages(combined, unit_number)
            theme_resp = openai_client.chat.completions.create(
                model=body.theme_model,
                messages=messages,
                temperature=body.temperature,
            )
            raw_theme = (theme_resp.choices[0].message.content or "").strip()
            # Normalise: keep only word chars, collapse spaces to underscore
            theme_slug = re.sub(r"\s+", "_", re.sub(r"[^\w\s]", "", raw_theme).strip()).lower()
            if not theme_slug:
                theme_slug = f"unit{unit_number}"

            log["theme"] = raw_theme

            # 4. Generate image
            image_prompt = (
                f"A vibrant, simple illustration representing the theme '{raw_theme}' "
                f"for a language learning app. Realistic style, no text in the image."
            )
            img_result = openai_client.images.generate(
                model=body.image_model,
                prompt=image_prompt,
                size="1024x1024",
            )
            if not img_result.data or not getattr(img_result.data[0], "b64_json", None):
                raise RuntimeError("No image data returned from image model.")
            image_bytes = base64.b64decode(img_result.data[0].b64_json)

            # 5. Upload to S3
            s3_key = f"{prefix}/unit{unit_number}_{theme_slug}.png"
            public_url = _upload_to_s3_public(
                image_bytes, s3_key, body.s3_bucket, body.aws_region,
                body.aws_access_key_id, body.aws_secret_access_key,
            )

            log["status"] = "success"
            log["image_url"] = public_url
            succeeded += 1

        except Exception as e:
            log["status"] = "failed"
            log["error"] = str(e)
            failed += 1

        results.append(log)

    conn.close()

    all_ok = failed == 0
    return JSONResponse(
        content={
            "status": "ok" if all_ok else "partial",
            "summary": {
                "total_units": len(unit_numbers),
                "succeeded": succeeded,
                "failed": failed,
                "skipped": skipped,
                "skipped_no_content": no_content,
            },
            "results": results,
        },
        status_code=200 if all_ok else 207,
    )


@app.post("/generate-article-questions")
def generate_article_questions(body: GenerateArticleQuestionsRequest):
    """
    For each lesson (filtered by lesson_type or lesson_ids):
      1. Fetches all articles for the lesson.
      2. If no articles exist, skips the lesson.
      3. Concatenates all article content and generates MCQ questions via GPT.
      4. Uploads the question JSON to s3_bucket/s3_output_prefix/<lesson_id>.json.
      5. Inserts Questions + Answers into the existing lesson in the database.

    Lessons whose output JSON already exists in S3 are skipped unless force=True.
    Requires at least one of lesson_type or lesson_ids.
    """
    if not body.lesson_type and not body.lesson_ids:
        raise HTTPException(
            status_code=400,
            detail="Provide at least one of 'lesson_type' or 'lesson_ids'.",
        )

    effective_api_key = body.openai_api_key or os.getenv("OPENAI_API_KEY")
    if not effective_api_key:
        raise HTTPException(status_code=400, detail="OpenAI API key is required.")

    openai_client = OpenAI(api_key=effective_api_key)
    s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    # --- Fetch target lessons ---
    try:
        with conn.cursor() as cur:
            lesson_rows: list[dict] = []

            if body.lesson_type:
                cur.execute(
                    "SELECT lesson_id, title FROM Lesson WHERE type = %s",
                    (body.lesson_type,),
                )
                lesson_rows.extend(cur.fetchall())

            if body.lesson_ids:
                fmt = ",".join(["%s"] * len(body.lesson_ids))
                cur.execute(
                    f"SELECT lesson_id, title FROM Lesson WHERE lesson_id IN ({fmt})",
                    tuple(body.lesson_ids),
                )
                lesson_rows.extend(cur.fetchall())

        # Deduplicate by lesson_id (in case lesson_type and lesson_ids overlap)
        seen: set[int] = set()
        unique_lessons: list[dict] = []
        for row in lesson_rows:
            if row["lesson_id"] not in seen:
                seen.add(row["lesson_id"])
                unique_lessons.append(row)

    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=f"Failed to fetch lessons: {e}")

    output_prefix = body.s3_output_prefix.rstrip("/")
    results = []
    succeeded = failed = skipped = no_articles = has_questions = 0

    for lesson in unique_lessons:
        if body.limit is not None and succeeded >= body.limit:
            break

        lesson_id = lesson["lesson_id"]
        lesson_title = lesson["title"]
        out_key = f"{output_prefix}/{lesson_id}.json"

        log: dict = {
            "lesson_id": lesson_id,
            "lesson_title": lesson_title,
            "output_key": out_key,
            "status": None,
            "num_questions": None,
            "error": None,
        }

        try:
            # 1. Fetch articles for this lesson
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT sequence_id, content FROM Article "
                    "WHERE lesson_id = %s ORDER BY sequence_id",
                    (lesson_id,),
                )
                articles = cur.fetchall()

            if not articles:
                log["status"] = "skipped_no_articles"
                no_articles += 1
                results.append(log)
                continue

            # 2. Skip if questions already exist in the DB and force=False
            if not body.force:
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT COUNT(*) AS cnt FROM Question WHERE lesson_id = %s",
                        (lesson_id,),
                    )
                    if cur.fetchone()["cnt"] > 0:
                        log["status"] = "skipped_has_questions"
                        has_questions += 1
                        results.append(log)
                        continue

            # 3. Skip if output already exists in S3 and force=False
            if not body.force and _s3_key_exists(s3, body.s3_bucket, out_key):
                log["status"] = "skipped"
                skipped += 1
                results.append(log)
                continue

            # 3. Concatenate article content
            combined_content = "\n\n".join(
                f"[Article {a['sequence_id']}]\n{a['content']}" for a in articles
            )

            # 4. Generate MCQ questions
            messages = _build_article_mcq_messages(combined_content, lesson_title, body.num_questions)
            resp = openai_client.chat.completions.create(
                model=body.model,
                messages=messages,
                temperature=body.temperature,
            )
            raw_content = resp.choices[0].message.content.strip()

            # 5. Parse and validate
            mcq_items = _parse_mcq_json_response(raw_content)
            _validate_article_mcq_list(mcq_items, body.num_questions)

            # 6. Upload JSON to S3
            s3.put_object(
                Bucket=body.s3_bucket,
                Key=out_key,
                Body=json.dumps(mcq_items, ensure_ascii=False, indent=2).encode("utf-8"),
                ContentType="application/json",
            )

            # 7. Insert Questions + Answers into the existing lesson
            _insert_article_lesson_questions(conn, lesson_id, mcq_items)
            conn.commit()

            log["status"] = "success"
            log["num_questions"] = len(mcq_items)
            succeeded += 1

        except Exception as e:
            conn.rollback()
            log["status"] = "failed"
            log["error"] = str(e)
            failed += 1

        results.append(log)

    conn.close()

    all_ok = failed == 0
    return JSONResponse(
        content={
            "status": "ok" if all_ok else "partial",
            "summary": {
                "total_lessons": len(unique_lessons),
                "succeeded": succeeded,
                "failed": failed,
                "skipped_s3_exists": skipped,
                "skipped_has_questions": has_questions,
                "skipped_no_articles": no_articles,
            },
            "results": results,
        },
        status_code=200 if all_ok else 207,
    )


@app.post("/migrate-audio-table")
def migrate_audio_table(body: DBConfig):
    """
    Adds missing columns to the Audio table for databases created from an older schema.
    Currently adds: audio_metadata JSON NULL (if absent).
    Safe to run multiple times — uses ALTER TABLE ... ADD COLUMN IF NOT EXISTS.
    """
    try:
        conn = connect_to_db(body)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    applied = []
    already_present = []

    migrations = [
        ("audio_metadata", "ALTER TABLE Audio ADD COLUMN audio_metadata JSON NULL"),
    ]

    try:
        with conn.cursor() as cur:
            for col_name, ddl in migrations:
                cur.execute("SHOW COLUMNS FROM Audio LIKE %s", (col_name,))
                if cur.fetchone():
                    already_present.append(col_name)
                else:
                    cur.execute(ddl)
                    applied.append(col_name)
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Migration failed: {e}")
    finally:
        conn.close()

    return JSONResponse(content={
        "status": "ok",
        "database": body.database,
        "columns_added": applied,
        "columns_already_present": already_present,
    })


# =============================================================================
#  /link-grammar-videos  (streaming, title-scoped, true replace)
#
#  Scans an S3 prefix for .mp4 files, extracts lesson_id from each filename
#  (first underscore-separated segment, e.g. '3608' from
#  '3608_..._unit12_grammar_generated.mp4'), and links each to its grammar lesson.
#
#  Scope:
#    titles = None        -> process every .mp4 found in the prefix
#    titles = [...]        -> resolve to lesson_ids; only process files for those lessons,
#                            and WARN for any requested title whose .mp4 isn't in S3
#
#  force:
#    False -> link only lessons with NO video yet; skip ones already linked
#    True  -> REPLACE: delete the lesson's existing VideoLesson + orphaned Video,
#             then insert the regenerated video (use after QA regenerates)
#
#  Streams NDJSON. Reuses connect_to_db / DBConfig / logger / _emit /
#  StreamingResponse / _make_s3_client / Path.
# =============================================================================


class LinkGrammarVideosRequest(BaseModel):
    db: DBConfig
    s3_bucket: str
    s3_prefix: str
    titles: list[str] | None = None                   # restrict to these grammar lessons
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    force: bool = False                               # True = replace existing video for the lesson


@app.post("/link-grammar-videos")
def link_grammar_videos(body: LinkGrammarVideosRequest):

    def stream():
        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)

        yield _emit("start", action="link-grammar-videos", database=body.db.database,
                    s3_prefix=body.s3_prefix, force=body.force,
                    titles=body.titles or None)

        # --- resolve title scope -> {lesson_id: title} ---
        wanted_ids: set | None = None
        title_by_id: dict = {}
        try:
            if body.titles:
                with conn.cursor() as cur:
                    ph = ",".join(["%s"] * len(body.titles))
                    cur.execute(
                        f"SELECT lesson_id, title FROM Lesson "
                        f"WHERE type='grammar' AND title IN ({ph})",
                        tuple(body.titles),
                    )
                    rows = cur.fetchall()
                title_by_id = {r["lesson_id"]: r["title"] for r in rows}
                wanted_ids = set(title_by_id.keys())
                found_titles = set(title_by_id.values())
                missing = [t for t in body.titles if t not in found_titles]
                if missing:
                    yield _emit("warning", message="titles not found as grammar lessons", titles=missing)
                if not wanted_ids:
                    conn.close()
                    yield _emit("error", message="none of the requested titles resolved to grammar lessons")
                    return
        except Exception as e:
            conn.close()
            yield _emit("error", message=f"Failed to resolve titles: {e}")
            return

        # --- list .mp4 keys under the prefix ---
        prefix = body.s3_prefix.rstrip("/") + "/"
        video_keys: list[str] = []
        try:
            paginator = s3.get_paginator("list_objects_v2")
            for page in paginator.paginate(Bucket=body.s3_bucket, Prefix=prefix):
                for obj in page.get("Contents", []):
                    key = obj["Key"]
                    if key.lower().endswith(".mp4"):
                        video_keys.append(key)
            video_keys.sort()
        except Exception as e:
            conn.close()
            yield _emit("error", message=f"Failed to list S3 objects: {e}")
            return

        yield _emit("found", total_mp4=len(video_keys))

        totals = {"inserted": 0, "replaced": 0, "skipped": 0, "errors": 0}
        seen_lesson_ids: set = set()

        try:
            for key in video_keys:
                filename = Path(key).name
                first_segment = filename.split("_")[0]
                if not first_segment.isdigit():
                    totals["errors"] += 1
                    yield _emit("file_error", key=key,
                                error=f"cannot parse lesson_id from filename '{filename}'")
                    continue
                lesson_id = int(first_segment)

                # title scope filter
                if wanted_ids is not None and lesson_id not in wanted_ids:
                    yield _emit("file_skipped", key=key, lesson_id=lesson_id,
                                reason="not in requested titles")
                    continue

                try:
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT lesson_id, title FROM Lesson WHERE lesson_id = %s AND type = 'grammar'",
                            (lesson_id,),
                        )
                        lrow = cur.fetchone()
                        if not lrow:
                            totals["errors"] += 1
                            yield _emit("file_error", key=key,
                                        error=f"no grammar lesson with lesson_id={lesson_id}")
                            continue
                        title = lrow["title"]

                        cur.execute("SELECT video_id FROM VideoLesson WHERE lesson_id = %s", (lesson_id,))
                        existing_video_ids = [r["video_id"] for r in cur.fetchall()]

                        was_replace = False
                        if existing_video_ids:
                            if not body.force:
                                totals["skipped"] += 1
                                yield _emit("skipped", key=key, lesson_id=lesson_id, title=title,
                                            reason="already linked (use force=True to replace)")
                                continue
                            cur.execute("DELETE FROM VideoLesson WHERE lesson_id = %s", (lesson_id,))
                            vfmt = ",".join(["%s"] * len(existing_video_ids))
                            cur.execute(
                                f"DELETE v FROM Video v "
                                f"LEFT JOIN VideoLesson vl ON vl.video_id = v.video_id "
                                f"WHERE v.video_id IN ({vfmt}) AND vl.video_id IS NULL",
                                tuple(existing_video_ids),
                            )
                            was_replace = True

                        video_url = f"https://{body.s3_bucket}.s3.{body.aws_region}.amazonaws.com/{key}"
                        cur.execute("INSERT INTO Video (url) VALUES (%s)", (video_url,))
                        video_id = cur.lastrowid
                        cur.execute(
                            "INSERT INTO VideoLesson (video_id, lesson_id) VALUES (%s, %s)",
                            (video_id, lesson_id),
                        )

                    conn.commit()
                    seen_lesson_ids.add(lesson_id)
                    if was_replace:
                        totals["replaced"] += 1
                        yield _emit("replaced", key=key, lesson_id=lesson_id, title=title,
                                    video_id=video_id, url=video_url, removed_video_ids=existing_video_ids)
                    else:
                        totals["inserted"] += 1
                        yield _emit("inserted", key=key, lesson_id=lesson_id, title=title,
                                    video_id=video_id, url=video_url)

                except Exception as e:
                    conn.rollback()
                    totals["errors"] += 1
                    yield _emit("file_error", key=key, lesson_id=lesson_id, error=str(e))

            # warn about requested titles that never matched a file
            if wanted_ids is not None:
                no_file = [title_by_id[i] for i in wanted_ids if i not in seen_lesson_ids]
                if no_file:
                    yield _emit("warning", message="requested titles had no matching .mp4 in S3",
                                titles=sorted(no_file))

            yield _emit("summary", totals=totals)
            logger.info("link-grammar-videos done | %s", totals)

        except Exception as e:
            conn.rollback()
            logger.error("link-grammar-videos crashed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e))
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


@app.post("/ingest-unit-images")
def ingest_unit_images(body: IngestUnitImagesRequest):
    """
    Scans an S3 prefix for .png files whose names follow the pattern
    unit{N}_{slug}.png (e.g. unit10_direct_clitic_pronouns.png), inserts a row
    into the UnitImages table for each file (url = public S3 URL, unit_number = N).

    Skips a file if a UnitImages row for that unit_number already exists,
    unless force=True (in which case a new row is always inserted).
    """
    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)

    # List all .png keys under the prefix
    prefix = body.s3_prefix.rstrip("/") + "/"
    png_keys: list[str] = []
    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=body.s3_bucket, Prefix=prefix):
        for obj in page.get("Contents", []):
            key = obj["Key"]
            if key.lower().endswith(".png"):
                png_keys.append(key)
    png_keys.sort()

    results = []
    skipped = []
    errors = []

    try:
        # --- Lesson-IDs path: link existing UnitImages to specific lessons ---
        if body.lesson_ids:
            for lesson_id in body.lesson_ids:
                log: dict = {
                    "lesson_id": lesson_id,
                    "unit_number": None,
                    "status": None,
                    "listening_questions_linked": 0,
                    "reading_lessons_linked": 0,
                }
                try:
                    with conn.cursor() as cur:
                        # 1. Resolve lesson → unit number
                        cur.execute(
                            "SELECT title, type FROM Lesson WHERE lesson_id = %s",
                            (lesson_id,),
                        )
                        lesson_row = cur.fetchone()
                        if not lesson_row:
                            log["status"] = "error"
                            log["error"] = f"Lesson {lesson_id} not found"
                            errors.append(log)
                            continue

                        unit_number = _extract_unit_num_from_title(lesson_row["title"])
                        if unit_number is None:
                            log["status"] = "error"
                            log["error"] = f"Cannot parse unit number from title '{lesson_row['title']}'"
                            errors.append(log)
                            continue

                        log["unit_number"] = unit_number

                        # 2. Find the existing UnitImages entry → get image_url via Image
                        cur.execute("""
                            SELECT i.image_url
                            FROM UnitImages ui
                            JOIN Image i ON i.image_id = ui.image_id
                            WHERE ui.unit_number = %s
                            LIMIT 1
                        """, (unit_number,))
                        ui_row = cur.fetchone()
                        if not ui_row:
                            log["status"] = "skipped"
                            log["reason"] = f"No UnitImages entry found for unit {unit_number}"
                            skipped.append(log)
                            continue

                        image_url = ui_row["image_url"]
                        lesson_type = lesson_row["type"]
                        listening_linked = 0
                        reading_linked = 0

                        # 3a. Listening: one Image row per unlinked question in this lesson
                        if lesson_type == "listening":
                            cur.execute("""
                                SELECT q.question_id
                                FROM Question q
                                LEFT JOIN Image i ON i.question_id = q.question_id
                                WHERE q.lesson_id = %s
                                  AND i.image_id IS NULL
                            """, (lesson_id,))
                            for q in cur.fetchall():
                                cur.execute(
                                    "INSERT INTO Image (lesson_id, question_id, image_url) VALUES (%s, %s, %s)",
                                    (lesson_id, q["question_id"], image_url),
                                )
                                listening_linked += 1

                        # 3b. Reading: one Image + LessonImages row if lesson not already linked
                        elif lesson_type == "reading" and "assessment" not in lesson_row["title"].lower():
                            cur.execute(
                                "SELECT image_id FROM LessonImages WHERE lesson_id = %s LIMIT 1",
                                (lesson_id,),
                            )
                            if not cur.fetchone():
                                cur.execute(
                                    "SELECT MIN(question_id) AS first_question_id FROM Question WHERE lesson_id = %s",
                                    (lesson_id,),
                                )
                                fq = cur.fetchone()
                                if fq and fq["first_question_id"]:
                                    cur.execute(
                                        "INSERT INTO Image (lesson_id, question_id, image_url) VALUES (%s, %s, %s)",
                                        (lesson_id, fq["first_question_id"], image_url),
                                    )
                                    img_id = cur.lastrowid
                                    cur.execute(
                                        "INSERT INTO LessonImages (lesson_id, image_id) VALUES (%s, %s)",
                                        (lesson_id, img_id),
                                    )
                                    reading_linked += 1

                    conn.commit()
                    log["status"] = "success"
                    log["listening_questions_linked"] = listening_linked
                    log["reading_lessons_linked"] = reading_linked
                    results.append(log)

                except Exception as e:
                    conn.rollback()
                    log["status"] = "error"
                    log["error"] = str(e)
                    errors.append(log)

            return JSONResponse(content={
                "status": "ok",
                "results": results,
                "skipped": skipped,
                "errors": errors,
                "summary": {
                    "total_lessons": len(body.lesson_ids),
                    "processed": len(results),
                    "skipped": len(skipped),
                    "errors": len(errors),
                    "total_listening_questions_linked": sum(r.get("listening_questions_linked", 0) for r in results),
                    "total_reading_lessons_linked": sum(r.get("reading_lessons_linked", 0) for r in results),
                },
            })

        # --- S3 scan path: process all PNG files under the prefix ---
        for key in png_keys:
            filename = Path(key).name
            unit_number = _extract_unit_num_from_title(filename)
            if unit_number is None:
                errors.append({"key": key, "error": f"Cannot parse unit number from filename '{filename}'"})
                continue

            image_url = f"https://{body.s3_bucket}.s3.{body.aws_region}.amazonaws.com/{key}"
            unit_like = f"unit{unit_number}\\_%"

            try:
                with conn.cursor() as cur:
                    # Skip entire unit if UnitImages already has an entry and force=False
                    if not body.force:
                        cur.execute(
                            "SELECT image_id FROM UnitImages WHERE unit_number = %s",
                            (unit_number,),
                        )
                        if cur.fetchone():
                            skipped.append({"key": key, "unit_number": unit_number, "reason": "UnitImages row already exists"})
                            continue

                    first_image_id = None
                    listening_linked = 0
                    reading_linked = 0

                    # --- Listening: one Image row per unlinked question ---
                    cur.execute("""
                        SELECT q.question_id, q.lesson_id
                        FROM Question q
                        JOIN Lesson l ON l.lesson_id = q.lesson_id
                        LEFT JOIN Image i ON i.question_id = q.question_id
                        WHERE l.type = 'listening'
                          AND l.title LIKE %s
                          AND i.image_id IS NULL
                    """, (unit_like,))
                    listening_questions = cur.fetchall()

                    for q in listening_questions:
                        cur.execute(
                            "INSERT INTO Image (lesson_id, question_id, image_url) VALUES (%s, %s, %s)",
                            (q["lesson_id"], q["question_id"], image_url),
                        )
                        img_id = cur.lastrowid
                        if first_image_id is None:
                            first_image_id = img_id
                        listening_linked += 1

                    # --- Reading: one Image + LessonImages row per unlinked lesson ---
                    # Fetch the first question_id per lesson to satisfy the NOT NULL FK on Image.question_id
                    cur.execute("""
                        SELECT l.lesson_id, MIN(q.question_id) AS first_question_id
                        FROM Lesson l
                        JOIN Question q ON q.lesson_id = l.lesson_id
                        LEFT JOIN LessonImages li ON li.lesson_id = l.lesson_id
                        WHERE l.type = 'reading'
                          AND l.title LIKE %s
                          AND l.title NOT LIKE '%%assessment%%'
                          AND li.image_id IS NULL
                        GROUP BY l.lesson_id
                    """, (unit_like,))
                    reading_lessons = cur.fetchall()

                    for r in reading_lessons:
                        cur.execute(
                            "INSERT INTO Image (lesson_id, question_id, image_url) VALUES (%s, %s, %s)",
                            (r["lesson_id"], r["first_question_id"], image_url),
                        )
                        img_id = cur.lastrowid
                        if first_image_id is None:
                            first_image_id = img_id
                        cur.execute(
                            "INSERT INTO LessonImages (lesson_id, image_id) VALUES (%s, %s)",
                            (r["lesson_id"], img_id),
                        )
                        reading_linked += 1

                    # --- UnitImages: one canonical row per unit ---
                    # If no new Image rows were created (everything already linked),
                    # fall back to any existing Image already tied to this unit's lessons.
                    if first_image_id is None:
                        cur.execute("""
                            SELECT i.image_id
                            FROM Image i
                            JOIN Lesson l ON l.lesson_id = i.lesson_id
                            WHERE l.title LIKE %s
                            LIMIT 1
                        """, (unit_like,))
                        row = cur.fetchone()
                        if row:
                            first_image_id = row["image_id"]

                    if first_image_id is not None:
                        cur.execute(
                            "INSERT INTO UnitImages (image_id, unit_number) VALUES (%s, %s)",
                            (first_image_id, unit_number),
                        )

                conn.commit()
                results.append({
                    "key": key,
                    "unit_number": unit_number,
                    "unit_image_id": first_image_id,
                    "listening_questions_linked": listening_linked,
                    "reading_lessons_linked": reading_linked,
                    "url": image_url,
                })

            except Exception as e:
                conn.rollback()
                errors.append({"key": key, "unit_number": unit_number, "error": str(e)})

    finally:
        conn.close()

    return JSONResponse(content={
        "status": "ok",
        "results": results,
        "skipped": skipped,
        "errors": errors,
        "summary": {
            "total_files": len(png_keys),
            "processed": len(results),
            "skipped": len(skipped),
            "errors": len(errors),
            "total_listening_questions_linked": sum(r["listening_questions_linked"] for r in results),
            "total_reading_lessons_linked": sum(r["reading_lessons_linked"] for r in results),
        },
    })


@app.post("/convert-to-listening")
def convert_to_listening(body: ConvertToListeningRequest):
    """
    For each lesson_id in the request:
      - Skips if the lesson is already type='listening'.
      - Skips if the lesson has no articles.
      - Runs GrammarScriptToAudio on each article (rewrites content → instructional
        script → MP3 bytes), uploads the MP3 to S3, inserts an Audio row, then
        flips Lesson.type to 'listening'.
    All articles for a lesson are committed together; if any article fails the
    lesson type is NOT changed and the error is reported.
    """
    generator = GrammarScriptToAudio(api_key=body.openai_api_key)

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)
    results = []
    total_succeeded = total_failed = total_skipped = 0

    try:
        for lesson_id in body.lesson_ids:
            log: dict = {
                "lesson_id": lesson_id,
                "status": None,
                "original_type": None,
                "articles_combined": 0,
                "audio_urls": [],
                "error": None,
            }

            try:
                with conn.cursor() as cur:
                    # 1. Fetch the lesson
                    cur.execute(
                        "SELECT lesson_id, title, type FROM Lesson WHERE lesson_id = %s",
                        (lesson_id,),
                    )
                    lesson = cur.fetchone()

                if not lesson:
                    log["status"] = "skipped"
                    log["error"] = f"Lesson {lesson_id} not found"
                    total_skipped += 1
                    results.append(log)
                    continue

                log["original_type"] = lesson["type"]

                if lesson["type"] == "listening":
                    log["status"] = "skipped"
                    log["error"] = "Already type='listening'"
                    total_skipped += 1
                    results.append(log)
                    continue

                # 2. Fetch all articles for this lesson
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT article_id, sequence_id, content "
                        "FROM Article WHERE lesson_id = %s ORDER BY sequence_id",
                        (lesson_id,),
                    )
                    articles = cur.fetchall()

                if not articles:
                    log["status"] = "skipped"
                    log["error"] = "No articles found"
                    total_skipped += 1
                    results.append(log)
                    continue

                # 3. Combine all articles; generate audio only if not already on S3
                lesson_title = lesson["title"]
                combined_content = "\n\n".join(a["content"] for a in articles)
                first_article_id = articles[0]["article_id"]

                s3_key = (
                    f"{body.s3_prefix.rstrip('/')}/"
                    f"{lesson_id}_{_safe_slug(lesson_title)}.mp3"
                )
                audio_url = (
                    f"https://{body.s3_bucket}.s3.{body.aws_region}.amazonaws.com/{s3_key}"
                )

                if not _s3_key_exists(s3, body.s3_bucket, s3_key):
                    script = generator.generate_script(combined_content)
                    audio_bytes = generator.generate_audio_bytes(script)
                    _upload_to_s3_public(
                        audio_bytes, s3_key, body.s3_bucket, body.aws_region,
                        body.aws_access_key_id, body.aws_secret_access_key,
                        content_type="audio/mpeg",
                    )

                # Insert Audio row; fall back gracefully if article_id column is absent
                metadata = json.dumps({
                    "source": "tts",
                    "tts_model": GrammarScriptToAudio.TTS_MODEL,
                    "voice": GrammarScriptToAudio.TTS_VOICE,
                })
                try:
                    _insert_audio_row(conn, lesson_id, first_article_id, audio_url, metadata, sequence_id=1)
                except pymysql.err.OperationalError as e:
                    if e.args[0] == 1054:  # Unknown column — minimal schema (lesson_id, sequence_id, audio_url only)
                        with conn.cursor() as cur:
                            cur.execute(
                                "INSERT INTO Audio (lesson_id, sequence_id, audio_url) VALUES (%s, %s, %s)",
                                (lesson_id, 1, audio_url),
                            )
                    else:
                        raise

                audio_urls = [audio_url]

                # 4. Flip lesson type to 'listening'
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE Lesson SET type = 'listening' WHERE lesson_id = %s",
                        (lesson_id,),
                    )

                conn.commit()

                log["status"] = "success"
                log["articles_combined"] = len(articles)
                log["audio_urls"] = audio_urls
                total_succeeded += 1

            except Exception as e:
                conn.rollback()
                log["status"] = "failed"
                log["error"] = str(e)
                total_failed += 1

            results.append(log)

    finally:
        conn.close()

    all_ok = total_failed == 0
    return JSONResponse(
        content={
            "status": "ok" if all_ok else "partial",
            "summary": {
                "total": len(body.lesson_ids),
                "succeeded": total_succeeded,
                "skipped": total_skipped,
                "failed": total_failed,
            },
            "results": results,
        },
        status_code=200 if all_ok else 207,
    )


# --- Writing-lesson helpers ---

def _build_writing_messages(unit_number: int, cefr_level: str,
                             context: str, num_questions: int) -> list[dict]:
    cefr_guidance = {
        "A1": "very simple and direct. Ask the learner to write 1-2 sentences about basic, concrete topics using vocabulary from the unit.",
        "A2": "simple and guided. Ask the learner to write 3-5 sentences about familiar topics, using structures seen in the unit.",
        "B1": "moderately open-ended. Ask the learner to write a short paragraph (5-8 sentences) describing, narrating, or giving an opinion on a topic from the unit.",
        "B2": "open-ended and analytical. Ask the learner to write a paragraph or short essay (8-12 sentences) that argues a point, compares ideas, or reflects on a theme from the unit.",
        "C1": "sophisticated and discursive. Ask the learner to produce a well-structured piece (one or more paragraphs) with nuanced vocabulary and complex grammar structures related to unit themes.",
        "C2": "highly open-ended and creative. Ask the learner to write an extended, polished piece that demonstrates mastery — e.g. a critical essay, creative narrative, or formal argument based on unit themes.",
    }.get(cefr_level.upper(), "appropriate to the learner's level.")

    system = (
        "You are an expert language-learning curriculum designer specializing in writing tasks. "
        "Your writing prompts should be engaging, pedagogically sound, and grounded in the unit content provided."
    )
    user = (
        f"UNIT: {unit_number}\n"
        f"CEFR LEVEL: {cefr_level}\n\n"
        f"UNIT CONTENT (articles, questions, and answers from existing lessons):\n"
        f"--------------------------------\n"
        f"{context}\n"
        f"--------------------------------\n\n"
        f"TASK:\n"
        f"- Create {num_questions} writing prompts for a {cefr_level} learner studying this unit.\n"
        f"- Each prompt should be {cefr_guidance}\n"
        f"- Prompts must be firmly grounded in the vocabulary, grammar, and themes of the unit content above.\n"
        f"- Do NOT include sample answers, model responses, or any evaluation criteria.\n"
        f"- Write prompts in ENGLISH.\n\n"
        f"OUTPUT FORMAT:\n"
        f"Return ONLY a JSON array (no surrounding prose, no code fences) where each element has exactly:\n"
        f"  - \"question\" : string  (the writing prompt shown to the learner)\n\n"
        f"IMPORTANT: valid JSON only, exactly {num_questions} items."
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


def _validate_writing_list(data: list, num_questions: int) -> None:
    if not isinstance(data, list):
        raise ValueError("Model did not return a JSON array.")
    if len(data) != num_questions:
        raise ValueError(f"Expected {num_questions} items; got {len(data)}.")
    for i, item in enumerate(data):
        if "question" not in item:
            raise ValueError(f"Item {i} missing key: 'question'.")


@app.post("/replace-speaking-articles")
def replace_speaking_articles(body: ReplaceSpeakingArticlesRequest):
    """
    For each JSON in the folder (e.g. unit19_model_conversation_cleaned.json):
      1. Extracts the unit number from the filename.
      2. Finds the single speaking lesson for that unit in the Lesson table.
      3. Deletes all existing Article rows for that lesson.
      4. Inserts new Article rows from the JSON's 'articles' array.
    """
    logger.info("replace-speaking-articles started | folder=%s limit=%s db=%s",
                body.folder, body.limit, body.db.database)

    folder = Path(body.folder)
    if not folder.is_dir():
        logger.error("Folder not found: %s", body.folder)
        raise HTTPException(status_code=400, detail=f"Folder not found: {body.folder}")

    json_files = sorted(folder.glob("*.json"))
    if not json_files:
        logger.error("No JSON files found in folder: %s", body.folder)
        raise HTTPException(status_code=400, detail="No JSON files found in folder")

    if body.limit is not None:
        json_files = json_files[: body.limit]

    logger.info("Files to process: %d", len(json_files))

    try:
        conn = connect_to_db(body.db)
        logger.info("Connected to database: %s", body.db.database)
    except Exception as e:
        logger.error("Could not connect to database: %s", e)
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    files_processed = []
    files_failed = []

    try:
        for json_file in json_files:
            logger.info("Processing file: %s", json_file.name)
            file_log = {
                "file": json_file.name,
                "status": None,
                "lesson_id": None,
                "articles_deleted": 0,
                "articles_inserted": 0,
                "errors": [],
            }

            # Extract unit number from filename (e.g. unit19_model_conversation_cleaned.json → 19)
            m = re.match(r"^unit(\d+)", json_file.name, re.IGNORECASE)
            if not m:
                msg = "Could not extract unit number from filename"
                logger.warning("[%s] %s", json_file.name, msg)
                file_log["status"] = "failed"
                file_log["errors"].append(msg)
                files_failed.append(file_log)
                continue
            unit_num = int(m.group(1))
            logger.info("[%s] Extracted unit number: %d", json_file.name, unit_num)

            # Parse JSON
            try:
                data = json.loads(json_file.read_text(encoding="utf-8-sig"))
                logger.info("[%s] JSON parsed successfully", json_file.name)
            except Exception as e:
                logger.warning("[%s] JSON parse error: %s", json_file.name, e)
                file_log["status"] = "failed"
                file_log["errors"].append(f"JSON parse error: {e}")
                files_failed.append(file_log)
                continue

            articles = data.get("articles", [])
            if not articles:
                msg = "No articles found in JSON"
                logger.warning("[%s] %s", json_file.name, msg)
                file_log["status"] = "failed"
                file_log["errors"].append(msg)
                files_failed.append(file_log)
                continue
            logger.info("[%s] Articles in JSON: %d", json_file.name, len(articles))

            try:
                with conn.cursor() as cur:
                    # Find the speaking lesson for this unit
                    cur.execute(
                        "SELECT lesson_id FROM Lesson "
                        "WHERE title LIKE %s AND type = 'speaking' LIMIT 1",
                        (f"unit{unit_num}\\_%",),
                    )
                    row = cur.fetchone()
                    if not row:
                        msg = f"No speaking lesson found for unit {unit_num}"
                        logger.warning("[%s] %s", json_file.name, msg)
                        file_log["status"] = "failed"
                        file_log["errors"].append(msg)
                        files_failed.append(file_log)
                        continue
                    lesson_id = row["lesson_id"]
                    file_log["lesson_id"] = lesson_id
                    logger.info("[%s] Found speaking lesson: lesson_id=%d", json_file.name, lesson_id)

                    # Delete existing articles
                    cur.execute("DELETE FROM Article WHERE lesson_id = %s", (lesson_id,))
                    file_log["articles_deleted"] = cur.rowcount
                    logger.info("[%s] Deleted %d existing articles for lesson_id=%d",
                                json_file.name, file_log["articles_deleted"], lesson_id)

                    # Insert new articles
                    for article in articles:
                        cur.execute(
                            "INSERT INTO Article (lesson_id, sequence_id, content) "
                            "VALUES (%s, %s, %s)",
                            (lesson_id, article["sequence_id"], article["text"]),
                        )
                        file_log["articles_inserted"] += 1
                    logger.info("[%s] Inserted %d articles for lesson_id=%d",
                                json_file.name, file_log["articles_inserted"], lesson_id)

                conn.commit()
                file_log["status"] = "success"
                files_processed.append(file_log)
                logger.info("[%s] Committed successfully", json_file.name)

            except Exception as e:
                conn.rollback()
                logger.error("[%s] DB error, rolling back: %s", json_file.name, e)
                file_log["status"] = "failed"
                file_log["errors"].append(str(e))
                files_failed.append(file_log)

    finally:
        conn.close()
        logger.info("DB connection closed")

    all_ok = len(files_failed) == 0
    logger.info("replace-speaking-articles finished | succeeded=%d failed=%d",
                len(files_processed), len(files_failed))
    return JSONResponse(
        content={
            "status": "ok" if all_ok else "partial",
            "summary": {
                "files_found": len(json_files),
                "files_succeeded": len(files_processed),
                "files_failed": len(files_failed),
                "articles_inserted": sum(f["articles_inserted"] for f in files_processed),
            },
            "files": files_processed + files_failed,
        },
        status_code=200 if all_ok else 207,
    )


@app.post("/generate-writing-lessons")
def generate_writing_lessons(body: GenerateWritingLessonsRequest):
    """
    For each unit number:
      1. Fetches all existing lessons whose title starts with 'unit{num}_'.
      2. Collects all articles, questions, and answers from those lessons as LLM context.
      3. Determines the CEFR level (from cefr_mapping if provided, else from existing DB lessons).
      4. Calls the LLM to generate writing prompts calibrated to that CEFR level.
      5. Inserts a new Lesson (type='writing'), Questions (type='writing', has_answer=0),
         and a blank Answer row for each question.
    Skips units that already have a writing lesson unless force=True.
    """
    effective_api_key = body.openai_api_key or os.getenv("OPENAI_API_KEY")
    if not effective_api_key:
        raise HTTPException(status_code=400, detail="OpenAI API key is required.")

    openai_client = OpenAI(api_key=effective_api_key)

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    results = []
    total_succeeded = total_failed = total_skipped = 0

    try:
        for unit_num in body.units:
            log: dict = {
                "unit": unit_num,
                "status": None,
                "cefr_level": None,
                "lesson_id": None,
                "questions_inserted": 0,
                "error": None,
            }

            try:
                with conn.cursor() as cur:
                    # 1. Skip if a writing lesson already exists for this unit
                    if not body.force:
                        cur.execute(
                            "SELECT lesson_id FROM Lesson WHERE title LIKE %s AND type = 'writing' LIMIT 1",
                            (f"unit{unit_num}\\_%",),
                        )
                        if cur.fetchone():
                            log["status"] = "skipped_already_exists"
                            total_skipped += 1
                            results.append(log)
                            continue

                    # 2. Fetch all existing lessons for this unit
                    cur.execute(
                        "SELECT lesson_id, title, cefr_level FROM Lesson "
                        "WHERE title LIKE %s AND type != 'writing'",
                        (f"unit{unit_num}\\_%",),
                    )
                    lessons = cur.fetchall()

                if not lessons:
                    log["status"] = "skipped_no_lessons"
                    total_skipped += 1
                    results.append(log)
                    continue

                # 3. Determine CEFR level
                if body.cefr_mapping:
                    cefr = resolve_cefr(f"unit{unit_num}_", body.cefr_mapping, body.cefr_level)
                else:
                    cefr = lessons[0]["cefr_level"] or body.cefr_level
                log["cefr_level"] = cefr

                # 4. Build context from all lessons in the unit
                context_parts: list[str] = []
                with conn.cursor() as cur:
                    for lesson in lessons:
                        lid = lesson["lesson_id"]
                        section_parts = [f"=== Lesson: {lesson['title']} ==="]

                        # Articles
                        cur.execute(
                            "SELECT sequence_id, content FROM Article "
                            "WHERE lesson_id = %s ORDER BY sequence_id",
                            (lid,),
                        )
                        for art in cur.fetchall():
                            section_parts.append(f"[Article {art['sequence_id']}]\n{art['content']}")

                        # Questions + Answers
                        cur.execute(
                            "SELECT question_id, sequence_id, question_text FROM Question "
                            "WHERE lesson_id = %s ORDER BY sequence_id",
                            (lid,),
                        )
                        for q in cur.fetchall():
                            section_parts.append(f"Q{q['sequence_id']}: {q['question_text']}")
                            cur.execute(
                                "SELECT answer_text, is_correct FROM Answer "
                                "WHERE question_id = %s ORDER BY answer_id",
                                (q["question_id"],),
                            )
                            for ans in cur.fetchall():
                                marker = "[correct]" if ans["is_correct"] else "[wrong]"
                                section_parts.append(f"  {marker} {ans['answer_text']}")

                        context_parts.append("\n".join(section_parts))

                context = "\n\n".join(context_parts)

                # 5. Generate writing prompts via LLM
                messages = _build_writing_messages(unit_num, cefr, context, body.num_questions)
                resp = openai_client.chat.completions.create(
                    model=body.model,
                    messages=messages,
                    temperature=body.temperature,
                )
                raw = resp.choices[0].message.content or ""
                writing_items = _parse_mcq_json_response(raw)
                _validate_writing_list(writing_items, body.num_questions)

                # 6. Insert writing lesson, questions, and blank answers
                lesson_title = f"unit{unit_num}_writing_{cefr.lower()}"
                with conn.cursor() as cur:
                    cur.execute(
                        "INSERT INTO Lesson (title, type, cefr_level, has_question) "
                        "VALUES (%s, 'writing', %s, 1)",
                        (lesson_title, cefr),
                    )
                    new_lesson_id = cur.lastrowid
                    log["lesson_id"] = new_lesson_id

                    for seq, item in enumerate(writing_items, start=1):
                        cur.execute(
                            "INSERT INTO Question (lesson_id, sequence_id, question_text, type, has_answer) "
                            "VALUES (%s, %s, %s, 'short_answer', 0)",
                            (new_lesson_id, seq, item["question"]),
                        )
                        question_id = cur.lastrowid
                        cur.execute(
                            "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                            "VALUES (%s, %s, '', 0)",
                            (new_lesson_id, question_id),
                        )
                        log["questions_inserted"] += 1

                conn.commit()
                log["status"] = "success"
                total_succeeded += 1

            except Exception as e:
                conn.rollback()
                log["status"] = "failed"
                log["error"] = str(e)
                total_failed += 1

            results.append(log)

    finally:
        conn.close()

    all_ok = total_failed == 0
    return JSONResponse(
        content={
            "status": "ok" if all_ok else "partial",
            "summary": {
                "units_requested": len(body.units),
                "succeeded": total_succeeded,
                "skipped": total_skipped,
                "failed": total_failed,
            },
            "results": results,
        },
        status_code=200 if all_ok else 207,
    )


@app.post("/backfill-answer-text")
def backfill_answer_text(body: BackfillAnswerTextRequest):
    """
    For each JSON file in a folder (or a single file), extract the lesson_id from
    the filename prefix (e.g. '3251_h_unit16_...json' → 3251), then for every
    question in questions_and_answers find the matching Question row by lesson_id +
    question_text, fetch its Answer rows ordered by insertion id, match them
    positionally to the JSON answers, verify is_correct alignment, and UPDATE
    answer_text.
    """
    if not body.folder and not body.filename:
        raise HTTPException(status_code=400, detail="Provide at least one of 'folder' or 'filename'")

    json_files: list[Path] = []

    if body.folder:
        folder = Path(body.folder)
        if not folder.is_dir():
            raise HTTPException(status_code=400, detail=f"Folder not found: {body.folder}")
        json_files.extend(sorted(folder.glob("*.json")))

    if body.filename:
        p = Path(body.filename)
        if not p.is_file():
            raise HTTPException(status_code=400, detail=f"File not found: {body.filename}")
        if p not in json_files:
            json_files.append(p)

    if not json_files:
        raise HTTPException(status_code=400, detail="No JSON files found")

    if body.limit is not None:
        json_files = json_files[: body.limit]

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    files_processed = []
    files_failed = []

    try:
        for json_file in json_files:
            file_log = {
                "file": json_file.name,
                "status": None,
                "lesson_id": None,
                "answers_updated": 0,
                "errors": [],
            }

            # Extract lesson_id from filename prefix (digits before first '_')
            m = re.match(r"^(\d+)_", json_file.name)
            if not m:
                file_log["status"] = "failed"
                file_log["errors"].append("Could not extract lesson_id from filename")
                files_failed.append(file_log)
                continue
            lesson_id = int(m.group(1))
            file_log["lesson_id"] = lesson_id

            # Parse JSON
            try:
                data = json.loads(json_file.read_text(encoding="utf-8-sig"))
            except Exception as e:
                file_log["status"] = "failed"
                file_log["errors"].append(f"JSON parse error: {e}")
                files_failed.append(file_log)
                continue

            try:
                with conn.cursor() as cur:
                    for qa in data.get("questions_and_answers", []):
                        question_text = qa.get("question", "")
                        json_answers = qa.get("answers", [])
                        if not json_answers:
                            continue

                        # Find the Question row
                        cur.execute(
                            "SELECT question_id FROM Question WHERE lesson_id = %s AND question_text = %s LIMIT 1",
                            (lesson_id, question_text),
                        )
                        q_row = cur.fetchone()
                        if not q_row:
                            file_log["errors"].append(
                                f"Question not found: lesson_id={lesson_id} text={question_text!r}"
                            )
                            continue
                        question_id = q_row["question_id"]

                        # Fetch Answer rows ordered by answer_id (insertion order)
                        cur.execute(
                            "SELECT answer_id, is_correct FROM Answer WHERE question_id = %s ORDER BY answer_id ASC",
                            (question_id,),
                        )
                        db_answers = cur.fetchall()

                        if len(db_answers) != len(json_answers):
                            file_log["errors"].append(
                                f"question_id={question_id}: DB has {len(db_answers)} answers "
                                f"but JSON has {len(json_answers)}"
                            )
                            continue

                        # Match by is_correct: correct DB row → correct JSON text;
                        # incorrect DB rows paired positionally with incorrect JSON answers
                        db_correct = [a for a in db_answers if a["is_correct"]]
                        db_incorrect = [a for a in db_answers if not a["is_correct"]]
                        json_correct = [a for a in json_answers if a.get("is_correct")]
                        json_incorrect = [a for a in json_answers if not a.get("is_correct")]

                        if len(db_correct) != len(json_correct):
                            file_log["errors"].append(
                                f"question_id={question_id}: DB has {len(db_correct)} correct answer(s) "
                                f"but JSON has {len(json_correct)}"
                            )
                            continue

                        for db_ans, json_ans in list(zip(db_correct, json_correct)) + list(zip(db_incorrect, json_incorrect)):
                            cur.execute(
                                "UPDATE Answer SET answer_text = %s WHERE answer_id = %s",
                                (json_ans.get("text", ""), db_ans["answer_id"]),
                            )
                            file_log["answers_updated"] += 1

                conn.commit()
                file_log["status"] = "success"
                files_processed.append(file_log)

            except Exception as e:
                conn.rollback()
                file_log["status"] = "failed"
                file_log["errors"].append(str(e))
                files_failed.append(file_log)

    finally:
        conn.close()

    all_ok = len(files_failed) == 0
    return JSONResponse(
        content={
            "status": "ok" if all_ok else "partial",
            "summary": {
                "files_found": len(json_files),
                "files_succeeded": len(files_processed),
                "files_failed": len(files_failed),
                "answers_updated": sum(f["answers_updated"] for f in files_processed),
            },
            "files": files_processed + files_failed,
        },
        status_code=200 if all_ok else 207,
    )

# =============================================================================
#  Add to main.py  —  Rebalance grammar lessons: split each lesson's questions
#  ~50/50 into `practice` and `learning`, converting multiple_choice questions
#  into short_answer along the way.
#
#  Requires the imports already present in main.py (FastAPI, pymysql, logger,
#  connect_to_db, DBConfig, JSONResponse, BaseModel).
# =============================================================================


# ----------------------------------------------------------------------------- 
# Request model  (place near the other Pydantic models)
# ----------------------------------------------------------------------------- 
class RebalanceGrammarRequest(BaseModel):
    db: DBConfig
    lesson_ids: list[int] | None = None     # restrict to these grammar lessons; None = ALL grammar lessons
    learning_ratio: float = 0.5             # target fraction of each lesson that should be `learning`
    handle_orphan_practice: bool = True     # Step A: MC-type questions sitting at answer_category='practice' → learning
    dry_run: bool = True                    # True = run everything in a transaction, verify, then ROLLBACK
    # German rule (post-French): a question is multiple_choice when it has >1 real answer
    # AND at least one is_correct. After this route runs, no grammar question keeps >1 answer,
    # so grammar can only resolve to practice or learning.


# ----------------------------------------------------------------------------- 
# Conversion helpers
# ----------------------------------------------------------------------------- 
def _convert_question_to_practice(cur, question_id: int) -> int:
    """
    Mold an MC question into a `practice` short_answer:
      - keep exactly ONE answer row (the correct one — lowest answer_id among
        is_correct=1; if none correct, the lowest answer_id overall),
      - delete every other answer row,
      - mark the survivor is_correct=1,
      - type='short_answer', has_answer=1.
    Returns the number of answer rows deleted.
    """
    cur.execute(
        "SELECT answer_id, is_correct FROM Answer WHERE question_id = %s ORDER BY answer_id ASC",
        (question_id,),
    )
    rows = cur.fetchall()
    if not rows:
        # No answers at all — nothing to keep; flag as answered short_answer (edge case).
        cur.execute(
            "UPDATE Question SET type = 'short_answer', has_answer = 1 WHERE question_id = %s",
            (question_id,),
        )
        return 0

    keep_id = next((r["answer_id"] for r in rows if r["is_correct"]), rows[0]["answer_id"])
    cur.execute(
        "DELETE FROM Answer WHERE question_id = %s AND answer_id <> %s",
        (question_id, keep_id),
    )
    deleted = cur.rowcount
    cur.execute("UPDATE Answer SET is_correct = 1 WHERE answer_id = %s", (keep_id,))
    cur.execute(
        "UPDATE Question SET type = 'short_answer', has_answer = 1 WHERE question_id = %s",
        (question_id,),
    )
    return deleted


def _convert_question_to_learning(cur, question_id: int) -> int:
    """
    Mold an MC question into a `learning` short_answer:
      - keep ONLY the lowest answer_id row, delete the rest,
      - blank that survivor (answer_text='', is_correct=0),
      - type='short_answer', has_answer=0.
    Returns the number of answer rows deleted.
    """
    cur.execute(
        "SELECT answer_id FROM Answer WHERE question_id = %s ORDER BY answer_id ASC",
        (question_id,),
    )
    rows = cur.fetchall()
    deleted = 0
    if rows:
        keep_id = rows[0]["answer_id"]
        cur.execute(
            "DELETE FROM Answer WHERE question_id = %s AND answer_id <> %s",
            (question_id, keep_id),
        )
        deleted = cur.rowcount
        cur.execute(
            "UPDATE Answer SET answer_text = '', is_correct = 0 WHERE answer_id = %s",
            (keep_id,),
        )
    cur.execute(
        "UPDATE Question SET type = 'short_answer', has_answer = 0 WHERE question_id = %s",
        (question_id,),
    )
    return deleted


# answer_category derivation (German / post-French rule), scoped via {ids} placeholder
_GRAMMAR_BACKFILL_SQL = """
UPDATE Question q
JOIN (
    SELECT q2.question_id,
           l.type AS lesson_type,
           COUNT(CASE WHEN TRIM(a.answer_text) <> '' THEN a.answer_id END) AS real_answers,
           SUM(CASE WHEN a.is_correct = 1 THEN 1 ELSE 0 END)               AS correct_count
    FROM Question q2
    LEFT JOIN Lesson l ON l.lesson_id   = q2.lesson_id
    LEFT JOIN Answer a ON a.question_id = q2.question_id
    WHERE q2.lesson_id IN ({ids})
    GROUP BY q2.question_id, l.type
) agg ON agg.question_id = q.question_id
SET q.answer_category = CASE
    WHEN agg.real_answers > 1 AND agg.correct_count >= 1 THEN 'multiple_choice'
    WHEN agg.real_answers >= 1                            THEN 'practice'
    WHEN agg.real_answers = 0
         AND agg.lesson_type IN ('vocabulary','grammar')  THEN 'learning'
    WHEN agg.real_answers = 0
         AND agg.lesson_type IN ('reading','writing','speaking','listening') THEN 'open_ended'
    ELSE NULL
END
WHERE q.lesson_id IN ({ids})
"""


# ----------------------------------------------------------------------------- 
# Route
# ----------------------------------------------------------------------------- 
@app.post("/rebalance-grammar-categories")
def rebalance_grammar_categories(body: RebalanceGrammarRequest):
    """
    Rebalance grammar-lesson questions so each lesson is ~50/50 practice/learning.

    Pipeline (single transaction):
      Step A  — convert MC-type questions currently sitting at answer_category='practice'
                (the malformed-MC / no-correct-flag group) into `learning`.
      Step B  — per grammar lesson: target_learning = floor(total * learning_ratio).
                Questions already in a learning state (has_answer=0, incl. Step A output)
                COUNT toward that target. Convert just enough of the remaining
                multiple_choice questions to learning to reach the target; the rest
                become practice. learning_needed is clamped to >= 0 and to the MC pool
                size, so a lesson already over-quota on learning (e.g. 6 of 10 from
                Step A) simply makes its remaining MC questions practice.
      Backfill— re-derive answer_category from the new structure (proves correctness).
      Verify  — per-lesson distribution + invariants, then COMMIT or (dry_run) ROLLBACK.

    Safe: neither user_attempts nor userResponses references Answer.answer_id, so
    deleting answer rows cannot strand user-activity data.
    """
    logger.info(
        "rebalance-grammar started | db=%s lesson_ids=%s ratio=%s handle_orphan=%s dry_run=%s",
        body.db.database, body.lesson_ids, body.learning_ratio,
        body.handle_orphan_practice, body.dry_run,
    )

    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        logger.error("Could not connect to database: %s", e)
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    step_a_log: list[dict] = []
    lesson_results: list[dict] = []
    skipped_non_grammar: list[dict] = []

    try:
        with conn.cursor() as cur:
            # --- Resolve in-scope grammar lessons ---
            if body.lesson_ids:
                fmt = ",".join(["%s"] * len(body.lesson_ids))
                cur.execute(
                    f"SELECT lesson_id, title, type FROM Lesson WHERE lesson_id IN ({fmt})",
                    tuple(body.lesson_ids),
                )
                rows = cur.fetchall()
                grammar_lessons = [r for r in rows if r["type"] == "grammar"]
                skipped_non_grammar = [
                    {"lesson_id": r["lesson_id"], "type": r["type"]}
                    for r in rows if r["type"] != "grammar"
                ]
            else:
                cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE type = 'grammar'")
                grammar_lessons = cur.fetchall()

            lesson_id_list = [r["lesson_id"] for r in grammar_lessons]
            if not lesson_id_list:
                conn.close()
                return JSONResponse(
                    status_code=400,
                    content={"status": "error", "detail": "No in-scope grammar lessons found."},
                )

            fmt_ids = ",".join(["%s"] * len(lesson_id_list))
            logger.info("In-scope grammar lessons: %d", len(lesson_id_list))

            # =================================================================
            # STEP A — orphan-practice MC questions → learning
            # =================================================================
            step_a_deleted = 0
            if body.handle_orphan_practice:
                cur.execute(
                    f"""
                    SELECT q.question_id, q.lesson_id
                    FROM Question q
                    WHERE q.lesson_id IN ({fmt_ids})
                      AND q.type = 'multiple_choice'
                      AND q.answer_category = 'practice'
                    ORDER BY q.lesson_id, q.question_id
                    """,
                    tuple(lesson_id_list),
                )
                step_a_targets = cur.fetchall()
                logger.info("Step A targets (MC-type @ practice): %d", len(step_a_targets))
                for t in step_a_targets:
                    d = _convert_question_to_learning(cur, t["question_id"])
                    step_a_deleted += d
                    step_a_log.append({
                        "question_id": t["question_id"],
                        "lesson_id": t["lesson_id"],
                        "answers_deleted": d,
                    })

            # =================================================================
            # STEP B — per-lesson 50/50 split of remaining MC questions
            # =================================================================
            step_b_deleted = 0
            for lesson in grammar_lessons:
                lid = lesson["lesson_id"]
                cur.execute(
                    "SELECT question_id, type, has_answer FROM Question "
                    "WHERE lesson_id = %s ORDER BY question_id ASC",
                    (lid,),
                )
                qs = cur.fetchall()
                total = len(qs)
                if total == 0:
                    lesson_results.append({
                        "lesson_id": lid, "title": lesson["title"], "total": 0,
                        "status": "empty",
                    })
                    continue

                existing_learning = sum(1 for q in qs if q["has_answer"] == 0)
                mc_pool = [q for q in qs if q["type"] == "multiple_choice"]

                target_learning = int(total * body.learning_ratio)   # floor
                learning_needed = target_learning - existing_learning
                if learning_needed < 0:
                    learning_needed = 0
                if learning_needed > len(mc_pool):
                    learning_needed = len(mc_pool)

                to_learning = mc_pool[:learning_needed]
                to_practice = mc_pool[learning_needed:]

                for q in to_learning:
                    step_b_deleted += _convert_question_to_learning(cur, q["question_id"])
                for q in to_practice:
                    step_b_deleted += _convert_question_to_practice(cur, q["question_id"])

                final_learning = existing_learning + len(to_learning)
                final_practice = total - final_learning

                rec = {
                    "lesson_id": lid,
                    "title": lesson["title"],
                    "total": total,
                    "target_learning": target_learning,
                    "existing_learning_before_stepB": existing_learning,
                    "mc_pool": len(mc_pool),
                    "converted_to_learning": len(to_learning),
                    "converted_to_practice": len(to_practice),
                    "final_learning": final_learning,
                    "final_practice": final_practice,
                    "lopsided": final_learning > final_practice,  # over-quota learning (e.g. Step A heavy)
                }
                lesson_results.append(rec)
                logger.info(
                    "[lesson %d] total=%d target_L=%d existing_L=%d pool=%d -> L+%d / P+%d => %dL/%dP%s",
                    lid, total, target_learning, existing_learning, len(mc_pool),
                    len(to_learning), len(to_practice), final_learning, final_practice,
                    "  *LOPSIDED*" if rec["lopsided"] else "",
                )

            # =================================================================
            # RE-RUN BACKFILL (derive answer_category from new structure)
            # =================================================================
            backfill_sql = _GRAMMAR_BACKFILL_SQL.format(ids=fmt_ids)
            cur.execute(backfill_sql, tuple(lesson_id_list) + tuple(lesson_id_list))
            backfill_updated = cur.rowcount
            logger.info("Backfill updated rows: %d", backfill_updated)

            # =================================================================
            # VERIFICATION (computed on real post-change state, pre-commit)
            # =================================================================
            cur.execute(
                f"SELECT lesson_id, answer_category, COUNT(*) AS cnt "
                f"FROM Question WHERE lesson_id IN ({fmt_ids}) "
                f"GROUP BY lesson_id, answer_category ORDER BY lesson_id",
                tuple(lesson_id_list),
            )
            per_lesson_categories = cur.fetchall()

            cur.execute(
                f"SELECT answer_category, COUNT(*) AS cnt "
                f"FROM Question WHERE lesson_id IN ({fmt_ids}) GROUP BY answer_category",
                tuple(lesson_id_list),
            )
            overall_categories = {r["answer_category"]: r["cnt"] for r in cur.fetchall()}

            cur.execute(
                f"SELECT COUNT(*) AS c FROM Question "
                f"WHERE lesson_id IN ({fmt_ids}) AND type = 'multiple_choice'",
                tuple(lesson_id_list),
            )
            remaining_mc_type = cur.fetchone()["c"]

            cur.execute(
                f"SELECT COUNT(*) AS c FROM Question "
                f"WHERE lesson_id IN ({fmt_ids}) AND answer_category = 'multiple_choice'",
                tuple(lesson_id_list),
            )
            remaining_mc_category = cur.fetchone()["c"]

            cur.execute(
                f"SELECT COUNT(*) AS c FROM Question "
                f"WHERE lesson_id IN ({fmt_ids}) AND answer_category IS NULL",
                tuple(lesson_id_list),
            )
            null_category = cur.fetchone()["c"]

            # Invariant: every in-scope grammar question must end with exactly ONE answer row
            cur.execute(
                f"""
                SELECT COUNT(*) AS c FROM (
                    SELECT q.question_id, COUNT(a.answer_id) AS n
                    FROM Question q
                    LEFT JOIN Answer a ON a.question_id = q.question_id
                    WHERE q.lesson_id IN ({fmt_ids})
                    GROUP BY q.question_id
                    HAVING n <> 1
                ) t
                """,
                tuple(lesson_id_list),
            )
            questions_not_single_answer = cur.fetchone()["c"]

            verification = {
                "overall_categories": overall_categories,
                "remaining_multiple_choice_type": remaining_mc_type,        # expect 0
                "remaining_multiple_choice_category": remaining_mc_category,  # expect 0
                "questions_with_null_category": null_category,               # expect 0
                "questions_not_exactly_one_answer": questions_not_single_answer,  # expect 0
                "lopsided_lessons": [r["lesson_id"] for r in lesson_results if r.get("lopsided")],
            }

            checks_pass = (
                remaining_mc_type == 0
                and remaining_mc_category == 0
                and null_category == 0
                and questions_not_single_answer == 0
            )

            # =================================================================
            # COMMIT / ROLLBACK gate
            # =================================================================
            if body.dry_run:
                conn.rollback()
                committed = False
                logger.info("DRY RUN — rolled back. checks_pass=%s", checks_pass)
            else:
                if not checks_pass:
                    conn.rollback()
                    committed = False
                    logger.error("Verification FAILED — rolled back instead of committing. %s", verification)
                else:
                    conn.commit()
                    committed = True
                    logger.info("COMMITTED.")

    except Exception as e:
        conn.rollback()
        conn.close()
        logger.error("rebalance-grammar failed, rolled back: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Rebalance failed (rolled back): {e}")

    conn.close()

    return JSONResponse(
        status_code=200,
        content={
            "status": "ok" if checks_pass else "verification_failed",
            "committed": committed,
            "dry_run": body.dry_run,
            "summary": {
                "grammar_lessons_in_scope": len(lesson_id_list),
                "skipped_non_grammar": skipped_non_grammar,
                "step_a_converted": len(step_a_log),
                "step_a_answers_deleted": step_a_deleted,
                "step_b_answers_deleted": step_b_deleted,
                "backfill_rows_updated": backfill_updated,
            },
            "verification": verification,
            "lessons": lesson_results,
            "step_a": step_a_log,
        },
    )


# =============================================================================
#  Export a QA workbook (.xlsx) for one language DB.
#
#  One workbook per language. One worksheet per unit (unit1, unit2, ...).
#  Within a unit tab, content is laid out top-down:
#     UNIT heading
#       TYPE heading (vocabulary, grammar, reading, writing, listening, speaking)
#         LESSON heading (Lesson.title, plus per-lesson url(s) where relevant)
#           per-type column block + rows
#
#  Requires openpyxl (pip install openpyxl).  Reuses connect_to_db / DBConfig /
#  logger / JSONResponse / HTTPException / json / re / Path already in main.py.
# =============================================================================

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _OPENPYXL_AVAILABLE = True
except ImportError:  # pragma: no cover
    Workbook = Font = PatternFill = Alignment = None  # type: ignore[assignment]
    _OPENPYXL_AVAILABLE = False

# Fixed display order of type sections within a unit
_TYPE_ORDER = ["vocabulary", "grammar", "reading", "writing", "listening", "speaking"]

# Styling (only available when openpyxl is installed)
if _OPENPYXL_AVAILABLE:
    _UNIT_FILL   = PatternFill("solid", fgColor="1F4E78")  # dark blue
    _TYPE_FILL   = PatternFill("solid", fgColor="2E75B6")  # medium blue
    _LESSON_FILL = PatternFill("solid", fgColor="BDD7EE")  # light blue
    _HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")  # grey column headers
    _WHITE_BOLD  = Font(name="Arial", bold=True, color="FFFFFF")
    _DARK_BOLD   = Font(name="Arial", bold=True, color="000000")
    _NORMAL      = Font(name="Arial")
    _WRAP_TOP    = Alignment(wrap_text=True, vertical="top")
    _WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
else:
    _UNIT_FILL = _TYPE_FILL = _LESSON_FILL = _HEADER_FILL = None
    _WHITE_BOLD = _DARK_BOLD = _NORMAL = None
    _WRAP_TOP = _WRAP_CENTER = None


class ExportQAWorkbookRequest(BaseModel):
    db: DBConfig
    language_label: str                 # used for the filename, e.g. "spanish"
    units: list[int] | None = None      # specific unit numbers; None = all units found
    output_dir: str = "/tmp"            # where the .xlsx is written


def _ids_json(lesson_id=None, question_id=None, answer_id=None, article_id=None) -> str:
    return json.dumps({
        "lesson_id": lesson_id,
        "question_id": question_id,
        "answer_id": answer_id,
        "article_id": article_id,
    })


def _unit_num(title: str):
    m = re.match(r"^unit(\d+)", title or "", re.IGNORECASE)
    return int(m.group(1)) if m else None


def _fetch_articles(cur, lesson_id):
    cur.execute(
        "SELECT article_id, sequence_id, content FROM Article "
        "WHERE lesson_id = %s ORDER BY sequence_id ASC, article_id ASC",
        (lesson_id,),
    )
    return cur.fetchall()


def _fetch_questions(cur, lesson_id):
    cur.execute(
        "SELECT question_id, sequence_id, question_text, answer_category "
        "FROM Question WHERE lesson_id = %s ORDER BY sequence_id ASC, question_id ASC",
        (lesson_id,),
    )
    return cur.fetchall()


def _fetch_answers(cur, question_id):
    cur.execute(
        "SELECT answer_id, answer_text, is_correct FROM Answer "
        "WHERE question_id = %s ORDER BY answer_id ASC",
        (question_id,),
    )
    return cur.fetchall()


def _question_image_url(cur, question_id):
    cur.execute(
        "SELECT image_url FROM Image WHERE question_id = %s ORDER BY image_id ASC LIMIT 1",
        (question_id,),
    )
    r = cur.fetchone()
    return r["image_url"] if r else ""


def _question_audio_url(cur, question_id):
    cur.execute(
        "SELECT audio_url FROM Audio WHERE question_id = %s ORDER BY audio_id ASC LIMIT 1",
        (question_id,),
    )
    r = cur.fetchone()
    return r["audio_url"] if r else ""


def _lesson_audio_url(cur, lesson_id):
    cur.execute(
        "SELECT audio_url FROM Audio WHERE lesson_id = %s ORDER BY audio_id ASC LIMIT 1",
        (lesson_id,),
    )
    r = cur.fetchone()
    return r["audio_url"] if r else ""


def _lesson_image_url(cur, lesson_id):
    cur.execute(
        "SELECT image_url FROM Image WHERE lesson_id = %s ORDER BY image_id ASC LIMIT 1",
        (lesson_id,),
    )
    r = cur.fetchone()
    return r["image_url"] if r else ""


def _lesson_video_url(cur, lesson_id):
    cur.execute(
        "SELECT v.url FROM Video v "
        "JOIN VideoLesson vl ON vl.video_id = v.video_id "
        "WHERE vl.lesson_id = %s ORDER BY v.video_id ASC LIMIT 1",
        (lesson_id,),
    )
    r = cur.fetchone()
    return r["url"] if r else ""


def _correct_index(answers):
    """1-based positions (by answer_id ascending) of is_correct=1 rows; '' if none."""
    idxs = [str(i) for i, a in enumerate(answers, start=1) if a["is_correct"]]
    return ",".join(idxs)


def _write_row(ws, row, values, font=None, fill=None, wrap=False, center_cols=None):
    center_cols = center_cols or set()
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font if font else _NORMAL
        if fill:
            c.fill = fill
        if col in center_cols:
            c.alignment = _WRAP_CENTER
        elif wrap:
            c.alignment = _WRAP_TOP
    return row + 1


def _heading(ws, row, text, fill, font, span=8):
    c = ws.cell(row=row, column=1, value=text)
    c.font = font
    c.fill = fill
    for col in range(2, span + 1):
        ws.cell(row=row, column=col).fill = fill
    return row + 1


# --- per-type block writers; each returns the next free row -------------------

def _block_vocabulary(ws, row, cur, lesson):
    lid = lesson["lesson_id"]
    row = _write_row(ws, row,
        ["answer_category", "sequence_id", "question_text", "answer_text",
         "image_url", "audio_url", "ids"],
        font=_DARK_BOLD, fill=_HEADER_FILL, center_cols={2})
    for q in _fetch_questions(cur, lid):
        answers = _fetch_answers(cur, q["question_id"])
        ans = answers[0] if answers else None
        row = _write_row(ws, row, [
            q["answer_category"], q["sequence_id"], q["question_text"],
            ans["answer_text"] if ans else "",
            _question_image_url(cur, q["question_id"]),
            _question_audio_url(cur, q["question_id"]),
            _ids_json(lid, q["question_id"], ans["answer_id"] if ans else None, None),
        ], wrap=True, center_cols={2})
    return row


def _block_writing(ws, row, cur, lesson):
    lid = lesson["lesson_id"]
    row = _write_row(ws, row,
        ["answer_category", "sequence_id", "question_text", "answer_text", "ids"],
        font=_DARK_BOLD, fill=_HEADER_FILL, center_cols={2})
    for q in _fetch_questions(cur, lid):
        answers = _fetch_answers(cur, q["question_id"])
        ans = answers[0] if answers else None
        row = _write_row(ws, row, [
            q["answer_category"], q["sequence_id"], q["question_text"],
            ans["answer_text"] if ans else "",
            _ids_json(lid, q["question_id"], ans["answer_id"] if ans else None, None),
        ], wrap=True, center_cols={2})
    return row


def _block_speaking(ws, row, cur, lesson):
    lid = lesson["lesson_id"]
    row = _write_row(ws, row, ["article_id", "article_text"],
                     font=_DARK_BOLD, fill=_HEADER_FILL)
    for a in _fetch_articles(cur, lid):
        row = _write_row(ws, row, [a["article_id"], a["content"]], wrap=True)
    return row


def _block_reading(ws, row, cur, lesson):
    """Article block | gap | MCQ block. image_url shown once at lesson heading."""
    lid = lesson["lesson_id"]
    # Article block + gap col + Q&A header in one header row
    row = _write_row(ws, row, [
        "article_id", "article_text", "",
        "answer_category", "sequence_id", "question_text",
        "Answer 1", "Answer 2", "Answer 3", "Answer 4", "Correct Answer", "ids",
    ], font=_DARK_BOLD, fill=_HEADER_FILL, center_cols={5})

    articles = _fetch_articles(cur, lid)
    questions = _fetch_questions(cur, lid)
    n = max(len(articles), len(questions))
    for i in range(n):
        art = articles[i] if i < len(articles) else None
        q = questions[i] if i < len(questions) else None
        left = [art["article_id"], art["content"]] if art else ["", ""]
        if q:
            answers = _fetch_answers(cur, q["question_id"])
            opts = [answers[j]["answer_text"] if j < len(answers) else "" for j in range(4)]
            correct_ans_id = next((a["answer_id"] for a in answers if a["is_correct"]), None)
            right = [q["answer_category"], q["sequence_id"], q["question_text"],
                     *opts, _correct_index(answers),
                     _ids_json(lid, q["question_id"], correct_ans_id,
                               art["article_id"] if art else None)]
        else:
            right = ["", "", "", "", "", "", "", "",
                     _ids_json(lid, None, None, art["article_id"] if art else None)]
        row = _write_row(ws, row, [*left, "", *right], wrap=True, center_cols={5})
    return row


def _block_listening(ws, row, cur, lesson):
    """One image_url + audio_url per lesson (shown at heading); MCQ rows below."""
    lid = lesson["lesson_id"]
    row = _write_row(ws, row, [
        "answer_category", "sequence_id", "question_text",
        "Answer 1", "Answer 2", "Answer 3", "Answer 4", "Correct Answer", "ids",
    ], font=_DARK_BOLD, fill=_HEADER_FILL, center_cols={2})
    for q in _fetch_questions(cur, lid):
        answers = _fetch_answers(cur, q["question_id"])
        opts = [answers[j]["answer_text"] if j < len(answers) else "" for j in range(4)]
        correct_ans_id = next((a["answer_id"] for a in answers if a["is_correct"]), None)
        row = _write_row(ws, row, [
            q["answer_category"], q["sequence_id"], q["question_text"],
            *opts, _correct_index(answers),
            _ids_json(lid, q["question_id"], correct_ans_id, None),
        ], wrap=True, center_cols={2})
    return row


def _block_grammar(ws, row, cur, lesson):
    """One video_url per lesson (shown at heading). Article block | gap | Q&A (short_answer)."""
    lid = lesson["lesson_id"]
    row = _write_row(ws, row, [
        "article_id", "article_text", "",
        "answer_category", "sequence_id", "question_text", "answer_text", "ids",
    ], font=_DARK_BOLD, fill=_HEADER_FILL, center_cols={5})

    articles = _fetch_articles(cur, lid)
    questions = _fetch_questions(cur, lid)
    n = max(len(articles), len(questions))
    for i in range(n):
        art = articles[i] if i < len(articles) else None
        q = questions[i] if i < len(questions) else None
        left = [art["article_id"], art["content"]] if art else ["", ""]
        if q:
            answers = _fetch_answers(cur, q["question_id"])
            ans = answers[0] if answers else None
            right = [q["answer_category"], q["sequence_id"], q["question_text"],
                     ans["answer_text"] if ans else "",
                     _ids_json(lid, q["question_id"], ans["answer_id"] if ans else None,
                               art["article_id"] if art else None)]
        else:
            right = ["", "", "", "",
                     _ids_json(lid, None, None, art["article_id"] if art else None)]
        row = _write_row(ws, row, [*left, "", *right], wrap=True, center_cols={5})
    return row


_BLOCK_WRITERS = {
    "vocabulary": _block_vocabulary,
    "writing": _block_writing,
    "speaking": _block_speaking,
    "reading": _block_reading,
    "listening": _block_listening,
    "grammar": _block_grammar,
}


@app.post("/export-qa-workbook")
def export_qa_workbook(body: ExportQAWorkbookRequest):
    """
    Build one .xlsx for a language DB: a tab per unit, per-type lesson blocks.
    Pass `units` to export a single unit (e.g. [1]) or a subset; omit for all.
    """
    logger.info("export-qa-workbook started | db=%s label=%s units=%s",
                body.db.database, body.language_label, body.units)
    if not _OPENPYXL_AVAILABLE:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required for /export-qa-workbook (pip install openpyxl).",
        )
    try:
        conn = connect_to_db(body.db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not connect to database: {e}")

    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT lesson_id, title, type FROM Lesson "
                "WHERE title REGEXP '^unit[0-9]' ORDER BY title ASC"
            )
            lessons = cur.fetchall()

            # Group lessons by unit number
            units: dict[int, list[dict]] = {}
            for l in lessons:
                u = _unit_num(l["title"])
                if u is None:
                    continue
                if body.units is not None and u not in body.units:
                    continue
                units.setdefault(u, []).append(l)

            if not units:
                conn.close()
                raise HTTPException(status_code=404, detail="No matching units found.")

            wb = Workbook()
            wb.remove(wb.active)

            for unit_num in sorted(units.keys()):
                ws = wb.create_sheet(title=f"unit{unit_num}")
                # column widths
                ws.column_dimensions["A"].width = 16
                ws.column_dimensions["B"].width = 50
                for col in "CDEFGHIJKL":
                    ws.column_dimensions[col].width = 24
                row = 1
                row = _heading(ws, row, f"UNIT {unit_num}", _UNIT_FILL, _WHITE_BOLD, span=12)
                row += 1

                unit_lessons = units[unit_num]
                for ltype in _TYPE_ORDER:
                    type_lessons = sorted(
                        [l for l in unit_lessons if l["type"] == ltype],
                        key=lambda x: x["title"],
                    )
                    if not type_lessons:
                        continue
                    row = _heading(ws, row, ltype.upper(), _TYPE_FILL, _WHITE_BOLD, span=12)

                    for lesson in type_lessons:
                        lid = lesson["lesson_id"]
                        # Lesson heading + per-lesson url(s) where relevant
                        extras = []
                        if ltype == "listening":
                            extras = [f"image_url: {_lesson_image_url(cur, lid)}",
                                      f"audio_url: {_lesson_audio_url(cur, lid)}"]
                        elif ltype == "reading":
                            extras = [f"image_url: {_lesson_image_url(cur, lid)}"]
                        elif ltype == "grammar":
                            extras = [f"video_url: {_lesson_video_url(cur, lid)}"]
                        head = f"Lesson: {lesson['title']}"
                        if extras:
                            head += "   |   " + "   |   ".join(extras)
                        row = _heading(ws, row, head, _LESSON_FILL, _DARK_BOLD, span=12)

                        writer = _BLOCK_WRITERS.get(ltype)
                        if writer:
                            row = writer(ws, row, cur, lesson)
                        row += 1  # spacer between lessons
                    row += 1  # spacer between type sections

                ws.freeze_panes = "A2"

        os.makedirs(body.output_dir, exist_ok=True)
        safe_label = re.sub(r"[^A-Za-z0-9_-]", "_", body.language_label)
        suffix = "" if body.units is None else "_units_" + "-".join(str(u) for u in sorted(body.units))
        out_path = os.path.join(body.output_dir, f"qa_{safe_label}{suffix}.xlsx")
        wb.save(out_path)
        logger.info("Workbook written: %s (%d unit tabs)", out_path, len(units))

    except HTTPException:
        conn.close()
        raise
    except Exception as e:
        conn.close()
        logger.error("export-qa-workbook failed: %s\n%s", e, traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Export failed: {e}")

    conn.close()
    return JSONResponse(content={
        "status": "ok",
        "language": body.language_label,
        "units_exported": sorted(units.keys()),
        "tabs": len(units),
        "output_path": out_path,
    })

    # =============================================================================
#  Streaming delete routes for QA cleanup.
#
#  /delete-questions : delete specific questions + everything attached
#  /delete-lessons   : delete whole lessons + everything under them
#
#  Both STREAM progress live as NDJSON (one JSON object per line) so you can
#  watch each step execute in real time instead of waiting for completion.
#
#  dry_run=True (default) does every delete inside a transaction, streams the
#  rowcounts, then ROLLS BACK — nothing is removed until you send dry_run=false.
#
#  Child-first deletion order (FKs in this schema are inconsistent, so we delete
#  every dependent explicitly rather than trusting ON DELETE CASCADE):
#     Image / Audio / userResponses / user_attempts / Answer  ->  Question
#     ... then LessonImages / lesson-level Image+Audio / VideoLesson(+orphan Video)
#     ... then Question / Article  ->  Lesson
#
#  Requires: from fastapi.responses import StreamingResponse
# =============================================================================

from fastapi.responses import StreamingResponse


class DeleteQuestionsRequest(BaseModel):
    db: DBConfig
    question_ids: list[int] | None = None
    keep_user_history: bool = False     # if True, leave userResponses / user_attempts rows intact
    dry_run: bool = True


class DeleteLessonsRequest(BaseModel):
    db: DBConfig
    titles: list[str] | None = None      
    lesson_ids: list[int] | None = None
    keep_user_history: bool = False
    dry_run: bool = True


def _emit(event: str, **fields) -> str:
    """Serialize one progress event as a single NDJSON line."""
    payload = {"event": event}
    payload.update(fields)
    return json.dumps(payload, ensure_ascii=False) + "\n"


def _in_clause(ids):
    return ",".join(["%s"] * len(ids))


# ----------------------------------------------------------------------------- 
# DELETE QUESTIONS
# ----------------------------------------------------------------------------- 
@app.post("/delete-questions")
def delete_questions(body: DeleteQuestionsRequest):

    def stream():
        if not body.question_ids:
            yield _emit("error", message="question_ids is empty")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        qids = body.question_ids
        fmt = _in_clause(qids)
        deleted_totals = {}
        yield _emit("start", action="delete-questions", count=len(qids),
                    dry_run=body.dry_run, keep_user_history=body.keep_user_history)
        logger.info("delete-questions started | n=%d dry_run=%s", len(qids), body.dry_run)

        try:
            with conn.cursor() as cur:
                # preview which actually exist
                cur.execute(f"SELECT question_id FROM Question WHERE question_id IN ({fmt})", tuple(qids))
                found = [r["question_id"] for r in cur.fetchall()]
                missing = [q for q in qids if q not in found]
                yield _emit("preview", existing=len(found), missing=missing)

                steps = [
                    ("Image (by question)",  f"DELETE FROM Image WHERE question_id IN ({fmt})", tuple(qids)),
                    ("Audio (by question)",  f"DELETE FROM Audio WHERE question_id IN ({fmt})", tuple(qids)),
                ]
                if not body.keep_user_history:
                    steps += [
                        ("userResponses", f"DELETE FROM userResponses WHERE question_id IN ({fmt})", tuple(qids)),
                        ("user_attempts", f"DELETE FROM user_attempts WHERE current_question_id IN ({fmt})", tuple(qids)),
                    ]
                steps += [
                    ("Answer",   f"DELETE FROM Answer WHERE question_id IN ({fmt})", tuple(qids)),
                    ("Question", f"DELETE FROM Question WHERE question_id IN ({fmt})", tuple(qids)),
                ]

                for label, sql, params in steps:
                    cur.execute(sql, params)
                    n = cur.rowcount
                    deleted_totals[label] = n
                    logger.info("delete-questions | %s -> %d rows", label, n)
                    yield _emit("step", table=label, deleted=n)

            if body.dry_run:
                conn.rollback()
                committed = False
                yield _emit("rollback", reason="dry_run")
            else:
                conn.commit()
                committed = True
                yield _emit("commit")

            yield _emit("summary", committed=committed, dry_run=body.dry_run,
                        deleted=deleted_totals)
            logger.info("delete-questions finished | committed=%s totals=%s", committed, deleted_totals)

        except Exception as e:
            conn.rollback()
            logger.error("delete-questions failed, rolled back: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# ----------------------------------------------------------------------------- 
# DELETE LESSONS
# -----------------------------------------------------------------------------
@app.post("/delete-lessons")
def delete_lessons(body: DeleteLessonsRequest):

    def stream():
        # open the connection FIRST — title resolution needs a cursor
        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        # build the target id list from lesson_ids + resolved titles
        target_ids = list(body.lesson_ids or [])
        try:
            if body.titles:
                with conn.cursor() as cur:
                    ph = ",".join(["%s"] * len(body.titles))
                    cur.execute(f"SELECT lesson_id, title FROM Lesson WHERE title IN ({ph})",
                                tuple(body.titles))
                    found = cur.fetchall()
                found_titles = {r["title"] for r in found}
                target_ids.extend(r["lesson_id"] for r in found)
                missing = [t for t in body.titles if t not in found_titles]
                if missing:
                    yield _emit("warning", message="titles not found", titles=missing)
        except Exception as e:
            conn.close()
            yield _emit("error", message=f"Failed to resolve titles: {e}")
            return

        target_ids = sorted(set(target_ids))   # dedup if a title and its id both given

        if not target_ids:
            conn.close()
            yield _emit("error", message="No lessons to delete (no valid titles or lesson_ids)")
            return

        grand_totals: dict = {}
        yield _emit("start", action="delete-lessons", count=len(target_ids),
                    dry_run=body.dry_run, keep_user_history=body.keep_user_history)
        logger.info("delete-lessons started | n=%d dry_run=%s", len(target_ids), body.dry_run)

        def bump(label, n):
            grand_totals[label] = grand_totals.get(label, 0) + n

        try:
            for lid in target_ids:
                with conn.cursor() as cur:
                    cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE lesson_id = %s", (lid,))
                    lesson = cur.fetchone()
                    if not lesson:
                        yield _emit("lesson_skipped", lesson_id=lid, reason="not found")
                        continue

                    cur.execute("SELECT question_id FROM Question WHERE lesson_id = %s", (lid,))
                    qids = [r["question_id"] for r in cur.fetchall()]
                    yield _emit("lesson_start", lesson_id=lid, title=lesson["title"],
                                type=lesson["type"], questions=len(qids))

                    per_lesson = {}

                    # 1) children of this lesson's questions
                    if qids:
                        qfmt = _in_clause(qids)
                        qsteps = [
                            ("Image(q)", f"DELETE FROM Image WHERE question_id IN ({qfmt})"),
                            ("Audio(q)", f"DELETE FROM Audio WHERE question_id IN ({qfmt})"),
                        ]
                        if not body.keep_user_history:
                            qsteps += [
                                ("userResponses", f"DELETE FROM userResponses WHERE question_id IN ({qfmt})"),
                                ("user_attempts", f"DELETE FROM user_attempts WHERE current_question_id IN ({qfmt})"),
                            ]
                        qsteps += [("Answer", f"DELETE FROM Answer WHERE question_id IN ({qfmt})")]
                        for label, sql in qsteps:
                            cur.execute(sql, tuple(qids))
                            n = cur.rowcount
                            per_lesson[label] = n; bump(label, n)
                            yield _emit("step", lesson_id=lid, table=label, deleted=n)

                    # 2) lesson-level media
                    lvl = [
                        ("LessonImages", "DELETE FROM LessonImages WHERE lesson_id = %s"),
                        ("Image(lesson)", "DELETE FROM Image WHERE lesson_id = %s"),
                        ("Audio(lesson)", "DELETE FROM Audio WHERE lesson_id = %s"),
                        ("VideoLesson", "DELETE FROM VideoLesson WHERE lesson_id = %s"),
                    ]
                    for label, sql in lvl:
                        cur.execute(sql, (lid,))
                        n = cur.rowcount
                        per_lesson[label] = n; bump(label, n)
                        yield _emit("step", lesson_id=lid, table=label, deleted=n)

                    # orphaned Video rows (no VideoLesson points to them anymore)
                    cur.execute(
                        "DELETE v FROM Video v "
                        "LEFT JOIN VideoLesson vl ON vl.video_id = v.video_id "
                        "WHERE vl.video_id IS NULL"
                    )
                    n = cur.rowcount
                    per_lesson["Video(orphan)"] = n; bump("Video(orphan)", n)
                    yield _emit("step", lesson_id=lid, table="Video(orphan)", deleted=n)

                    # 3) questions + articles
                    for label, sql in [
                        ("Question", "DELETE FROM Question WHERE lesson_id = %s"),
                        ("Article",  "DELETE FROM Article WHERE lesson_id = %s"),
                    ]:
                        cur.execute(sql, (lid,))
                        n = cur.rowcount
                        per_lesson[label] = n; bump(label, n)
                        yield _emit("step", lesson_id=lid, table=label, deleted=n)

                    # 4) the lesson
                    cur.execute("DELETE FROM Lesson WHERE lesson_id = %s", (lid,))
                    n = cur.rowcount
                    per_lesson["Lesson"] = n; bump("Lesson", n)
                    yield _emit("step", lesson_id=lid, table="Lesson", deleted=n)

                    yield _emit("lesson_done", lesson_id=lid, deleted=per_lesson)
                    logger.info("delete-lessons | lesson %d done | %s", lid, per_lesson)

            if body.dry_run:
                conn.rollback()
                committed = False
                yield _emit("rollback", reason="dry_run")
            else:
                conn.commit()
                committed = True
                yield _emit("commit")

            yield _emit("summary", committed=committed, dry_run=body.dry_run,
                        grand_totals=grand_totals)
            logger.info("delete-lessons finished | committed=%s totals=%s", committed, grand_totals)

        except Exception as e:
            conn.rollback()
            logger.error("delete-lessons failed, rolled back: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# =============================================================================
#  /regenerate-lesson-media
#
#  Re-hydrate media for lessons after a re-insert, OR top-up media for a few
#  newly-added questions, WITHOUT needing to know database-generated ids.
#
#  Select lessons by (priority order):
#     lesson_ids  -> explicit ids (if you happen to know them)
#     titles      -> exact Lesson.title match  (best for re-insertion: titles are stable)
#     units       -> every unitN_ lesson in those unit numbers
#
#  Optional question_sequence_ids: narrow vocab work to specific questions by their
#  JSON sequence_id (stable & human-known), resolved to current question_id at runtime.
#
#  Behaviour: SKIP-IF-PRESENT. Generates media only for questions/lessons that are
#  currently missing it; never overwrites existing media. This makes whole-lesson
#  re-hydrate and partial top-up the same safe call.
#
#  Per type:
#     vocabulary -> images (+link) and audio (+link) for questions missing them
#     reading    -> link per-lesson image via unit-image ingest (if missing)
#     listening  -> link per-question images via unit-image ingest (if missing)
#     grammar    -> NO generation; emits a manual_todo checklist (audio is generated
#                   separately, video is made by an external tool, then linked)
#     speaking   -> nothing (text-only; no media in schema)
#
#  Streams NDJSON progress (one JSON object per line) so you can watch it live.
#  dry_run=True (default): generation routines are NOT called; it reports the plan
#  (what's missing, what it would generate) and rolls back any lookups.
#
#  Reuses existing helpers/classes already in main.py:
#     VocabToPictures, _resize_to_256_png_bytes, _upload_to_s3_public, _safe_slug,
#     _insert_image_row, _make_s3_client, _generate_tts_bytes, _translate_text,
#     _insert_vocab_audio_row, _extract_unit_num_from_title, connect_to_db, _emit
#  Requires: from fastapi.responses import StreamingResponse  (already imported)
# =============================================================================


class RegenerateLessonMediaRequest(BaseModel):
    db: DBConfig
    # --- selection (use ONE; checked in this priority order) ---
    lesson_ids: list[int] | None = None
    titles: list[str] | None = None
    units: list[int] | None = None
    # --- optional narrowing for partial top-up (vocab questions by JSON sequence_id) ---
    question_sequence_ids: list[int] | None = None
    # --- S3 / credentials (same fields as the other generation routes) ---
    s3_bucket: str
    s3_image_prefix: str = "images"
    s3_audio_prefix: str = "audio"
    aws_region: str = "us-east-1"
    aws_access_key_id: str | None = None
    aws_secret_access_key: str | None = None
    openai_api_key: str | None = None
    # vocab audio translation (only used when an answer is blank)
    translate_vocab_audio: bool = False
    source_language: str = "en"
    target_language: str = "es"
    tts_model: str = "gpt-4o-mini-tts"
    voice: str = "alloy"
    tts_instructions: str | None = "Speak slowly and clearly with a warm, friendly, teacher-like tone."
    translate_images: bool = False     # passed to VocabToPictures.generate_one
    dry_run: bool = True


@app.post("/regenerate-lesson-media")
def regenerate_lesson_media(body: RegenerateLessonMediaRequest):

    def stream():
        # --- resolve selection to current lessons ---
        if not (body.lesson_ids or body.titles or body.units):
            yield _emit("error", message="Provide one of lesson_ids, titles, or units")
            return
        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="regenerate-lesson-media", dry_run=body.dry_run,
                    selector=("lesson_ids" if body.lesson_ids else
                              "titles" if body.titles else "units"))

        try:
            with conn.cursor() as cur:
                if body.lesson_ids:
                    fmt = ",".join(["%s"] * len(body.lesson_ids))
                    cur.execute(
                        f"SELECT lesson_id, title, type FROM Lesson WHERE lesson_id IN ({fmt})",
                        tuple(body.lesson_ids),
                    )
                    lessons = cur.fetchall()
                elif body.titles:
                    fmt = ",".join(["%s"] * len(body.titles))
                    cur.execute(
                        f"SELECT lesson_id, title, type FROM Lesson WHERE title IN ({fmt})",
                        tuple(body.titles),
                    )
                    lessons = cur.fetchall()
                else:  # units
                    likes = " OR ".join(["title LIKE %s"] * len(body.units))
                    params = [f"unit{u}\\_%" for u in body.units]
                    cur.execute(
                        f"SELECT lesson_id, title, type FROM Lesson WHERE {likes} ORDER BY title",
                        tuple(params),
                    )
                    lessons = cur.fetchall()

            if not lessons:
                yield _emit("error", message="No lessons matched the selection")
                conn.close()
                return

            yield _emit("resolved", count=len(lessons),
                        lessons=[{"lesson_id": l["lesson_id"], "title": l["title"],
                                  "type": l["type"]} for l in lessons])

            # --- shared clients (only built when not dry_run) ---
            s3 = None
            img_gen = None
            openai_client = None
            translate_client = None
            if not body.dry_run:
                s3 = _make_s3_client(body.aws_region, body.aws_access_key_id, body.aws_secret_access_key)
                img_gen = VocabToPictures(api_key=body.openai_api_key, model="gpt-image-1", size="1024x1024")
                openai_client = OpenAI(api_key=body.openai_api_key or os.getenv("OPENAI_API_KEY"))
                tkw = {"region_name": body.aws_region}
                if body.aws_access_key_id and body.aws_secret_access_key:
                    tkw["aws_access_key_id"] = body.aws_access_key_id
                    tkw["aws_secret_access_key"] = body.aws_secret_access_key
                translate_client = boto3.client("translate", **tkw)

            grand = {"images": 0, "audio": 0, "skipped": 0, "manual_grammar": 0, "errors": 0}

            for lesson in lessons:
                lid = lesson["lesson_id"]
                ltype = lesson["type"]
                title = lesson["title"]
                yield _emit("lesson_start", lesson_id=lid, title=title, type=ltype)

                # ---------------- speaking: nothing ----------------
                if ltype == "speaking":
                    yield _emit("lesson_done", lesson_id=lid, note="speaking is text-only; no media")
                    continue

                # ---------------- grammar: manual checklist ----------------
                if ltype == "grammar":
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT v.url FROM Video v JOIN VideoLesson vl ON vl.video_id=v.video_id "
                            "WHERE vl.lesson_id=%s LIMIT 1", (lid,))
                        has_video = cur.fetchone() is not None
                    grand["manual_grammar"] += 1
                    yield _emit("manual_todo", lesson_id=lid, title=title,
                                video_linked=has_video,
                                steps=[
                                    "run /generate-grammar-audio for this lesson",
                                    "create the video with the external tool (named with THIS lesson_id)",
                                    "upload the .mp4 to S3, then run /link-grammar-videos",
                                ])
                    yield _emit("lesson_done", lesson_id=lid)
                    continue

                # ---------------- vocabulary: images + audio ----------------
                if ltype == "vocabulary":
                    # questions missing an image / missing audio (skip-if-present)
                    with conn.cursor() as cur:
                        seq_filter = ""
                        params = [lid]
                        if body.question_sequence_ids:
                            seqfmt = ",".join(["%s"] * len(body.question_sequence_ids))
                            seq_filter = f" AND q.sequence_id IN ({seqfmt})"
                            params += list(body.question_sequence_ids)

                        cur.execute(
                            f"""SELECT q.question_id, q.sequence_id, q.question_text,
                                       (SELECT COUNT(*) FROM Image i WHERE i.question_id=q.question_id) AS has_img,
                                       (SELECT COUNT(*) FROM Audio a WHERE a.question_id=q.question_id) AS has_aud
                                FROM Question q
                                WHERE q.lesson_id=%s{seq_filter}
                                ORDER BY q.sequence_id""",
                            tuple(params),
                        )
                        qrows = cur.fetchall()

                    need_img = [q for q in qrows if not q["has_img"]]
                    need_aud = [q for q in qrows if not q["has_aud"]]
                    yield _emit("vocab_plan", lesson_id=lid,
                                questions=len(qrows), need_image=len(need_img), need_audio=len(need_aud))

                    if body.dry_run:
                        grand["skipped"] += len(qrows)
                        yield _emit("lesson_done", lesson_id=lid, note="dry_run: no generation")
                        continue

                    # --- images ---
                    for q in need_img:
                        try:
                            gen = img_gen.generate_one(
                                word=q["question_text"],
                                translate=body.translate_images,
                            )
                            if not gen or not gen.get("image_bytes"):
                                raise RuntimeError("no image returned")
                            png = _resize_to_256_png_bytes(gen["image_bytes"])
                            key = f"{body.s3_image_prefix}/{lid}_{q['question_id']}_{_safe_slug(title)}.png"
                            url = _upload_to_s3_public(png, key, body.s3_bucket, body.aws_region,
                                                       body.aws_access_key_id, body.aws_secret_access_key)
                            meta = {"original_text": gen["original_text"],
                                    "translated_text": gen["translated_text"],
                                    "s3_key": key, "public_url": url, "lesson_title": title}
                            _insert_image_row(conn, lesson_id=lid, question_id=q["question_id"],
                                              image_url=url, image_metadata_json=json.dumps(meta, ensure_ascii=False))
                            conn.commit()
                            grand["images"] += 1
                            yield _emit("image", lesson_id=lid, question_id=q["question_id"], url=url)
                        except Exception as e:
                            conn.rollback(); grand["errors"] += 1
                            yield _emit("image_error", lesson_id=lid, question_id=q["question_id"], error=str(e))

                    # --- audio ---
                    for q in need_aud:
                        try:
                            with conn.cursor() as cur:
                                cur.execute(
                                    "SELECT answer_text FROM Answer WHERE question_id=%s AND is_correct=1 LIMIT 1",
                                    (q["question_id"],))
                                arow = cur.fetchone()
                                answer_text = (arow["answer_text"] if arow else "") or ""

                            if not answer_text.strip() and body.translate_vocab_audio:
                                translated = _translate_text(translate_client, q["question_text"],
                                                             body.source_language, body.target_language)
                                with conn.cursor() as cur:
                                    cur.execute("SELECT COUNT(*) AS c FROM Answer WHERE question_id=%s",
                                                (q["question_id"],))
                                    if cur.fetchone()["c"] > 0:
                                        cur.execute("UPDATE Answer SET answer_text=%s WHERE question_id=%s AND is_correct=1",
                                                    (translated, q["question_id"]))
                                    else:
                                        cur.execute("INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                                    "VALUES (%s,%s,%s,1)", (lid, q["question_id"], translated))
                                answer_text = translated

                            speak = answer_text.strip() or q["question_text"]
                            audio_bytes = _generate_tts_bytes(openai_client, speak,
                                                              body.tts_model, body.voice, body.tts_instructions)
                            key = f"{body.s3_audio_prefix}/{lid}_{q['question_id']}_{_safe_slug(title)}.mp3"
                            s3.put_object(Bucket=body.s3_bucket, Key=key, Body=audio_bytes, ContentType="audio/mpeg")
                            url = f"https://{body.s3_bucket}.s3.{body.aws_region}.amazonaws.com/{key}"
                            _insert_vocab_audio_row(conn, lid, q["question_id"], q["sequence_id"], url,
                                                    body.tts_model, body.voice)
                            conn.commit()
                            grand["audio"] += 1
                            yield _emit("audio", lesson_id=lid, question_id=q["question_id"], url=url)
                        except Exception as e:
                            conn.rollback(); grand["errors"] += 1
                            yield _emit("audio_error", lesson_id=lid, question_id=q["question_id"], error=str(e))

                    yield _emit("lesson_done", lesson_id=lid)
                    continue

                # ---------------- reading / listening: image linking ----------------
                if ltype in ("reading", "listening"):
                    # report whether a lesson-level / question-level image already exists;
                    # actual linking is handled by /ingest-unit-images (unit-scoped).
                    with conn.cursor() as cur:
                        cur.execute("SELECT COUNT(*) AS c FROM Image WHERE lesson_id=%s", (lid,))
                        has_img = cur.fetchone()["c"]
                    unit = _extract_unit_num_from_title(title)
                    yield _emit("link_via_ingest", lesson_id=lid, type=ltype, unit=unit,
                                images_present=has_img,
                                note=f"run /ingest-unit-images with units=[{unit}] to (re)link the unit image")
                    yield _emit("lesson_done", lesson_id=lid)
                    continue

                # ---------------- unknown type ----------------
                yield _emit("lesson_done", lesson_id=lid, note=f"no media rule for type '{ltype}'")

            if body.dry_run:
                conn.rollback()
            yield _emit("summary", dry_run=body.dry_run, totals=grand)
            logger.info("regenerate-lesson-media finished | dry_run=%s totals=%s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("regenerate-lesson-media failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e))
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")

# =============================================================================
#  /insert-content
#
#  Append question(s) and/or article(s) to an EXISTING lesson, without touching
#  the rest of the lesson. Reuses the same JSON item shapes as /insert-lessons.
#
#  You send ONLY the new items, not the whole lesson:
#     - new questions  -> "questions_and_answers": [ {question_text, answers:[...]} , ... ]
#     - new articles   -> "articles": [ {text} , ... ]
#
#  Target the lesson by `title` (stable across re-inserts) or `lesson_id`.
#
#  Auto-handled (do NOT put in payload):
#     - sequence_id   : appended as max(sequence_id)+1, incrementing per item
#                       (you MAY pass sequence_id on an item to force a value)
#     - lesson_id     : filled from the resolved lesson
#     - has_answer    : set from whether the question has a real (non-blank) answer
#     - answer_category : derived (learning/open_ended/practice/multiple_choice)
#
#  Streams NDJSON progress. dry_run=True (default) inserts inside a transaction,
#  streams what it WOULD create, then ROLLS BACK.
#
#  Does NOT generate media — follow with /regenerate-lesson-media (by title +
#  the new question sequence_ids) to create images/audio for new vocab questions.
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse.
# =============================================================================


class InsertContentRequest(BaseModel):
    db: DBConfig
    # --- target (use ONE) ---
    title: str | None = None
    lesson_id: int | None = None
    # --- new content (same shapes as /insert-lessons) ---
    questions_and_answers: list[dict] | None = None
    articles: list[dict] | None = None
    dry_run: bool = True


def _derive_answer_category(lesson_type: str, real_answers: int, correct_count: int) -> str | None:
    """Same rule as the grammar backfill, generalized for any lesson type."""
    if real_answers > 1 and correct_count >= 1:
        return "multiple_choice"
    if real_answers >= 1:
        return "practice"
    if real_answers == 0 and lesson_type in ("vocabulary", "grammar"):
        return "learning"
    if real_answers == 0 and lesson_type in ("reading", "writing", "speaking", "listening"):
        return "open_ended"
    return None


@app.post("/insert-content")
def insert_content(body: InsertContentRequest):

    def stream():
        if not body.title and not body.lesson_id:
            yield _emit("error", message="Provide a target: title or lesson_id")
            return
        if not body.questions_and_answers and not body.articles:
            yield _emit("error", message="Provide questions_and_answers and/or articles to insert")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="insert-content", dry_run=body.dry_run)

        try:
            with conn.cursor() as cur:
                # --- resolve target lesson ---
                if body.lesson_id:
                    cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE lesson_id=%s", (body.lesson_id,))
                else:
                    cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE title=%s", (body.title,))
                lesson = cur.fetchone()
                if not lesson:
                    yield _emit("error", message="Target lesson not found")
                    conn.close(); return

                lid = lesson["lesson_id"]
                ltype = lesson["type"]
                yield _emit("resolved", lesson_id=lid, title=lesson["title"], type=ltype)

                # --- current max sequence ids ---
                cur.execute("SELECT COALESCE(MAX(sequence_id),0) AS m FROM Article WHERE lesson_id=%s", (lid,))
                art_seq = cur.fetchone()["m"]
                cur.execute("SELECT COALESCE(MAX(sequence_id),0) AS m FROM Question WHERE lesson_id=%s", (lid,))
                q_seq = cur.fetchone()["m"]

                inserted = {"articles": [], "questions": []}

                # --- articles ---
                for art in (body.articles or []):
                    text = art.get("text", art.get("content", ""))
                    seq = art.get("sequence_id")
                    if seq is None:
                        art_seq += 1
                        seq = art_seq
                    cur.execute(
                        "INSERT INTO Article (lesson_id, sequence_id, content) VALUES (%s,%s,%s)",
                        (lid, seq, text),
                    )
                    aid = cur.lastrowid
                    inserted["articles"].append({"article_id": aid, "sequence_id": seq})
                    yield _emit("article_inserted", lesson_id=lid, article_id=aid, sequence_id=seq)

                # --- questions (+ answers) ---
                for qa in (body.questions_and_answers or []):
                    question_text = qa.get("question_text") or qa.get("question", "")
                    seq = qa.get("sequence_id")
                    if seq is None:
                        q_seq += 1
                        seq = q_seq

                    # normalize answers into a list of {text, is_correct}
                    answers = []
                    if "answers" in qa and isinstance(qa["answers"], list):
                        for ans in qa["answers"]:
                            answers.append({
                                "text": ans.get("answer_text") or ans.get("answer") or ans.get("text", ""),
                                "is_correct": int(ans.get("is_correct", False)),
                            })
                    elif "answer" in qa:
                        raw = qa["answer"]
                        if isinstance(raw, str):
                            answers.append({"text": raw, "is_correct": int(qa.get("is_correct", True))})
                        elif isinstance(raw, dict):
                            answers.append({"text": raw.get("text") or raw.get("answer", ""),
                                            "is_correct": int(raw.get("is_correct", True))})
                        elif isinstance(raw, list):
                            for ans in raw:
                                answers.append({"text": ans.get("answer", ""),
                                                "is_correct": int(ans.get("is_correct", True))})

                    real_answers = sum(1 for a in answers if a["text"].strip() != "")
                    correct_count = sum(1 for a in answers if a["is_correct"])
                    has_answer = 1 if real_answers >= 1 else 0
                    qtype = qa.get("type") or ("multiple_choice" if len(answers) > 1 else "short_answer")
                    category = _derive_answer_category(ltype, real_answers, correct_count)

                    cur.execute(
                        "INSERT INTO Question (lesson_id, sequence_id, question_text, type, has_answer, answer_category) "
                        "VALUES (%s,%s,%s,%s,%s,%s)",
                        (lid, seq, question_text, qtype, has_answer, category),
                    )
                    qid = cur.lastrowid
                    for a in answers:
                        cur.execute(
                            "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) VALUES (%s,%s,%s,%s)",
                            (lid, qid, a["text"], a["is_correct"]),
                        )
                    inserted["questions"].append({
                        "question_id": qid, "sequence_id": seq, "type": qtype,
                        "answers": len(answers), "has_answer": has_answer, "answer_category": category,
                    })
                    yield _emit("question_inserted", lesson_id=lid, question_id=qid, sequence_id=seq,
                                type=qtype, answers=len(answers), answer_category=category)

                # keep Lesson.has_question correct if we added the first questions
                if inserted["questions"]:
                    cur.execute("UPDATE Lesson SET has_question=1 WHERE lesson_id=%s", (lid,))

            if body.dry_run:
                conn.rollback()
                committed = False
                yield _emit("rollback", reason="dry_run")
            else:
                conn.commit()
                committed = True
                yield _emit("commit")

            new_q_seqs = [q["sequence_id"] for q in inserted["questions"]]
            yield _emit("summary", committed=committed, dry_run=body.dry_run,
                        lesson_id=lid, title=lesson["title"],
                        articles_inserted=len(inserted["articles"]),
                        questions_inserted=len(inserted["questions"]),
                        new_question_sequence_ids=new_q_seqs,
                        next_step=(
                            "run /regenerate-lesson-media with this title and "
                            f"question_sequence_ids={new_q_seqs} to create media for new vocab questions"
                            if ltype == "vocabulary" and new_q_seqs else None
                        ))
            logger.info("insert-content | lesson=%s committed=%s +%dq +%da",
                        lid, committed, len(inserted["questions"]), len(inserted["articles"]))

        except Exception as e:
            conn.rollback()
            logger.error("insert-content failed, rolled back: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")

# =============================================================================
#  Add to main.py  —  /rebalance-grammar-practice-learning
#
#  Spanish situation: each grammar lesson is currently EITHER all-practice OR
#  all-learning. Rebalance each lesson to ~50/50.
#
#  Per lesson (target_learning = floor(total/2)):
#     - all-practice lesson  -> blank the FIRST half  -> learning
#                               (keep one row, answer_text='', is_correct=0, has_answer=0)
#     - all-learning lesson   -> GENERATE answers for the SECOND half -> practice
#                               (fill the blank row via LLM, is_correct=1, has_answer=1)
#     - mixed lesson          -> generalises: move just enough to hit target_learning,
#                               blanking practice->learning or generating learning->practice
#
#  learning -> practice REQUIRES inventing an answer (the original was deleted when the
#  question became learning). Answers are generated from question_text via OpenAI and are
#  GUESSES — every generated question_id is reported so QA can verify them.
#
#  After mutation, answer_category is re-derived from structure (the grammar backfill).
#
#  Streams NDJSON. dry_run=True (default): reports the plan (no LLM calls, no writes),
#  then rolls back.
#
#  Reuses: connect_to_db, DBConfig, logger, _emit, StreamingResponse,
#          _GRAMMAR_BACKFILL_SQL, OpenAI, os.
# =============================================================================


class RebalancePracticeLearningRequest(BaseModel):
    db: DBConfig
    lesson_ids: list[int] | None = None     # restrict to these grammar lessons; None = ALL grammar
    learning_ratio: float = 0.5
    openai_api_key: str | None = None       # falls back to OPENAI_API_KEY env var
    model: str = "gpt-4o-mini"
    dry_run: bool = True


def _generate_grammar_answer(client, model: str, question_text: str) -> str:
    """Ask the LLM for the single correct answer to a grammar question. Returns plain text."""
    resp = client.chat.completions.create(
        model=model,
        temperature=0,
        messages=[
            {"role": "system", "content": (
                "You are a precise language-learning answer key. Given a grammar practice "
                "question, reply with ONLY the single correct answer — no explanation, no "
                "punctuation beyond what the answer itself needs, no quotes, no preamble."
            )},
            {"role": "user", "content": question_text},
        ],
    )
    return (resp.choices[0].message.content or "").strip()


def _to_learning(cur, question_id: int) -> int:
    """Blank a question into learning: keep lowest answer_id row, blank it, has_answer=0."""
    cur.execute("SELECT answer_id FROM Answer WHERE question_id=%s ORDER BY answer_id ASC", (question_id,))
    rows = cur.fetchall()
    deleted = 0
    if rows:
        keep = rows[0]["answer_id"]
        cur.execute("DELETE FROM Answer WHERE question_id=%s AND answer_id<>%s", (question_id, keep))
        deleted = cur.rowcount
        cur.execute("UPDATE Answer SET answer_text='', is_correct=0 WHERE answer_id=%s", (keep,))
    cur.execute("UPDATE Question SET type='short_answer', has_answer=0 WHERE question_id=%s", (question_id,))
    return deleted


def _to_practice_with_answer(cur, question_id: int, lesson_id: int, answer_text: str) -> None:
    """Fill a (blank) question's answer to make it practice. Keeps exactly one row."""
    cur.execute("SELECT answer_id FROM Answer WHERE question_id=%s ORDER BY answer_id ASC", (question_id,))
    rows = cur.fetchall()
    if rows:
        keep = rows[0]["answer_id"]
        cur.execute("DELETE FROM Answer WHERE question_id=%s AND answer_id<>%s", (question_id, keep))
        cur.execute("UPDATE Answer SET answer_text=%s, is_correct=1 WHERE answer_id=%s", (answer_text, keep))
    else:
        cur.execute(
            "INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) VALUES (%s,%s,%s,1)",
            (lesson_id, question_id, answer_text),
        )
    cur.execute("UPDATE Question SET type='short_answer', has_answer=1 WHERE question_id=%s", (question_id,))


@app.post("/rebalance-grammar-practice-learning")
def rebalance_grammar_practice_learning(body: RebalancePracticeLearningRequest):

    def stream():
        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="rebalance-practice-learning",
                    dry_run=body.dry_run, learning_ratio=body.learning_ratio)

        client = None
        if not body.dry_run:
            client = OpenAI(api_key=body.openai_api_key or os.getenv("OPENAI_API_KEY"))

        generated_ids = []
        grand = {"to_learning": 0, "to_practice_generated": 0, "lessons": 0, "errors": 0}

        try:
            with conn.cursor() as cur:
                if body.lesson_ids:
                    fmt = ",".join(["%s"] * len(body.lesson_ids))
                    cur.execute(f"SELECT lesson_id, title FROM Lesson WHERE type='grammar' AND lesson_id IN ({fmt})",
                                tuple(body.lesson_ids))
                else:
                    cur.execute("SELECT lesson_id, title FROM Lesson WHERE type='grammar' ORDER BY title")
                grammar = cur.fetchall()

            if not grammar:
                yield _emit("error", message="No grammar lessons in scope")
                conn.close(); return

            lesson_id_list = [g["lesson_id"] for g in grammar]
            yield _emit("resolved", lessons=len(grammar))

            for g in grammar:
                lid = g["lesson_id"]
                # current state of each question: is it learning (no real answer) or practice?
                cur = conn.cursor()
                cur.execute(
                    """SELECT q.question_id, q.question_text,
                              COUNT(CASE WHEN TRIM(a.answer_text) <> '' THEN a.answer_id END) AS real_answers
                       FROM Question q
                       LEFT JOIN Answer a ON a.question_id = q.question_id
                       WHERE q.lesson_id = %s
                       GROUP BY q.question_id, q.question_text
                       ORDER BY q.question_id ASC""",
                    (lid,),
                )
                qs = cur.fetchall()
                cur.close()
                total = len(qs)
                if total == 0:
                    continue
                grand["lessons"] += 1

                learning_q = [q for q in qs if q["real_answers"] == 0]
                practice_q = [q for q in qs if q["real_answers"] >= 1]
                target_learning = int(total * body.learning_ratio)

                plan = {"lesson_id": lid, "title": g["title"], "total": total,
                        "current_learning": len(learning_q), "current_practice": len(practice_q),
                        "target_learning": target_learning}

                # decide moves
                make_learning = []   # practice -> learning (blank)
                make_practice = []   # learning -> practice (generate)

                if len(learning_q) < target_learning:
                    # need more learning: blank the first (target - current) practice questions
                    need = target_learning - len(learning_q)
                    make_learning = practice_q[:need]
                elif len(learning_q) > target_learning:
                    # too much learning: generate answers for the excess (turn into practice)
                    excess = len(learning_q) - target_learning
                    # "second half as practice" -> take from the end of the learning list
                    make_practice = learning_q[-excess:]

                plan["will_blank_to_learning"] = len(make_learning)
                plan["will_generate_to_practice"] = len(make_practice)
                yield _emit("lesson_plan", **plan)

                if body.dry_run:
                    continue

                cur = conn.cursor()
                # practice -> learning (cheap, no LLM)
                for q in make_learning:
                    _to_learning(cur, q["question_id"])
                    grand["to_learning"] += 1
                    yield _emit("blanked", lesson_id=lid, question_id=q["question_id"])

                # learning -> practice (LLM-generated answer)
                for q in make_practice:
                    try:
                        ans = _generate_grammar_answer(client, body.model, q["question_text"])
                        if not ans:
                            raise RuntimeError("LLM returned empty answer")
                        _to_practice_with_answer(cur, q["question_id"], lid, ans)
                        grand["to_practice_generated"] += 1
                        generated_ids.append(q["question_id"])
                        yield _emit("generated", lesson_id=lid, question_id=q["question_id"],
                                    answer=ans)
                    except Exception as e:
                        grand["errors"] += 1
                        yield _emit("generate_error", lesson_id=lid, question_id=q["question_id"], error=str(e))
                conn.commit()
                cur.close()

            # re-derive answer_category from new structure
            if not body.dry_run:
                fmt_ids = ",".join(["%s"] * len(lesson_id_list))
                with conn.cursor() as cur:
                    cur.execute(_GRAMMAR_BACKFILL_SQL.format(ids=fmt_ids),
                                tuple(lesson_id_list) + tuple(lesson_id_list))
                    backfilled = cur.rowcount
                conn.commit()
                yield _emit("backfill", rows_updated=backfilled)

                # verification
                with conn.cursor() as cur:
                    cur.execute(
                        f"SELECT answer_category, COUNT(*) AS c FROM Question "
                        f"WHERE lesson_id IN ({fmt_ids}) GROUP BY answer_category",
                        tuple(lesson_id_list))
                    cats = {r["answer_category"]: r["c"] for r in cur.fetchall()}
                yield _emit("verification", overall_categories=cats)

            if body.dry_run:
                conn.rollback()

            yield _emit("summary", dry_run=body.dry_run, totals=grand,
                        generated_question_ids=generated_ids,
                        note="generated_question_ids contain AI-written answers — QA should verify them")
            logger.info("rebalance-practice-learning done | dry_run=%s totals=%s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("rebalance-practice-learning failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# =============================================================================
#  Add to main.py  —  /sync-vocab-lesson
#
#  Reconcile a RE-EXTRACTED vocabulary lesson against the existing DB lesson,
#  IN PLACE, reusing correct images and only paying to regenerate what changed.
#
#  Per lesson (matched by title), one LLM call matches questions by MEANING:
#     - PERFECT (matched + identical text) -> nothing changes (image + audio kept)
#     - FUZZY   (matched + text differs)   -> update question_text & answer_text;
#                                             KEEP image; DELETE audio row (so it regenerates)
#     - NEW     (new question, no match)   -> insert question + answer (image+audio later)
#     - REMOVED (db question, no match)    -> delete it + its image/audio links
#
#  The LLM only matches by meaning using index labels (D1.., N1..) and returns
#  index pairs — it never sees ids and never handles data. We decide perfect-vs-fuzzy
#  ourselves by comparing the actual text. One-to-one is enforced: a reused index is
#  treated as NOT matched (so it becomes NEW / REMOVED) and logged.
#
#  This route does NO media generation. Afterwards run /regenerate-lesson-media by
#  title with the reported question_sequence_ids: NEW questions get image+audio,
#  FUZZY questions get audio regenerated (image already present is skipped).
#
#  Streams NDJSON. dry_run=True (default): makes the (cheap) LLM call so you can
#  judge the matching, reports every fuzzy/new/removed decision, but writes nothing.
#
#  Input: re-extracted JSON files (same shape as /insert-lessons): each file has
#  "title" and "questions_and_answers": [{question_text, answers:[{answer_text}]}].
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse / OpenAI / os / Path.
# =============================================================================


class SyncVocabLessonRequest(BaseModel):
    db: DBConfig
    folder: str | None = None          # directory of re-extracted JSON files
    files: list[str] | None = None     # or explicit file paths
    model: str = "gpt-4o-mini"
    openai_api_key: str | None = None
    keep_user_history: bool = False    # if True, don't delete userResponses/user_attempts for removed questions
    dry_run: bool = True


def _norm_text(s: str) -> str:
    """Light normalization for the perfect-vs-fuzzy decision: trim + collapse whitespace."""
    return re.sub(r"\s+", " ", (s or "").strip())


def _llm_match_questions(client, model: str, db_texts: list[str], new_texts: list[str]) -> list[dict]:
    """
    Ask the LLM to match by meaning. db_texts/new_texts are plain strings; we label them
    D1.. and N1.. in the prompt. Returns list of {"db": "D3", "new": "N5"}.
    """
    db_block = "\n".join(f"D{i+1}: {t}" for i, t in enumerate(db_texts))
    new_block = "\n".join(f"N{i+1}: {t}" for i, t in enumerate(new_texts))
    system = (
        "You match vocabulary items between two lists by MEANING. Two entries match if they "
        "refer to the same vocabulary item even if the wording differs slightly (e.g. 'Dog' vs "
        "'The Dog', or a small spelling/accent difference). Be conservative: only pair items you "
        "are confident are the same item. Return ONLY JSON."
    )
    user = (
        f"DB LIST:\n{db_block}\n\n"
        f"NEW LIST:\n{new_block}\n\n"
        "Return a JSON array of matched pairs, each like {\"db\":\"D1\",\"new\":\"N2\"}. "
        "Each D label and each N label may appear AT MOST ONCE. Omit anything with no confident match. "
        "No prose, no code fences."
    )
    resp = client.chat.completions.create(
        model=model, temperature=0,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
    )
    content = (resp.choices[0].message.content or "").strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
    data = json.loads(content)
    if not isinstance(data, list):
        raise ValueError("LLM did not return a JSON array of pairs")
    return data


@app.post("/sync-vocab-lesson")
def sync_vocab_lesson(body: SyncVocabLessonRequest):

    def stream():
        # gather files
        files: list[Path] = []
        if body.folder:
            p = Path(body.folder)
            if not p.is_dir():
                yield _emit("error", message=f"Folder not found: {body.folder}"); return
            files.extend(sorted(p.glob("*.json")))
        if body.files:
            for f in body.files:
                fp = Path(f)
                if fp.is_file() and fp not in files:
                    files.append(fp)
        if not files:
            yield _emit("error", message="No re-extracted JSON files provided"); return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}"); return

        client = OpenAI(api_key=body.openai_api_key or os.getenv("OPENAI_API_KEY"))

        yield _emit("start", action="sync-vocab-lesson", files=len(files), dry_run=body.dry_run)

        grand = {"perfect": 0, "fuzzy": 0, "new": 0, "removed": 0,
                 "duplicates_ignored": 0, "lessons": 0, "errors": 0}
        regen_plan = []  # [{title, question_sequence_ids: [...]}]

        try:
            for fp in files:
                try:
                    data = json.loads(fp.read_text(encoding="utf-8-sig"))
                except Exception as e:
                    grand["errors"] += 1
                    yield _emit("file_error", file=fp.name, error=f"JSON parse: {e}")
                    continue

                title = data.get("title")
                new_qas = data.get("questions_and_answers", []) or []
                if not title:
                    grand["errors"] += 1
                    yield _emit("file_error", file=fp.name, error="missing title")
                    continue

                # resolve existing lesson by title
                with conn.cursor() as cur:
                    cur.execute("SELECT lesson_id, type FROM Lesson WHERE title=%s", (title,))
                    lrow = cur.fetchone()
                if not lrow:
                    grand["errors"] += 1
                    yield _emit("lesson_error", title=title, error="lesson not found in DB")
                    continue
                if lrow["type"] != "vocabulary":
                    yield _emit("lesson_error", title=title, error=f"lesson type is '{lrow['type']}', expected vocabulary")
                    continue
                lid = lrow["lesson_id"]

                # fetch DB questions + their single answer + media presence
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT q.question_id, q.sequence_id, q.question_text,
                                  (SELECT a.answer_id FROM Answer a WHERE a.question_id=q.question_id
                                   ORDER BY a.answer_id ASC LIMIT 1) AS answer_id,
                                  (SELECT COUNT(*) FROM Audio au WHERE au.question_id=q.question_id) AS has_aud
                           FROM Question q WHERE q.lesson_id=%s ORDER BY q.question_id ASC""",
                        (lid,),
                    )
                    db_q = cur.fetchall()

                # new questions: extract text + answer text
                new_q = []
                for qa in new_qas:
                    qtext = qa.get("question_text") or qa.get("question", "")
                    ans = ""
                    if isinstance(qa.get("answers"), list) and qa["answers"]:
                        a0 = qa["answers"][0]
                        ans = a0.get("answer_text") or a0.get("answer") or a0.get("text", "")
                    elif isinstance(qa.get("answer"), str):
                        ans = qa["answer"]
                    new_q.append({"question_text": qtext, "answer_text": ans})

                db_texts = [r["question_text"] for r in db_q]
                new_texts = [q["question_text"] for q in new_q]
                yield _emit("lesson_start", title=title, lesson_id=lid,
                            db_questions=len(db_q), new_questions=len(new_q))

                # --- LLM match (always, even dry_run) ---
                try:
                    pairs = _llm_match_questions(client, body.model, db_texts, new_texts)
                except Exception as e:
                    grand["errors"] += 1
                    yield _emit("lesson_error", title=title, error=f"LLM match failed: {e}")
                    continue

                # resolve pairs to indices, enforce one-to-one
                used_d, used_n = set(), set()
                matches = []  # (db_idx, new_idx)
                for p in pairs:
                    d = str(p.get("db", "")); n = str(p.get("new", ""))
                    md = re.match(r"D(\d+)$", d); mn = re.match(r"N(\d+)$", n)
                    if not md or not mn:
                        continue
                    di = int(md.group(1)) - 1; ni = int(mn.group(1)) - 1
                    if not (0 <= di < len(db_q)) or not (0 <= ni < len(new_q)):
                        continue
                    if di in used_d or ni in used_n:
                        grand["duplicates_ignored"] += 1
                        yield _emit("duplicate_ignored", title=title, db=d, new=n)
                        continue
                    used_d.add(di); used_n.add(ni)
                    matches.append((di, ni))

                # classify
                perfect, fuzzy = [], []
                for di, ni in matches:
                    if _norm_text(db_q[di]["question_text"]) == _norm_text(new_q[ni]["question_text"]):
                        perfect.append((di, ni))
                    else:
                        fuzzy.append((di, ni))
                new_only = [i for i in range(len(new_q)) if i not in used_n]
                removed = [i for i in range(len(db_q)) if i not in used_d]

                yield _emit("llm_matched", title=title,
                            perfect=len(perfect), fuzzy=len(fuzzy),
                            new=len(new_only), removed=len(removed))

                lesson_regen_seqs = []

                # --- apply ---
                cur = conn.cursor()

                # next sequence id for appends
                next_seq = (max([r["sequence_id"] for r in db_q], default=0)) + 1

                # FUZZY: update text, delete audio row (keep image)
                for di, ni in fuzzy:
                    dbq = db_q[di]; nq = new_q[ni]
                    yield _emit("fuzzy", title=title, question_id=dbq["question_id"],
                                sequence_id=dbq["sequence_id"],
                                old_text=dbq["question_text"], new_text=nq["question_text"],
                                action="update text, keep image, regenerate audio")
                    if not body.dry_run:
                        cur.execute("UPDATE Question SET question_text=%s WHERE question_id=%s",
                                    (nq["question_text"], dbq["question_id"]))
                        if dbq["answer_id"]:
                            cur.execute("UPDATE Answer SET answer_text=%s WHERE answer_id=%s",
                                        (nq["answer_text"], dbq["answer_id"]))
                        else:
                            cur.execute("INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                        "VALUES (%s,%s,%s,1)", (lid, dbq["question_id"], nq["answer_text"]))
                        # drop audio link so regenerate recreates it
                        cur.execute("DELETE FROM Audio WHERE question_id=%s", (dbq["question_id"],))
                    grand["fuzzy"] += 1
                    lesson_regen_seqs.append(dbq["sequence_id"])

                # NEW: insert question + answer (append)
                for ni in new_only:
                    nq = new_q[ni]
                    seq = next_seq; next_seq += 1
                    yield _emit("new", title=title, sequence_id=seq,
                                question_text=nq["question_text"], action="insert, generate image+audio")
                    if not body.dry_run:
                        cur.execute(
                            "INSERT INTO Question (lesson_id, sequence_id, question_text, type, has_answer, answer_category) "
                            "VALUES (%s,%s,%s,'short_answer',1,'practice')",
                            (lid, seq, nq["question_text"]),
                        )
                        qid = cur.lastrowid
                        cur.execute("INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                    "VALUES (%s,%s,%s,1)", (lid, qid, nq["answer_text"]))
                    grand["new"] += 1
                    lesson_regen_seqs.append(seq)

                # REMOVED: delete db question + media + (optionally) user rows
                for di in removed:
                    dbq = db_q[di]
                    yield _emit("removed", title=title, question_id=dbq["question_id"],
                                sequence_id=dbq["sequence_id"], text=dbq["question_text"],
                                action="delete question + image + audio links")
                    if not body.dry_run:
                        qid = dbq["question_id"]
                        cur.execute("DELETE FROM Image WHERE question_id=%s", (qid,))
                        cur.execute("DELETE FROM Audio WHERE question_id=%s", (qid,))
                        if not body.keep_user_history:
                            cur.execute("DELETE FROM userResponses WHERE question_id=%s", (qid,))
                            cur.execute("DELETE FROM user_attempts WHERE current_question_id=%s", (qid,))
                        cur.execute("DELETE FROM Answer WHERE question_id=%s", (qid,))
                        cur.execute("DELETE FROM Question WHERE question_id=%s", (qid,))
                    grand["removed"] += 1

                grand["perfect"] += len(perfect)
                grand["lessons"] += 1

                if body.dry_run:
                    conn.rollback()
                else:
                    conn.commit()
                cur.close()

                yield _emit("lesson_summary", title=title,
                            perfect=len(perfect), fuzzy=len(fuzzy),
                            new=len(new_only), removed=len(removed),
                            regenerate_sequence_ids=sorted(lesson_regen_seqs))
                if lesson_regen_seqs:
                    regen_plan.append({"title": title,
                                       "question_sequence_ids": sorted(lesson_regen_seqs)})

            yield _emit("summary", dry_run=body.dry_run, totals=grand,
                        regenerate_plan=regen_plan,
                        note="run /regenerate-lesson-media per title with these sequence_ids "
                             "(NEW -> image+audio, FUZZY -> audio only, image kept)")
            logger.info("sync-vocab-lesson done | dry_run=%s totals=%s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("sync-vocab-lesson failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")

# =============================================================================
#  Add to main.py  —  /link-listening-unit-images
#
#  For each unit, stamp that unit's image onto EVERY question in EVERY listening
#  lesson of that unit (question-level images).
#
#  Per unit (unit_number is a VARCHAR in UnitImages, so we match as a string):
#    1. UnitImages   -> image_id for this unit_number
#    2. Image        -> image_url for that image_id   (the source url to copy)
#    3. Lesson       -> type='listening' AND title REGEXP '^unit{N}_'  -> their questions
#    4. Per question:
#         - Image row exists -> UPDATE image_url   (overwrite, per spec)
#         - no Image row      -> INSERT one via _insert_image_row
#
#  Streams NDJSON. dry_run=True (default): reports the plan per unit, writes nothing,
#  rolls back. Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse /
#  _insert_image_row / pymysql / json.
# =============================================================================


class LinkListeningUnitImagesRequest(BaseModel):
    db: DBConfig
    units: list[int]                 # unit numbers; matched as strings against UnitImages.unit_number
    dry_run: bool = True


@app.post("/link-listening-unit-images")
def link_listening_unit_images(body: LinkListeningUnitImagesRequest):

    def stream():
        if not body.units:
            yield _emit("error", message="Provide at least one unit in 'units'")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="link-listening-unit-images",
                    database=body.db.database, units=body.units, dry_run=body.dry_run)

        grand = {"units": 0, "updated": 0, "inserted": 0,
                 "questions": 0, "lessons": 0, "skipped_units": 0}

        try:
            for unit in body.units:
                unit_str = str(unit)

                # 1) UnitImages -> image_id for this unit
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT image_id FROM UnitImages WHERE unit_number = %s "
                        "ORDER BY id DESC LIMIT 1",
                        (unit_str,),
                    )
                    uirow = cur.fetchone()
                if not uirow or not uirow.get("image_id"):
                    grand["skipped_units"] += 1
                    yield _emit("unit_skipped", unit=unit, reason="no row in UnitImages")
                    continue
                source_image_id = uirow["image_id"]

                # 2) Image -> image_url for that image_id
                with conn.cursor() as cur:
                    cur.execute("SELECT image_url FROM Image WHERE image_id = %s", (source_image_id,))
                    imgrow = cur.fetchone()
                source_url = (imgrow or {}).get("image_url")
                if not source_url or not str(source_url).strip():
                    grand["skipped_units"] += 1
                    yield _emit("unit_skipped", unit=unit, image_id=source_image_id,
                                reason="source Image has no usable image_url")
                    continue

                # 3) listening lessons in this unit + their questions
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT lesson_id, title FROM Lesson "
                        "WHERE type = 'listening' AND title REGEXP %s ORDER BY lesson_id",
                        (f"^unit{unit}_",),
                    )
                    lessons = cur.fetchall()

                if not lessons:
                    yield _emit("unit_done", unit=unit, image_id=source_image_id,
                                lessons=0, questions=0, updated=0, inserted=0,
                                note="no listening lessons matched")
                    continue

                yield _emit("unit_start", unit=unit, image_id=source_image_id,
                            image_url=source_url, listening_lessons=len(lessons))

                unit_updated = unit_inserted = unit_questions = 0

                for lrow in lessons:
                    lid = lrow["lesson_id"]
                    with conn.cursor() as cur:
                        cur.execute("SELECT question_id FROM Question WHERE lesson_id = %s ORDER BY sequence_id", (lid,))
                        qids = [r["question_id"] for r in cur.fetchall()]

                    l_upd = l_ins = 0
                    for qid in qids:
                        unit_questions += 1
                        # does this question already have an Image row?
                        with conn.cursor() as cur:
                            cur.execute(
                                "SELECT image_id FROM Image WHERE lesson_id = %s AND question_id = %s "
                                "ORDER BY image_id ASC LIMIT 1",
                                (lid, qid),
                            )
                            existing = cur.fetchone()

                        if not body.dry_run:
                            if existing:
                                with conn.cursor() as cur:
                                    cur.execute(
                                        "UPDATE Image SET image_url = %s WHERE image_id = %s",
                                        (source_url, existing["image_id"]),
                                    )
                            else:
                                meta = json.dumps({
                                    "source": "unit_image",
                                    "unit": unit,
                                    "source_image_id": source_image_id,
                                }, ensure_ascii=False)
                                _insert_image_row(
                                    conn,
                                    lesson_id=lid,
                                    question_id=qid,
                                    image_url=source_url,
                                    image_metadata_json=meta,
                                )

                        if existing:
                            l_upd += 1; unit_updated += 1
                        else:
                            l_ins += 1; unit_inserted += 1

                    yield _emit("lesson_done", unit=unit, lesson_id=lid, title=lrow["title"],
                                questions=len(qids), updated=l_upd, inserted=l_ins)

                # commit per unit (or roll back on dry run)
                if body.dry_run:
                    conn.rollback()
                else:
                    conn.commit()

                grand["units"] += 1
                grand["lessons"] += len(lessons)
                grand["questions"] += unit_questions
                grand["updated"] += unit_updated
                grand["inserted"] += unit_inserted

                yield _emit("unit_done", unit=unit, image_id=source_image_id,
                            lessons=len(lessons), questions=unit_questions,
                            updated=unit_updated, inserted=unit_inserted)

            yield _emit("summary", dry_run=body.dry_run, totals=grand)
            logger.info("link-listening-unit-images done | dry_run=%s totals=%s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("link-listening-unit-images failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")

# =============================================================================
#  /link-reading-lesson-images
#
#  For new READING lessons that exist in Lesson but have no image wired up yet:
#  create one Image row for the lesson and one LessonImages row pointing at it.
#
#  Per lesson (targeted by lesson_id OR title, each with its own image_url):
#    1. resolve lesson; must be type='reading'            -> else warn + skip
#    2. if it already has a LessonImages row              -> warn + skip (safe re-runs)
#    3. find first question (lowest sequence_id)          -> none? warn + skip
#    4. INSERT Image (lesson_id, question_id, sequence_id, image_url, metadata) -> new image_id
#    5. INSERT LessonImages (image_id, lesson_id)
#
#  Streams NDJSON. dry_run=True (default): reports the plan, writes nothing, rolls back.
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse / json.
# =============================================================================


class ReadingImageTarget(BaseModel):
    lesson_id: int | None = None
    title: str | None = None
    image_url: str


class LinkReadingLessonImagesRequest(BaseModel):
    db: DBConfig
    lessons: list[ReadingImageTarget]      # each: {lesson_id | title, image_url}
    dry_run: bool = True


@app.post("/link-reading-lesson-images")
def link_reading_lesson_images(body: LinkReadingLessonImagesRequest):

    def stream():
        if not body.lessons:
            yield _emit("error", message="Provide at least one lesson in 'lessons'")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="link-reading-lesson-images",
                    database=body.db.database, lessons=len(body.lessons), dry_run=body.dry_run)

        grand = {"created": 0, "skipped": 0, "errors": 0}

        try:
            for tgt in body.lessons:
                if not tgt.lesson_id and not tgt.title:
                    grand["errors"] += 1
                    yield _emit("lesson_error", error="target needs lesson_id or title")
                    continue
                if not tgt.image_url or not tgt.image_url.strip():
                    grand["errors"] += 1
                    yield _emit("lesson_error", lesson_id=tgt.lesson_id, title=tgt.title,
                                error="image_url is empty")
                    continue

                # 1) resolve lesson
                with conn.cursor() as cur:
                    if tgt.lesson_id:
                        cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE lesson_id = %s",
                                    (tgt.lesson_id,))
                    else:
                        cur.execute("SELECT lesson_id, title, type FROM Lesson WHERE title = %s",
                                    (tgt.title,))
                    lrow = cur.fetchone()

                if not lrow:
                    grand["skipped"] += 1
                    yield _emit("lesson_skipped", lesson_id=tgt.lesson_id, title=tgt.title,
                                reason="lesson not found")
                    continue
                if lrow["type"] != "reading":
                    grand["skipped"] += 1
                    yield _emit("lesson_skipped", lesson_id=lrow["lesson_id"], title=lrow["title"],
                                reason=f"lesson type is '{lrow['type']}', expected reading")
                    continue

                lid = lrow["lesson_id"]

                # 2) already linked?
                with conn.cursor() as cur:
                    cur.execute("SELECT id FROM LessonImages WHERE lesson_id = %s LIMIT 1", (lid,))
                    if cur.fetchone():
                        grand["skipped"] += 1
                        yield _emit("lesson_skipped", lesson_id=lid, title=lrow["title"],
                                    reason="already has a LessonImages row")
                        continue

                # 3) first question (lowest sequence_id)
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT question_id, sequence_id FROM Question WHERE lesson_id = %s "
                        "ORDER BY sequence_id ASC, question_id ASC LIMIT 1",
                        (lid,),
                    )
                    qrow = cur.fetchone()
                if not qrow:
                    grand["skipped"] += 1
                    yield _emit("lesson_skipped", lesson_id=lid, title=lrow["title"],
                                reason="lesson has no questions")
                    continue

                first_qid = qrow["question_id"]
                first_seq = qrow["sequence_id"]

                yield _emit("lesson_plan", lesson_id=lid, title=lrow["title"],
                            first_question_id=first_qid, sequence_id=first_seq,
                            image_url=tgt.image_url)

                if not body.dry_run:
                    meta = json.dumps({"source": "reading_lesson_image", "lesson_id": lid},
                                      ensure_ascii=False)
                    with conn.cursor() as cur:
                        # 4) Image row for the lesson
                        cur.execute(
                            "INSERT INTO Image (lesson_id, question_id, sequence_id, image_url, image_metadata) "
                            "VALUES (%s, %s, %s, %s, CAST(%s AS JSON))",
                            (lid, first_qid, first_seq, tgt.image_url, meta),
                        )
                        new_image_id = cur.lastrowid

                        # 5) LessonImages row linking image -> lesson
                        cur.execute(
                            "INSERT INTO LessonImages (image_id, lesson_id) VALUES (%s, %s)",
                            (new_image_id, lid),
                        )
                    conn.commit()
                else:
                    new_image_id = None

                grand["created"] += 1
                yield _emit("lesson_created", lesson_id=lid, title=lrow["title"],
                            image_id=new_image_id, question_id=first_qid,
                            sequence_id=first_seq, image_url=tgt.image_url)
                logger.info("link-reading-lesson-images | lesson %d -> image %s",
                            lid, new_image_id)

            if body.dry_run:
                conn.rollback()

            yield _emit("summary", dry_run=body.dry_run, totals=grand)
            logger.info("link-reading-lesson-images done | dry_run=%s totals=%s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("link-reading-lesson-images failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# =============================================================================
#  /swap-vocab-question-answer
#
#  For one or more vocabulary lessons (by title), swap question_text <-> answer_text
#  for every question: old question_text becomes the answer, old answer_text becomes
#  the question.
#
#  Safety:
#    - vocabulary lessons only (skips + reports any non-vocab / missing title)
#    - skips any question with no real (non-blank) answer to swap with (would blank it)
#    - flags questions with >1 answer row (ambiguous) and skips them
#    - dry_run=True (default): streams every old->new swap, writes nothing, rolls back
#    - commits per lesson on a live run
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse.
# =============================================================================


class SwapVocabQARequest(BaseModel):
    db: DBConfig
    titles: list[str]
    dry_run: bool = True


@app.post("/swap-vocab-question-answer")
def swap_vocab_question_answer(body: SwapVocabQARequest):

    def stream():
        if not body.titles:
            yield _emit("error", message="Provide at least one title in 'titles'")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="swap-vocab-question-answer",
                    titles=body.titles, dry_run=body.dry_run)

        grand = {"lessons": 0, "swapped": 0,
                 "skipped_no_answer": 0, "skipped_multi_answer": 0, "skipped_lessons": 0}

        try:
            for title in body.titles:
                # resolve lesson, must be vocabulary
                with conn.cursor() as cur:
                    cur.execute("SELECT lesson_id, type FROM Lesson WHERE title = %s", (title,))
                    lrow = cur.fetchone()
                if not lrow:
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title, reason="lesson not found")
                    continue
                if lrow["type"] != "vocabulary":
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title,
                                reason=f"lesson type is '{lrow['type']}', expected vocabulary")
                    continue
                lid = lrow["lesson_id"]

                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT question_id, sequence_id, question_text FROM Question "
                        "WHERE lesson_id = %s ORDER BY sequence_id, question_id",
                        (lid,),
                    )
                    questions = cur.fetchall()

                yield _emit("lesson_start", title=title, lesson_id=lid, questions=len(questions))

                l_tot = {"swapped": 0, "skipped_no_answer": 0, "skipped_multi_answer": 0}

                for q in questions:
                    qid = q["question_id"]
                    with conn.cursor() as cur:
                        cur.execute(
                            "SELECT answer_id, answer_text FROM Answer WHERE question_id = %s "
                            "ORDER BY answer_id ASC",
                            (qid,),
                        )
                        answers = cur.fetchall()

                    if len(answers) > 1:
                        l_tot["skipped_multi_answer"] += 1
                        yield _emit("skipped", title=title, question_id=qid,
                                    sequence_id=q["sequence_id"],
                                    reason="multiple answer rows (ambiguous)", answer_rows=len(answers))
                        continue

                    ans = answers[0] if answers else None
                    ans_text = (ans["answer_text"] if ans else "") or ""
                    q_text = q["question_text"] or ""

                    if ans is None or ans_text.strip() == "":
                        l_tot["skipped_no_answer"] += 1
                        yield _emit("skipped", title=title, question_id=qid,
                                    sequence_id=q["sequence_id"],
                                    reason="no non-blank answer to swap with")
                        continue

                    new_question_text = ans_text
                    new_answer_text = q_text

                    yield _emit("swap", title=title, question_id=qid, sequence_id=q["sequence_id"],
                                old_question=q_text, old_answer=ans_text,
                                new_question=new_question_text, new_answer=new_answer_text)

                    if not body.dry_run:
                        with conn.cursor() as cur:
                            cur.execute("UPDATE Question SET question_text = %s WHERE question_id = %s",
                                        (new_question_text, qid))
                            cur.execute("UPDATE Answer SET answer_text = %s WHERE answer_id = %s",
                                        (new_answer_text, ans["answer_id"]))
                    l_tot["swapped"] += 1

                if body.dry_run:
                    conn.rollback()
                else:
                    conn.commit()

                grand["lessons"] += 1
                grand["swapped"] += l_tot["swapped"]
                grand["skipped_no_answer"] += l_tot["skipped_no_answer"]
                grand["skipped_multi_answer"] += l_tot["skipped_multi_answer"]

                yield _emit("lesson_done", title=title, lesson_id=lid, totals=l_tot)

            yield _emit("summary", dry_run=body.dry_run, totals=grand)
            logger.info("swap-vocab-qa | dry_run=%s %s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("swap-vocab-qa failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# =============================================================================
#  /renumber-vocab-sequence
#
#  For one or more VOCABULARY lessons (by title), renumber every question's
#  sequence_id to a clean 1..n — closing gaps and resolving duplicate sequence_ids
#  left behind by QA edits/deletes.
#
#  Order: sequence_id ASC, then question_id ASC (stable tie-break for duplicates).
#  Safe two-pass update (offset to a temporary high range, then set final 1..n) so
#  it works even with duplicate sequence_ids or a unique index on (lesson_id, sequence_id).
#
#  Safety:
#    - vocabulary lessons only (skips + reports non-vocab / missing titles)
#    - dry_run=True (default): streams every old->new mapping, writes nothing, rolls back
#    - commits per lesson on a live run
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse.
# =============================================================================


class RenumberVocabSequenceRequest(BaseModel):
    db: DBConfig
    titles: list[str]
    dry_run: bool = True


# offset large enough to never collide with real 1..n values during pass 1
_SEQ_TEMP_OFFSET = 1000000


# =============================================================================
#  /renumber-vocab-sequence
#
#  For one or more VOCABULARY lessons (by title), renumber every question's
#  sequence_id to a clean 1..n — closing gaps and resolving duplicate sequence_ids
#  left behind by QA edits/deletes.
#
#  Order: question_id ASC (creation order) -> first-created question becomes 1, etc.
#  Safe two-pass update (offset to a temporary high range, then set final 1..n) so
#  it works even with duplicate sequence_ids or a unique index on (lesson_id, sequence_id).
#
#  Safety:
#    - vocabulary lessons only (skips + reports non-vocab / missing titles)
#    - dry_run=True (default): streams every old->new mapping, writes nothing, rolls back
#    - commits per lesson on a live run
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse.
# =============================================================================


class RenumberVocabSequenceRequest(BaseModel):
    db: DBConfig
    titles: list[str]
    dry_run: bool = True


# offset large enough to never collide with real 1..n values during pass 1
_SEQ_TEMP_OFFSET = 1000000


@app.post("/renumber-vocab-sequence")
def renumber_vocab_sequence(body: RenumberVocabSequenceRequest):

    def stream():
        if not body.titles:
            yield _emit("error", message="Provide at least one title in 'titles'")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        yield _emit("start", action="renumber-vocab-sequence",
                    titles=body.titles, dry_run=body.dry_run)

        grand = {"lessons": 0, "renumbered_questions": 0,
                 "unchanged_questions": 0, "skipped_lessons": 0}

        try:
            for title in body.titles:
                with conn.cursor() as cur:
                    cur.execute("SELECT lesson_id, type FROM Lesson WHERE title = %s", (title,))
                    lrow = cur.fetchone()
                if not lrow:
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title, reason="lesson not found")
                    continue
                if lrow["type"] != "vocabulary":
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title,
                                reason=f"lesson type is '{lrow['type']}', expected vocabulary")
                    continue
                lid = lrow["lesson_id"]

                # read questions in the target order
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT question_id, sequence_id FROM Question "
                        "WHERE lesson_id = %s ORDER BY question_id ASC",
                        (lid,),
                    )
                    questions = cur.fetchall()

                yield _emit("lesson_start", title=title, lesson_id=lid, questions=len(questions))

                # compute the new mapping (1..n)
                mapping = []   # (question_id, old_seq, new_seq)
                changed = 0
                for new_seq, q in enumerate(questions, start=1):
                    old_seq = q["sequence_id"]
                    mapping.append((q["question_id"], old_seq, new_seq))
                    if old_seq != new_seq:
                        changed += 1
                        yield _emit("renumber", title=title, question_id=q["question_id"],
                                    old=old_seq, new=new_seq)

                # apply with a safe two-pass update (only if there is something to do)
                if not body.dry_run and changed > 0:
                    with conn.cursor() as cur:
                        # pass 1: move everyone to a temporary, collision-free range
                        for qid, _old, new_seq in mapping:
                            cur.execute(
                                "UPDATE Question SET sequence_id = %s WHERE question_id = %s",
                                (new_seq + _SEQ_TEMP_OFFSET, qid),
                            )
                        # pass 2: bring them down to the final 1..n
                        for qid, _old, new_seq in mapping:
                            cur.execute(
                                "UPDATE Question SET sequence_id = %s WHERE question_id = %s",
                                (new_seq, qid),
                            )
                    conn.commit()
                elif body.dry_run:
                    conn.rollback()
                # changed == 0 -> nothing to write

                grand["lessons"] += 1
                grand["renumbered_questions"] += changed
                grand["unchanged_questions"] += (len(questions) - changed)

                yield _emit("lesson_done", title=title, lesson_id=lid,
                            total=len(questions), renumbered=changed,
                            unchanged=len(questions) - changed)

            yield _emit("summary", dry_run=body.dry_run, totals=grand)
            logger.info("renumber-vocab-sequence | dry_run=%s %s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("renumber-vocab-sequence failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")


# =============================================================================
#  /fill-vocab-answers
#
#  For one or more VOCABULARY lessons (by title), fill in MISSING answers by
#  translating question_text -> English via OpenAI.
#
#  Per question:
#    - real answer already present  -> left untouched
#    - Answer row exists but blank   -> UPDATE its answer_text with the translation
#    - no Answer row                 -> INSERT one (is_correct=1)
#  Then: Question.has_answer = 1 and answer_category = 'practice'.
#
#  Safety:
#    - vocabulary lessons only (skips + reports otherwise)
#    - dry_run=True (default): makes the (cheap) OpenAI calls so you can review every
#      proposed translation, but writes nothing and rolls back
#    - commits per lesson on a live run
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse / OpenAI / os.
# =============================================================================


class FillVocabAnswersRequest(BaseModel):
    db: DBConfig
    titles: list[str]
    source_language: str = "German"        # human-readable; helps the model translate correctly
    model: str = "gpt-4o-mini"
    openai_api_key: str | None = None
    dry_run: bool = True


def _translate_to_english(client, model: str, text: str, source_language: str) -> str:
    """Translate a single vocab word/phrase from source_language to English. Returns plain text."""
    resp = client.chat.completions.create(
        model=model,
        temperature=0,
        messages=[
            {"role": "system", "content": (
                "You translate single vocabulary words or short phrases from a language-learning "
                f"textbook. Translate the given {source_language} term into English. Reply with ONLY "
                "the English translation — no quotes, no explanation, no trailing punctuation unless "
                "it is part of the term."
            )},
            {"role": "user", "content": text},
        ],
    )
    return (resp.choices[0].message.content or "").strip()


@app.post("/fill-vocab-answers")
def fill_vocab_answers(body: FillVocabAnswersRequest):

    def stream():
        if not body.titles:
            yield _emit("error", message="Provide at least one title in 'titles'")
            return

        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        client = OpenAI(api_key=body.openai_api_key or os.getenv("OPENAI_API_KEY"))

        yield _emit("start", action="fill-vocab-answers", titles=body.titles,
                    source_language=body.source_language, model=body.model, dry_run=body.dry_run)

        grand = {"lessons": 0, "filled": 0, "inserted": 0, "updated": 0,
                 "already_had_answer": 0, "skipped_lessons": 0, "errors": 0}

        try:
            for title in body.titles:
                with conn.cursor() as cur:
                    cur.execute("SELECT lesson_id, type FROM Lesson WHERE title = %s", (title,))
                    lrow = cur.fetchone()
                if not lrow:
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title, reason="lesson not found")
                    continue
                if lrow["type"] != "vocabulary":
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title,
                                reason=f"lesson type is '{lrow['type']}', expected vocabulary")
                    continue
                lid = lrow["lesson_id"]

                # questions + their current (single) answer row, in order
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT q.question_id, q.sequence_id, q.question_text,
                                  (SELECT a.answer_id   FROM Answer a WHERE a.question_id = q.question_id
                                   ORDER BY a.answer_id ASC LIMIT 1) AS answer_id,
                                  (SELECT a.answer_text FROM Answer a WHERE a.question_id = q.question_id
                                   ORDER BY a.answer_id ASC LIMIT 1) AS answer_text
                           FROM Question q WHERE q.lesson_id = %s
                           ORDER BY q.sequence_id, q.question_id""",
                        (lid,),
                    )
                    questions = cur.fetchall()

                yield _emit("lesson_start", title=title, lesson_id=lid, questions=len(questions))

                l_filled = l_ins = l_upd = l_had = 0

                for q in questions:
                    qid = q["question_id"]
                    has_real_answer = (q["answer_text"] or "").strip() != ""
                    if has_real_answer:
                        l_had += 1
                        continue  # leave existing answers alone

                    qtext = (q["question_text"] or "").strip()
                    if not qtext:
                        grand["errors"] += 1
                        yield _emit("skipped", title=title, question_id=qid,
                                    sequence_id=q["sequence_id"], reason="question_text is blank")
                        continue

                    # translate (cheap; we do this even on dry_run so you can review)
                    try:
                        english = _translate_to_english(client, body.model, qtext, body.source_language)
                        if not english:
                            raise RuntimeError("empty translation")
                    except Exception as e:
                        grand["errors"] += 1
                        yield _emit("translate_error", title=title, question_id=qid,
                                    sequence_id=q["sequence_id"], question_text=qtext, error=str(e))
                        continue

                    has_blank_row = q["answer_id"] is not None
                    action = "update_blank" if has_blank_row else "insert"
                    yield _emit("fill", title=title, question_id=qid, sequence_id=q["sequence_id"],
                                question_text=qtext, translation=english, action=action)

                    if not body.dry_run:
                        with conn.cursor() as cur:
                            if has_blank_row:
                                cur.execute("UPDATE Answer SET answer_text = %s, is_correct = 1 "
                                            "WHERE answer_id = %s", (english, q["answer_id"]))
                                l_upd += 1
                            else:
                                cur.execute("INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                            "VALUES (%s, %s, %s, 1)", (lid, qid, english))
                                l_ins += 1
                            cur.execute("UPDATE Question SET has_answer = 1, answer_category = 'practice' "
                                        "WHERE question_id = %s", (qid,))
                    else:
                        if has_blank_row: l_upd += 1
                        else: l_ins += 1

                    l_filled += 1

                if body.dry_run:
                    conn.rollback()
                else:
                    conn.commit()

                grand["lessons"] += 1
                grand["filled"] += l_filled
                grand["inserted"] += l_ins
                grand["updated"] += l_upd
                grand["already_had_answer"] += l_had

                yield _emit("lesson_done", title=title, lesson_id=lid,
                            filled=l_filled, inserted=l_ins, updated=l_upd,
                            already_had_answer=l_had, total=len(questions))

            yield _emit("summary", dry_run=body.dry_run, totals=grand,
                        note="filled answers are machine translations — review before relying on them")
            logger.info("fill-vocab-answers | dry_run=%s %s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("fill-vocab-answers failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")

# =============================================================================
#  /regenerate-grammar-questions
#
#  Grammar MCQs were rebalanced into single-answer learning/practice questions, but
#  their TEXT still reads as multiple-choice nonsense. This route rewrites the text
#  IN PLACE from the lesson's article, keeping the same learning/practice split and
#  the same question rows (so images/audio/user history stay attached).
#
#  Per grammar lesson (by title):
#    1. read all Article text + classify questions: learning (no real answer) / practice (has answer)
#    2. ONE OpenAI call: from the article, generate L learning questions + P practice Q&A pairs
#    3. rewrite each learning row's text (stays answerless); rewrite each practice row's text
#       and write the generated answer (is_correct=1, has_answer=1)
#    4. re-run the grammar backfill so answer_category stays consistent
#
#  Streams NDJSON. dry_run=True (default): makes the LLM call so you can review the
#  rewrites, writes nothing, rolls back. Commits per lesson on a live run.
#
#  Reuses connect_to_db / DBConfig / logger / _emit / StreamingResponse / OpenAI / os /
#         _GRAMMAR_BACKFILL_SQL.
# =============================================================================


class RegenerateGrammarQuestionsRequest(BaseModel):
    db: DBConfig
    titles: list[str]
    model: str = "gpt-4o-mini"
    openai_api_key: str | None = None
    dry_run: bool = True


def _generate_grammar_question_set(client, model: str, article_text: str,
                                   n_learning: int, n_practice: int) -> dict:
    """
    One call: from the article, produce n_learning learning questions (no answer needed)
    and n_practice practice questions each with a correct answer. Returns:
        {"learning": ["q", ...], "practice": [{"question": "...", "answer": "..."}, ...]}
    """
    system = (
        "You write grammar exercises for a language-learning app, based ONLY on the provided "
        "grammar lesson article. Produce two kinds of questions:\n"
        "- LEARNING questions: standalone prompts that help a learner think about / recall the "
        "grammar concept. They do NOT need a checkable answer.\n"
        "- PRACTICE questions: each has ONE clear correct answer derivable from the article.\n"
        "Every question must make sense on its own (no 'which of the following', no multiple-choice "
        "phrasing, no references to options). Return ONLY JSON, no prose, no code fences."
    )
    user = (
        f"ARTICLE:\n{article_text}\n\n"
        f"Generate exactly {n_learning} LEARNING questions and {n_practice} PRACTICE questions.\n"
        "Return JSON of this exact shape:\n"
        '{"learning": ["question", ...], '
        '"practice": [{"question": "question", "answer": "the correct answer"}, ...]}'
    )
    resp = client.chat.completions.create(
        model=model, temperature=0.2,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user}],
    )
    content = (resp.choices[0].message.content or "").strip()
    if content.startswith("```"):
        content = re.sub(r"^```(?:json)?\s*", "", content)
        content = re.sub(r"\s*```$", "", content)
    data = json.loads(content)
    learning = [str(q).strip() for q in data.get("learning", []) if str(q).strip()]
    practice = []
    for item in data.get("practice", []):
        q = str(item.get("question", "")).strip()
        a = str(item.get("answer", "")).strip()
        if q and a:
            practice.append({"question": q, "answer": a})
    return {"learning": learning, "practice": practice}


@app.post("/regenerate-grammar-questions")
def regenerate_grammar_questions(body: RegenerateGrammarQuestionsRequest):

    def stream():
        if not body.titles:
            yield _emit("error", message="Provide at least one title in 'titles'")
            return
        try:
            conn = connect_to_db(body.db)
        except Exception as e:
            yield _emit("error", message=f"DB connection failed: {e}")
            return

        effective_api_key = body.openai_api_key or os.getenv("OPENAI_API_KEY")
        if not effective_api_key:
            conn.close()
            yield _emit("error", message="OpenAI API key is required (send openai_api_key in the request).")
            return
        client = OpenAI(api_key=effective_api_key)

        yield _emit("start", action="regenerate-grammar-questions",
                    titles=body.titles, model=body.model, dry_run=body.dry_run)

        grand = {"lessons": 0, "learning_rewritten": 0, "practice_rewritten": 0,
                 "shortfalls": 0, "skipped_lessons": 0, "errors": 0}
        touched_lesson_ids: list = []

        try:
            for title in body.titles:
                with conn.cursor() as cur:
                    cur.execute("SELECT lesson_id, type FROM Lesson WHERE title = %s", (title,))
                    lrow = cur.fetchone()
                if not lrow:
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title, reason="lesson not found"); continue
                if lrow["type"] != "grammar":
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title,
                                reason=f"lesson type is '{lrow['type']}', expected grammar"); continue
                lid = lrow["lesson_id"]

                # article text (all articles for the lesson, in order)
                with conn.cursor() as cur:
                    cur.execute(
                        "SELECT content FROM Article WHERE lesson_id = %s ORDER BY sequence_id, article_id",
                        (lid,),
                    )
                    article_text = "\n\n".join((r["content"] or "").strip() for r in cur.fetchall() if r["content"])

                if not article_text.strip():
                    grand["skipped_lessons"] += 1
                    yield _emit("lesson_skipped", title=title, reason="lesson has no article text"); continue

                # classify questions by structure: learning (no real answer) / practice (has answer)
                with conn.cursor() as cur:
                    cur.execute(
                        """SELECT q.question_id, q.sequence_id,
                                  (SELECT a.answer_id FROM Answer a WHERE a.question_id=q.question_id
                                   ORDER BY a.answer_id ASC LIMIT 1) AS answer_id,
                                  (SELECT COUNT(*) FROM Answer a
                                   WHERE a.question_id=q.question_id AND TRIM(a.answer_text) <> '') AS real_answers
                           FROM Question q WHERE q.lesson_id = %s
                           ORDER BY q.sequence_id, q.question_id""",
                        (lid,),
                    )
                    qs = cur.fetchall()

                learning_rows = [q for q in qs if q["real_answers"] == 0]
                practice_rows = [q for q in qs if q["real_answers"] >= 1]
                n_learning, n_practice = len(learning_rows), len(practice_rows)

                yield _emit("lesson_start", title=title, lesson_id=lid,
                            learning=n_learning, practice=n_practice)

                if n_learning == 0 and n_practice == 0:
                    yield _emit("lesson_done", title=title, lesson_id=lid,
                                learning_rewritten=0, practice_rewritten=0, note="no questions")
                    continue

                # one LLM call for the whole lesson
                try:
                    gen = _generate_grammar_question_set(client, body.model, article_text,
                                                         n_learning, n_practice)
                except Exception as e:
                    grand["errors"] += 1
                    yield _emit("lesson_error", title=title, error=f"LLM generation failed: {e}")
                    continue

                gen_learning = gen["learning"]
                gen_practice = gen["practice"]

                # map by count (report any shortfall rather than mismatching)
                l_pairs = list(zip(learning_rows, gen_learning))
                p_pairs = list(zip(practice_rows, gen_practice))
                if len(gen_learning) < n_learning or len(gen_practice) < n_practice:
                    grand["shortfalls"] += 1
                    yield _emit("shortfall", title=title,
                                wanted_learning=n_learning, got_learning=len(gen_learning),
                                wanted_practice=n_practice, got_practice=len(gen_practice))

                l_done = p_done = 0
                cur = conn.cursor()

                # learning: rewrite text only, stays answerless
                for row, new_q in l_pairs:
                    yield _emit("learning_rewrite", title=title, question_id=row["question_id"],
                                new_question=new_q)
                    if not body.dry_run:
                        cur.execute("UPDATE Question SET question_text=%s, has_answer=0 WHERE question_id=%s",
                                    (new_q, row["question_id"]))
                    l_done += 1

                # practice: rewrite text + write the answer
                for row, item in p_pairs:
                    yield _emit("practice_rewrite", title=title, question_id=row["question_id"],
                                new_question=item["question"], new_answer=item["answer"])
                    if not body.dry_run:
                        cur.execute("UPDATE Question SET question_text=%s, has_answer=1 WHERE question_id=%s",
                                    (item["question"], row["question_id"]))
                        if row["answer_id"]:
                            cur.execute("UPDATE Answer SET answer_text=%s, is_correct=1 WHERE answer_id=%s",
                                        (item["answer"], row["answer_id"]))
                        else:
                            cur.execute("INSERT INTO Answer (lesson_id, question_id, answer_text, is_correct) "
                                        "VALUES (%s,%s,%s,1)", (lid, row["question_id"], item["answer"]))
                    p_done += 1

                cur.close()

                if body.dry_run:
                    conn.rollback()
                else:
                    conn.commit()
                    touched_lesson_ids.append(lid)

                grand["lessons"] += 1
                grand["learning_rewritten"] += l_done
                grand["practice_rewritten"] += p_done
                yield _emit("lesson_done", title=title, lesson_id=lid,
                            learning_rewritten=l_done, practice_rewritten=p_done)

            # re-run grammar backfill on the lessons we changed
            if not body.dry_run and touched_lesson_ids:
                fmt = ",".join(["%s"] * len(touched_lesson_ids))
                with conn.cursor() as cur:
                    cur.execute(_GRAMMAR_BACKFILL_SQL.format(ids=fmt),
                                tuple(touched_lesson_ids) + tuple(touched_lesson_ids))
                conn.commit()
                yield _emit("backfill", lessons=len(touched_lesson_ids))

            yield _emit("summary", dry_run=body.dry_run, totals=grand)
            logger.info("regenerate-grammar-questions | dry_run=%s %s", body.dry_run, grand)

        except Exception as e:
            conn.rollback()
            logger.error("regenerate-grammar-questions failed: %s\n%s", e, traceback.format_exc())
            yield _emit("error", message=str(e), rolled_back=True)
        finally:
            conn.close()

    return StreamingResponse(stream(), media_type="application/x-ndjson")

lambda_handler = Mangum(app)
